"""Torch profiler trace analyzer for kernel execution analysis.

Analyzes torch profiler trace files (.trace.json.gz) and provides breakdowns by:
- Stage: prefill vs decode
- Kernel type: attention, MoE, quantization, communication, linear, memory, other
- Layer-level analysis with per-layer kernel sequences
"""

import argparse
import csv
import gzip
import json
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


class Stage(Enum):
    """Execution stage for a kernel."""

    PREFILL = "prefill"
    DECODE = "decode"
    UNKNOWN = "unknown"


class KernelType(Enum):
    """Classification of kernel by operation type."""

    ATTENTION = "attention"
    MOE = "moe"
    QUANTIZATION = "quantization"
    COMMUNICATION = "communication"
    LINEAR = "linear"
    MEMORY = "memory"
    OTHER = "other"


class LayerType(Enum):
    """Type of transformer layer."""

    MLA_FC = "MLA+FC"  # MLA attention + fully connected
    MLA_MOE = "MLA+MoE"  # MLA attention + MoE
    MHA_FC = "MHA+FC"  # MHA attention + fully connected
    MHA_MOE = "MHA+MoE"  # MHA attention + MoE
    GDN_FC = "GDN+FC"  # GDN (gated delta network) attention + fully connected
    GDN_MOE = "GDN+MoE"  # GDN attention + MoE
    ATTN = "Attn"  # Attention-only half-layer (no MoE/FC)
    MOE = "MoE"  # MoE-only half-layer (no attention)
    FC = "FC"  # FC-only half-layer (no attention)
    UNKNOWN = "unknown"


@dataclass
class KernelEvent:
    """Single kernel execution event from the trace."""

    name: str
    timestamp_us: float  # microseconds
    duration_us: float  # microseconds
    stage: Stage
    kernel_type: KernelType
    initial_stage: Stage
    grid: Optional[Tuple[int, int, int]] = None
    block: Optional[Tuple[int, int, int]] = None
    simplified_name: str = ""  # Short name for display

    @property
    def duration_ms(self) -> float:
        return self.duration_us / 1000.0


@dataclass
class KernelStats:
    """Aggregated statistics for a group of kernels."""

    count: int = 0
    total_time_us: float = 0.0
    min_time_us: float = float("inf")
    max_time_us: float = 0.0

    def add(self, duration_us: float) -> None:
        self.count += 1
        self.total_time_us += duration_us
        self.min_time_us = min(self.min_time_us, duration_us)
        self.max_time_us = max(self.max_time_us, duration_us)

    @property
    def avg_time_us(self) -> float:
        return self.total_time_us / self.count if self.count > 0 else 0.0

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def avg_time_ms(self) -> float:
        return self.avg_time_us / 1000.0

    @property
    def min_time_ms(self) -> float:
        return self.min_time_us / 1000.0 if self.min_time_us != float("inf") else 0.0

    @property
    def max_time_ms(self) -> float:
        return self.max_time_us / 1000.0


@dataclass
class LayerEvent:
    """Single layer's execution with all its kernels."""

    layer_idx: int
    layer_type: LayerType
    stage: Stage
    kernels: List[KernelEvent] = field(default_factory=list)

    @property
    def total_time_us(self) -> float:
        return sum(k.duration_us for k in self.kernels)

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def attention_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.ATTENTION
        )

    @property
    def moe_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.MOE
        )

    @property
    def linear_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.LINEAR
        )

    @property
    def communication_time_us(self) -> float:
        return sum(
            k.duration_us
            for k in self.kernels
            if k.kernel_type == KernelType.COMMUNICATION
        )

    @property
    def quantization_time_us(self) -> float:
        return sum(
            k.duration_us
            for k in self.kernels
            if k.kernel_type == KernelType.QUANTIZATION
        )

    def get_breakdown(self) -> Dict[KernelType, float]:
        """Get time breakdown by kernel type."""
        breakdown = {}
        for k in self.kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us
        return breakdown


@dataclass
class AnalysisResult:
    """Complete analysis result with all breakdowns."""

    # All kernel events (time-ordered)
    events: List[KernelEvent] = field(default_factory=list)

    # Per-kernel stats (keyed by kernel name)
    per_kernel_stats: Dict[str, KernelStats] = field(default_factory=dict)

    # Per-stage stats
    per_stage_stats: Dict[Stage, KernelStats] = field(default_factory=dict)

    # Per-type stats
    per_type_stats: Dict[KernelType, KernelStats] = field(default_factory=dict)

    # Per-stage per-type stats
    per_stage_type_stats: Dict[Tuple[Stage, KernelType], KernelStats] = field(
        default_factory=dict
    )

    # Per-stage per-kernel stats
    per_stage_kernel_stats: Dict[Tuple[Stage, str], KernelStats] = field(
        default_factory=dict
    )

    # Layer-level analysis
    layers: List[LayerEvent] = field(default_factory=list)

    @property
    def total_time_us(self) -> float:
        return sum(e.duration_us for e in self.events)

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def total_time_s(self) -> float:
        return self.total_time_us / 1_000_000.0


class KernelClassifier:
    """Classifies kernels by stage, type, and provides simplified names."""

    def __init__(self, grid_threshold: int = 10000):
        self.grid_threshold = grid_threshold

        # Stage detection patterns (priority order)
        self._stage_force_unknown_patterns = [
            re.compile(
                r"_fused_rms_fp8|_fused_rms_mxfp4|_gemm_a8w8|_batched_gemm_a8w8|"
                r"_gemm_afp4wfp4|_batched_gemm_a16wfp4",
                re.IGNORECASE,
            ),
        ]
        self._stage_patterns = [
            (re.compile(r"set_mla_kv_buffer_kernel|concat_and_cast_mha_k_kernel", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"qseqlen[1-4][^0-9]|qseqlen[1-4]$", re.IGNORECASE), Stage.DECODE),
            (re.compile(r"qseqlen[5-9]|qseqlen\d{2,}", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"_decode_|decode_attention|paged_attention_ll4mi", re.IGNORECASE), Stage.DECODE),
            (re.compile(r"_prefill_|flash_attn|fmha_fwd|FmhaBatchPrefill", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"generate_draft_decode|create_extend_after_decode"), Stage.DECODE),
        ]

        # Kernel type patterns (order matters — first match wins)
        self._type_patterns = [
            # MoE patterns (before attention to avoid moeSoftmax → attention)
            (
                re.compile(
                    r"fused_moe|moe_align|topk|expert|MoeFlatmm|MoeSorting|"
                    r"kernel_moe_gemm|kernel_moe_mxgemm|shared_experts|grouped_topk|"
                    r"moe_fused_gate|_moe_mxfp4_sort|moeSoftmax",
                    re.IGNORECASE,
                ),
                KernelType.MOE,
            ),
            # Attention patterns (including GDN / gated delta network)
            (
                re.compile(
                    r"aiter::mla_|mla_a8w8|decode_attention|flash_attn|attention|softmax|"
                    r"fmha_|FmhaBatchPrefill|mla_reduce|kv_cache|flashinfer|set_mla_kv|"
                    r"kn_entry_2c_sbhd|kn_get_mla_metadata|qk_rope|"
                    r"concat_and_cast_mha|kn_mla_reduce|paged_attention|"
                    r"chunk_gated_delta_rule|chunk_fwd_kernel_o|chunk_local_cumsum|"
                    r"chunk_scaled_dot_kkt|fused_gdn_gating|_causal_conv1d_|"
                    r"fused_sigmoid_gating_delta_rule|fused_qkvzba_split|"
                    r"l2norm_fwd|_layer_norm_fwd_1pass|solve_tril|"
                    r"merge_16x16_to_64x64|recompute_w_u_fwd",
                    re.IGNORECASE,
                ),
                KernelType.ATTENTION,
            ),
            # Quantization patterns
            (
                re.compile(
                    r"mxfp4|fp8|quant|_gemm_afp4wfp4|_fused_rms_mxfp4|"
                    r"_batched_gemm_a16wfp4|dynamic_per_group_scaled_quant|"
                    r"_dynamic_mxfp4|_fused_flatten|fp4x2|_gemm_a8w8|"
                    r"_batched_gemm_a8w8|_fused_rms_fp8",
                    re.IGNORECASE,
                ),
                KernelType.QUANTIZATION,
            ),
            # Communication patterns
            (
                re.compile(
                    r"all_reduce|cross_device_reduce|nccl|rccl|broadcast|"
                    r"allgather|reduce_scatter|rcclGenericKernel",
                    re.IGNORECASE,
                ),
                KernelType.COMMUNICATION,
            ),
            # Linear/GEMM patterns (after quant to avoid overlap)
            (
                re.compile(
                    r"Cijk_Alik_Bljk|Cijk_Ailk_Bljk|Cijk_SB_|"
                    r"_gemm_a16_w16|Custom_Cijk",
                    re.IGNORECASE,
                ),
                KernelType.LINEAR,
            ),
            # Memory patterns
            (
                re.compile(
                    r"memcpy|memset|__amd_rocclr_copyBuffer|"
                    r"bfloat16_copy|float8_copy",
                    re.IGNORECASE,
                ),
                KernelType.MEMORY,
            ),
        ]

        # Simplified name patterns for display
        self._simplify_patterns = [
            (re.compile(r"ncclDevKernel|rcclGenericKernel|cross_device_reduce"), "ALLREDUCE"), 
            (re.compile(r"fmha_fwd"), "FMHA"),
            (re.compile(r"mla_a8w8.*qseqlen1"), "MLA_DECODE"),
            (
                re.compile(r"mla_a8w8.*qseqlen[2-9]|mla_a8w8.*qseqlen\d{2,}"),
                "MLA_PREFILL",
            ),
            (re.compile(r"kn_mla_reduce"), "MLA_REDUCE"),
            (re.compile(r"_fused_rms_mxfp4_quant"), "RMS_MXFP4"),
            (re.compile(r"_fused_rms_fp8"), "RMS_FP8"),
            (re.compile(r"_gemm_afp4wfp4.*reduce"), "GEMM_FP4_RED"),
            (re.compile(r"_gemm_afp4wfp4"), "GEMM_FP4"),
            (re.compile(r"_gemm_a8w8_blockscale.*reduce"), "GEMM_FP8_RED"),
            (re.compile(r"_gemm_a8w8_blockscale"), "GEMM_FP8"),
            (re.compile(r"Rmsnorm2dFwd|fused_dual_residual_rmsnorm_kernel|fused_rmsnorm_kernel"), "RMSNORM"),
            (re.compile(r"kn_entry_2c_sbhd"), "KV_CACHE"),
            (re.compile(r"set_mla_kv_buffer"), "MLA_KV_SET"),
            (re.compile(r"_dynamic_mxfp4_quant"), "DYN_MXFP4"),
            (re.compile(r"concat_and_cast_mha"), "MHA_CONCAT"),
            (re.compile(r"kernel_moe_mxgemm|kernel_moe_gemm"), "MOE_GEMM"),
            (re.compile(r"MoeSorting|moe_fused_gate|_moe_mxfp4_sort"), "MOE_SORT"),
            (re.compile(r"dynamic_per_group_scaled_quant"), "DYN_QUANT"),
            (re.compile(r"fused_append_shared_experts"), "SHARED_EXP"),
            (re.compile(r"act_and_mul_kernel"), "ACT_MUL"),
            (re.compile(r"Cijk_Alik|Cijk_Ailk|Cijk_SB_|Custom_Cijk"), "HIPBLAS_GEMM"),
            (re.compile(r"paged_attention_ll4mi_QKV"), "PA_DECODE"),
            (re.compile(r"paged_attention_ll4mi_reduce"), "PA_REDUCE"),
            (re.compile(r"FmhaBatchPrefill"), "FMHA_PREFILL"),
            (re.compile(r"fused_dual_residual_rmsnorm"), "DUAL_RMSNORM"),
            (re.compile(r"gelu_and_mul_kernel"), "GELU_MUL"),
            (re.compile(r"rotary_embedding_kernel"), "ROPE"),
            (re.compile(r"fused_softcap_kernel"), "SOFTCAP"),
            (re.compile(r"grouped_topk"), "TOPK"),
            (re.compile(r"bfloat16_copy"), "BF16_COPY"),
            (re.compile(r"float8_copy"), "FP8_COPY"),
            (re.compile(r"create_flashinfer|generate_draft"), "SPEC_INDEX"),
            (re.compile(r"qk_rope_cat_and_cache"), "ROPE_CACHE"),
            (re.compile(r"_batched_gemm_a16wfp4"), "BATCH_GEMM_FP4"),
            (re.compile(r"_batched_gemm_a8w8"), "BATCH_GEMM_FP8"),
            (re.compile(r"_gemm_a16_w16"), "GEMM_A16W16"),
            (re.compile(r"_fused_flatten"), "FLATTEN_QUANT"),
            (re.compile(r"MoeFlatmm"), "MOE_FLATMM"),
            (re.compile(r"moeSoftmax"), "MOE_SOFTMAX"),
            # GDN (gated delta network) kernels — Qwen3-Coder-Next
            (re.compile(r"chunk_gated_delta_rule"), "GDN_DELTA"),
            (re.compile(r"chunk_fwd_kernel_o"), "GDN_OUTPUT"),
            (re.compile(r"fused_gdn_gating"), "GDN_GATE"),
            (re.compile(r"fused_sigmoid_gating_delta_rule"), "GDN_RECUR"),
            (re.compile(r"fused_qkvzba_split"), "GDN_QKV_SPLIT"),
            (re.compile(r"_causal_conv1d_fwd"), "CONV1D_FWD"),
            (re.compile(r"_causal_conv1d_update"), "CONV1D_UPD"),
            (re.compile(r"chunk_local_cumsum"), "GDN_CUMSUM"),
            (re.compile(r"chunk_scaled_dot_kkt"), "GDN_DOT"),
            (re.compile(r"solve_tril"), "GDN_TRIL"),
            (re.compile(r"merge_16x16_to_64x64"), "GDN_MERGE"),
            (re.compile(r"recompute_w_u_fwd"), "GDN_RECOMP"),
            (re.compile(r"l2norm_fwd"), "L2NORM"),
            (re.compile(r"_layer_norm_fwd_1pass"), "LAYERNORM"),
        ]

    def classify_stage(
        self, name: str, grid: Optional[Tuple[int, int, int]] = None
    ) -> Stage:
        """Classify kernel stage based on name and grid dimensions."""
        for pattern in self._stage_force_unknown_patterns:
            if pattern.search(name):
                return Stage.UNKNOWN
        for pattern, stage in self._stage_patterns:
            if pattern.search(name):
                return stage

        if grid is not None and len(grid) >= 1:
            grid_0 = grid[0]
            if grid_0 > self.grid_threshold:
                return Stage.DECODE
            elif grid_0 < 200:
                return Stage.PREFILL

        return Stage.UNKNOWN

    def classify_type(self, name: str) -> KernelType:
        """Classify kernel type based on name patterns."""
        for pattern, kernel_type in self._type_patterns:
            if pattern.search(name):
                return kernel_type
        return KernelType.OTHER

    def simplify_name(self, name: str) -> str:
        """Get simplified kernel name for display."""
        for pattern, simple_name in self._simplify_patterns:
            if pattern.search(name):
                return simple_name
        # Truncate unknown names
        return name[:25] if len(name) > 25 else name


class LayerDetector:
    """Detects layer boundaries in kernel sequences."""

    def __init__(self, mtp_qseqlen_decode: bool = False):
        self._mtp_qseqlen_decode = mtp_qseqlen_decode
        # Layer start markers (after ALLREDUCE)
        self._layer_start_patterns = [
            re.compile(
                r"_fused_rms_mxfp4_quant|_fused_rms_fp8|fused_dual_residual_rmsnorm_kernel|fused_rmsnorm_kernel"
            ),  # Prefill/decode layernorm (quantized fused or standard)
        ]

        # MoE block markers
        self._moe_markers = [
            re.compile(
                r"MoeSorting|moe_fused_gate|kernel_moe_gemm|kernel_moe_mxgemm|_moe_mxfp4_sort"
            ),
        ]

        # FC block markers (activation function after linear)
        self._fc_markers = [
            re.compile(r"act_and_mul_kernel|silu_kernel|gelu_and_mul_kernel"),
        ]

        # MLA-specific attention markers (vs generic MHA/FMHA)
        self._mla_markers = [
            re.compile(
                r"aiter::mla_|mla_a8w8|set_mla_kv_buffer|kn_mla_reduce|"
                r"mla_reduce|concat_and_cast_mha_k"
            ),
        ]

        # GDN (gated delta network) attention markers — Qwen3-Coder-Next
        self._gdn_markers = [
            re.compile(
                r"chunk_gated_delta_rule|fused_gdn_gating|"
                r"fused_sigmoid_gating_delta_rule|"
                r"_causal_conv1d_|chunk_fwd_kernel_o|"
                r"chunk_local_cumsum|chunk_scaled_dot_kkt"
            ),
        ]

        # Attention stage hints for layer stage inference
        self._attention_decode_patterns = [
            re.compile(r"qseqlen1[^0-9]|qseqlen1$|decode_attention|paged_attention_ll4mi", re.IGNORECASE),
            re.compile(r"mla_a8w8.*qseqlen(1|4)", re.IGNORECASE),
            # GDN decode: recurrent update kernels (vs chunk kernels for prefill)
            re.compile(r"fused_sigmoid_gating_delta_rule_update|_causal_conv1d_update", re.IGNORECASE),
        ]
        if self._mtp_qseqlen_decode:
            self._attention_decode_patterns.append(
                re.compile(r"qseqlen[2-9]|qseqlen\d{2,}", re.IGNORECASE)
            )
        self._attention_prefill_patterns = [
            re.compile(r"fmha_fwd|flash_attn|FmhaBatchPrefill", re.IGNORECASE),
            # GDN prefill: chunk-based kernels (vs recurrent update for decode)
            re.compile(r"chunk_gated_delta_rule_fwd|fused_gdn_gating|_causal_conv1d_fwd", re.IGNORECASE),
        ]
        if not self._mtp_qseqlen_decode:
            self._attention_prefill_patterns.append(
                re.compile(r"qseqlen[2-9]|qseqlen\d{2,}", re.IGNORECASE)
            )

    @staticmethod
    def _preceded_by_allreduce(kernels: List[KernelEvent], lookback: int = 15) -> bool:
        """Check if ALLREDUCE appeared recently with only OTHER/MEMORY kernels between.

        Models with unfused RMSNorm (e.g. Qwen3-Coder-Next) have a sequence of
        generic elementwise kernels between ALLREDUCE and the first computational
        kernel.  This helper looks backward through the kernel list and returns
        True if ALLREDUCE is found within *lookback* positions with only
        OTHER / MEMORY typed kernels in between.
        """
        for i in range(len(kernels) - 1, max(-1, len(kernels) - 1 - lookback), -1):
            k = kernels[i]
            if k.simplified_name == "ALLREDUCE":
                return True
            if k.kernel_type not in (KernelType.OTHER, KernelType.MEMORY):
                return False
        return False

    def detect_layers(self, events: List[KernelEvent]) -> List[LayerEvent]:
        """Detect layer boundaries and classify layers.

        Layer stage is determined by the dominant stage of kernels within each layer.

        Two boundary detection strategies:
        1. Primary: a fused RMSNorm kernel immediately after ALLREDUCE
           (DeepSeek, Grok, etc.)
        2. Fallback: first LINEAR (GEMM) kernel after ALLREDUCE when only
           OTHER/MEMORY kernels appear in between — handles models with
           unfused normalization (Qwen3-Coder-Next).
        """
        if not events:
            return []

        layers = []
        current_layer_kernels: List[KernelEvent] = []
        layer_idx = 0
        in_layer = False
        saw_attention = False
        saw_mla = False
        saw_gdn = False
        saw_moe = False
        saw_fc = False

        for i, event in enumerate(events):
            name = event.name
            simplified = event.simplified_name

            # Detect layer start: ALLREDUCE followed by RMS layernorm
            is_allreduce = (
                simplified == "ALLREDUCE"
                or "rcclGenericKernel" in name
                or "cross_device_reduce" in name
            )
            is_layer_start = any(p.search(name) for p in self._layer_start_patterns)

            # Fallback: first GEMM after ALLREDUCE + unfused norm sequence.
            # Only applies when no fused RMSNorm kernel appeared between
            # ALLREDUCE and this GEMM (otherwise the primary detection
            # handles it).
            is_gemm_after_allreduce = False
            if (
                not is_layer_start
                and event.kernel_type == KernelType.LINEAR
                and len(current_layer_kernels) > 0
                and self._preceded_by_allreduce(current_layer_kernels)
            ):
                # Verify no layer_start pattern kernel exists between
                # ALLREDUCE and here (otherwise primary path should handle)
                has_fused_norm = False
                for j in range(len(current_layer_kernels) - 1, -1, -1):
                    k = current_layer_kernels[j]
                    if k.simplified_name == "ALLREDUCE":
                        break
                    if any(p.search(k.name) for p in self._layer_start_patterns):
                        has_fused_norm = True
                        break
                if not has_fused_norm:
                    is_gemm_after_allreduce = True

            # Check if this is a MoE, FC, MLA, or GDN marker
            is_moe = any(p.search(name) for p in self._moe_markers)
            is_fc = any(p.search(name) for p in self._fc_markers)
            is_attention = event.kernel_type == KernelType.ATTENTION
            is_mla = any(p.search(name) for p in self._mla_markers)
            is_gdn = any(p.search(name) for p in self._gdn_markers)

            # Layer boundary: named pattern after ALLREDUCE, or fallback GEMM
            is_boundary = False
            if is_layer_start and (
                not in_layer
                or (
                    len(current_layer_kernels) > 0
                    and (
                        current_layer_kernels[-1].simplified_name == "ALLREDUCE"
                        or self._preceded_by_allreduce(current_layer_kernels)
                    )
                )
            ):
                is_boundary = True
            elif is_gemm_after_allreduce:
                is_boundary = True

            if is_boundary:
                # Save previous layer if exists
                if current_layer_kernels:
                    layer_type = self._classify_layer_type(
                        saw_attention, saw_mla, saw_gdn, saw_moe, saw_fc
                    )
                    # Determine stage from kernels
                    stage = self._determine_layer_stage(current_layer_kernels)
                    layers.append(
                        LayerEvent(
                            layer_idx=layer_idx,
                            layer_type=layer_type,
                            stage=stage,
                            kernels=current_layer_kernels.copy(),
                        )
                    )
                    layer_idx += 1

                # Start new layer — include preceding ALLREDUCE + unfused
                # norm kernels so they belong to the new layer.
                allreduce_pos = None
                for j in range(len(current_layer_kernels) - 1, -1, -1):
                    if current_layer_kernels[j].simplified_name == "ALLREDUCE":
                        allreduce_pos = j
                        break
                    if current_layer_kernels[j].kernel_type not in (
                        KernelType.OTHER,
                        KernelType.MEMORY,
                    ):
                        break

                if allreduce_pos is not None:
                    current_layer_kernels = (
                        current_layer_kernels[allreduce_pos:] + [event]
                    )
                else:
                    current_layer_kernels = [event]
                in_layer = True
                saw_attention = False
                saw_mla = False
                saw_gdn = False
                saw_moe = False
                saw_fc = False
            else:
                current_layer_kernels.append(event)

            # Track what we've seen in this layer
            if is_attention:
                saw_attention = True
            if is_mla:
                saw_mla = True
            if is_gdn:
                saw_gdn = True
            if is_moe:
                saw_moe = True
            if is_fc:
                saw_fc = True

        # Save last layer
        if current_layer_kernels:
            layer_type = self._classify_layer_type(
                saw_attention, saw_mla, saw_gdn, saw_moe, saw_fc
            )
            stage = self._determine_layer_stage(current_layer_kernels)
            layers.append(
                LayerEvent(
                    layer_idx=layer_idx,
                    layer_type=layer_type,
                    stage=stage,
                    kernels=current_layer_kernels,
                )
            )

        return layers

    @staticmethod
    def _classify_layer_type(
        saw_attention: bool,
        saw_mla: bool,
        saw_gdn: bool,
        saw_moe: bool,
        saw_fc: bool,
    ) -> LayerType:
        """Classify layer type from what markers were observed.

        Attention variant priority: MLA > GDN > MHA (generic).
        """
        if saw_attention:
            if saw_mla:
                if saw_moe:
                    return LayerType.MLA_MOE
                if saw_fc:
                    return LayerType.MLA_FC
            elif saw_gdn:
                if saw_moe:
                    return LayerType.GDN_MOE
                if saw_fc:
                    return LayerType.GDN_FC
            else:
                if saw_moe:
                    return LayerType.MHA_MOE
                if saw_fc:
                    return LayerType.MHA_FC
            return LayerType.ATTN
        # No attention — half-layer
        if saw_moe:
            return LayerType.MOE
        if saw_fc:
            return LayerType.FC
        return LayerType.UNKNOWN

    def _determine_layer_stage(self, kernels: List[KernelEvent]) -> Stage:
        """Determine layer stage based on dominant stage of its kernels."""
        prefill_count = sum(1 for k in kernels if k.stage == Stage.PREFILL)
        decode_count = sum(1 for k in kernels if k.stage == Stage.DECODE)

        if prefill_count > decode_count:
            return Stage.PREFILL
        if decode_count > prefill_count:
            return Stage.DECODE

        return Stage.UNKNOWN

    def merge_half_layers(self, layers: List[LayerEvent]) -> List[LayerEvent]:
        """Merge adjacent Attn+MoE or Attn+FC half-layers into full layers.

        Models like Grok-2 produce two half-layers per transformer block
        (one attention-only, one MoE/FC-only).  This pass recombines them
        into single full layers so layer counts match the model's
        num_hidden_layers.  The merged type is MLA or MHA based on
        whether MLA-specific kernels appear in the attention half.
        """
        merged: List[LayerEvent] = []
        i = 0
        idx = 0
        while i < len(layers):
            layer = layers[i]
            if (
                layer.layer_type == LayerType.ATTN
                and i + 1 < len(layers)
                and layers[i + 1].layer_type in (LayerType.MOE, LayerType.FC)
            ):
                next_layer = layers[i + 1]
                has_mla = any(
                    p.search(k.name)
                    for k in layer.kernels
                    for p in self._mla_markers
                )
                has_gdn = any(
                    p.search(k.name)
                    for k in layer.kernels
                    for p in self._gdn_markers
                )
                is_moe = next_layer.layer_type == LayerType.MOE
                if has_mla:
                    merged_type = LayerType.MLA_MOE if is_moe else LayerType.MLA_FC
                elif has_gdn:
                    merged_type = LayerType.GDN_MOE if is_moe else LayerType.GDN_FC
                else:
                    merged_type = LayerType.MHA_MOE if is_moe else LayerType.MHA_FC
                combined_kernels = layer.kernels + next_layer.kernels
                stage = self._determine_layer_stage(combined_kernels)
                merged.append(
                    LayerEvent(
                        layer_idx=idx,
                        layer_type=merged_type,
                        stage=stage,
                        kernels=combined_kernels,
                    )
                )
                i += 2
            else:
                merged.append(
                    LayerEvent(
                        layer_idx=idx,
                        layer_type=layer.layer_type,
                        stage=layer.stage,
                        kernels=layer.kernels,
                    )
                )
                i += 1
            idx += 1
        return merged

    def _infer_stage_from_names(self, kernels: List[KernelEvent]) -> Stage:
        for k in kernels:
            name = k.name
            if any(p.search(name) for p in self._attention_decode_patterns):
                return Stage.DECODE
            if any(p.search(name) for p in self._attention_prefill_patterns):
                return Stage.PREFILL
        return Stage.UNKNOWN


class TraceAnalyzer:
    """Main analyzer that loads and processes trace files."""

    def __init__(
        self,
        grid_threshold: int = 10000,
        mtp_qseqlen_decode: bool = False,
        split_half_layers: bool = False,
    ):
        self.classifier = KernelClassifier(grid_threshold=grid_threshold)
        self.layer_detector = LayerDetector(mtp_qseqlen_decode=mtp_qseqlen_decode)
        self.split_half_layers = split_half_layers

    def load_trace(self, path: str) -> Dict[str, Any]:
        """Load trace file (supports .json.gz and .json)."""
        logger.info(f"Loading trace file: {path}")

        if path.endswith(".gz"):
            with gzip.open(path, "rt", encoding="utf-8") as f:
                return json.load(f)
        else:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

    def _parse_grid(self, grid_data: Any) -> Optional[Tuple[int, int, int]]:
        """Parse grid dimensions from trace data."""
        if grid_data is None:
            return None

        if isinstance(grid_data, list) and len(grid_data) >= 3:
            return (int(grid_data[0]), int(grid_data[1]), int(grid_data[2]))

        if isinstance(grid_data, str):
            try:
                parsed = json.loads(grid_data)
                if isinstance(parsed, list) and len(parsed) >= 3:
                    return (int(parsed[0]), int(parsed[1]), int(parsed[2]))
            except (json.JSONDecodeError, ValueError):
                pass

        return None

    def analyze(
        self, trace_data: Dict[str, Any], detect_layers: bool = False
    ) -> AnalysisResult:
        """Analyze trace data and compute statistics."""
        result = AnalysisResult()

        events = trace_data.get("traceEvents", [])
        logger.info(f"Total events in trace: {len(events)}")

        # Filter kernel events
        kernel_events = [
            e for e in events if e.get("ph") == "X" and e.get("cat") == "kernel"
        ]
        logger.info(f"Kernel events found: {len(kernel_events)}")

        # Sort by timestamp
        kernel_events.sort(key=lambda x: x.get("ts", 0))

        # Process each kernel event
        for event in kernel_events:
            name = event.get("name", "")
            timestamp_us = event.get("ts", 0)
            duration_us = event.get("dur", 0)

            args = event.get("args", {})
            grid = self._parse_grid(args.get("grid"))
            block = self._parse_grid(args.get("block"))

            # Classify
            stage = self.classifier.classify_stage(name, grid)
            kernel_type = self.classifier.classify_type(name)
            simplified_name = self.classifier.simplify_name(name)

            if kernel_type in (
                KernelType.COMMUNICATION,
                KernelType.QUANTIZATION,
                KernelType.MOE,
                KernelType.LINEAR,
                KernelType.MEMORY,
            ):
                stage = Stage.UNKNOWN

            # Create event
            kernel_event = KernelEvent(
                name=name,
                timestamp_us=timestamp_us,
                duration_us=duration_us,
                stage=stage,
                kernel_type=kernel_type,
                initial_stage=stage,
                grid=grid,
                block=block,
                simplified_name=simplified_name,
            )
            result.events.append(kernel_event)

            # Update stats
            if name not in result.per_kernel_stats:
                result.per_kernel_stats[name] = KernelStats()
            result.per_kernel_stats[name].add(duration_us)

            if stage not in result.per_stage_stats:
                result.per_stage_stats[stage] = KernelStats()
            result.per_stage_stats[stage].add(duration_us)

            if kernel_type not in result.per_type_stats:
                result.per_type_stats[kernel_type] = KernelStats()
            result.per_type_stats[kernel_type].add(duration_us)

            stage_type_key = (stage, kernel_type)
            if stage_type_key not in result.per_stage_type_stats:
                result.per_stage_type_stats[stage_type_key] = KernelStats()
            result.per_stage_type_stats[stage_type_key].add(duration_us)

            stage_kernel_key = (stage, name)
            if stage_kernel_key not in result.per_stage_kernel_stats:
                result.per_stage_kernel_stats[stage_kernel_key] = KernelStats()
            result.per_stage_kernel_stats[stage_kernel_key].add(duration_us)

        # Detect layers if requested
        if detect_layers:
            # Detect layers from ALL events (don't filter by stage)
            # Layer stage is determined by dominant stage within each layer
            all_layers = self.layer_detector.detect_layers(result.events)

            # Merge adjacent Attn+MoE / Attn+FC half-layers into full
            # layers when requested (e.g. Grok-2 dual-norm architecture).
            if not self.split_half_layers:
                before = len(all_layers)
                all_layers = self.layer_detector.merge_half_layers(all_layers)
                logger.info(
                    f"Merged half-layers: {before} -> {len(all_layers)} layers"
                )

            # Stage propagation: MoE/FC half-layers inherit stage from
            # the immediately preceding layer (the attention half of the
            # same transformer block).  This fixes models like Grok-2 where
            # dual-residual RMSNorm fires twice per block, producing an
            # attention half-layer followed by a MoE/FC half-layer whose
            # tied vote would otherwise leave stage=UNKNOWN.
            for i, layer in enumerate(all_layers):
                if layer.stage == Stage.UNKNOWN and i > 0:
                    prev = all_layers[i - 1]
                    if prev.stage in (Stage.PREFILL, Stage.DECODE):
                        layer.stage = prev.stage

            # Count layers by stage
            prefill_count = sum(1 for l in all_layers if l.stage == Stage.PREFILL)
            decode_count = sum(1 for l in all_layers if l.stage == Stage.DECODE)
            unknown_count = sum(1 for l in all_layers if l.stage == Stage.UNKNOWN)

            result.layers = all_layers
            logger.info(
                f"Detected {prefill_count} prefill layers, {decode_count} decode layers"
                + (f", {unknown_count} unknown layers" if unknown_count else "")
            )
            # Re-assign kernel stages based on layer vote (only for prefill/decode layers)
            for layer in result.layers:
                if layer.stage in (Stage.PREFILL, Stage.DECODE):
                    for k in layer.kernels:
                        k.stage = layer.stage

            # Recompute stage-related stats after layer assignment
            result.per_stage_stats = {}
            result.per_stage_type_stats = {}
            result.per_stage_kernel_stats = {}
            for event in result.events:
                stage = event.stage
                kernel_type = event.kernel_type
                name = event.name
                duration_us = event.duration_us

                if stage not in result.per_stage_stats:
                    result.per_stage_stats[stage] = KernelStats()
                result.per_stage_stats[stage].add(duration_us)

                stage_type_key = (stage, kernel_type)
                if stage_type_key not in result.per_stage_type_stats:
                    result.per_stage_type_stats[stage_type_key] = KernelStats()
                result.per_stage_type_stats[stage_type_key].add(duration_us)

                stage_kernel_key = (stage, name)
                if stage_kernel_key not in result.per_stage_kernel_stats:
                    result.per_stage_kernel_stats[stage_kernel_key] = KernelStats()
                result.per_stage_kernel_stats[stage_kernel_key].add(duration_us)

        return result


class ReportFormatter:
    """Formats analysis results for console output."""

    def __init__(self, result: AnalysisResult):
        self.result = result

    def format_time(self, time_ms: float) -> str:
        """Format time value with appropriate unit."""
        if time_ms >= 1000:
            return f"{time_ms / 1000:.3f} s"
        elif time_ms >= 1:
            return f"{time_ms:.3f} ms"
        else:
            return f"{time_ms * 1000:.3f} us"

    def format_time_us(self, time_us: float) -> str:
        """Format time in microseconds with appropriate unit."""
        return self.format_time(time_us / 1000.0)

    def format_percentage(self, part: float, total: float) -> str:
        """Format percentage."""
        if total == 0:
            return "0.0%"
        return f"{100.0 * part / total:.1f}%"

    def print_summary(self) -> None:
        """Print overall summary."""
        print("=" * 80)
        print("TRACE ANALYSIS SUMMARY")
        print("=" * 80)
        print()

        total_time_ms = self.result.total_time_ms
        print(f"Total Kernel Time:   {self.format_time(total_time_ms)}")

        for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
            stats = self.result.per_stage_stats.get(stage, KernelStats())
            stage_time_ms = stats.total_time_ms
            pct = self.format_percentage(stage_time_ms, total_time_ms)
            print(
                f"  - {stage.value.capitalize():12s} {self.format_time(stage_time_ms):>12s} ({pct})"
            )

        print()

    def print_type_breakdown(self) -> None:
        """Print breakdown by kernel type."""
        print("-" * 80)
        print("BREAKDOWN BY KERNEL TYPE")
        print("-" * 80)
        print()
        print(f"{'Type':<16} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}")
        print("-" * 64)

        total_time_ms = self.result.total_time_ms

        sorted_types = sorted(
            self.result.per_type_stats.items(),
            key=lambda x: x[1].total_time_us,
            reverse=True,
        )

        for kernel_type, stats in sorted_types:
            pct = self.format_percentage(stats.total_time_ms, total_time_ms)
            print(
                f"{kernel_type.value:<16} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}"
            )

        print()

    def print_stage_type_breakdown(self, stage: Stage) -> None:
        """Print type breakdown for a specific stage."""
        print("-" * 80)
        print(f"BREAKDOWN BY TYPE ({stage.value.upper()})")
        print("-" * 80)
        print()
        print(f"{'Type':<16} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}")
        print("-" * 64)

        stage_stats = self.result.per_stage_stats.get(stage, KernelStats())
        stage_total_ms = stage_stats.total_time_ms

        stage_type_items = [
            (kt, stats)
            for (s, kt), stats in self.result.per_stage_type_stats.items()
            if s == stage
        ]
        sorted_items = sorted(
            stage_type_items, key=lambda x: x[1].total_time_us, reverse=True
        )

        for kernel_type, stats in sorted_items:
            pct = self.format_percentage(stats.total_time_ms, stage_total_ms)
            print(
                f"{kernel_type.value:<16} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}"
            )

        print()

    def print_top_kernels(self, stage: Optional[Stage] = None, top_n: int = 20) -> None:
        """Print top kernels by total time."""
        if stage is not None:
            title = f"TOP {top_n} KERNELS ({stage.value.upper()})"
            kernel_items = [
                (name, stats)
                for (s, name), stats in self.result.per_stage_kernel_stats.items()
                if s == stage
            ]
            stage_stats = self.result.per_stage_stats.get(stage, KernelStats())
            total_time_ms = stage_stats.total_time_ms
        else:
            title = f"TOP {top_n} KERNELS (ALL)"
            kernel_items = list(self.result.per_kernel_stats.items())
            total_time_ms = self.result.total_time_ms

        print("-" * 80)
        print(title)
        print("-" * 80)
        print()
        print(
            f"{'#':<4} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}  {'Kernel Name'}"
        )
        print("-" * 80)

        sorted_kernels = sorted(
            kernel_items, key=lambda x: x[1].total_time_us, reverse=True
        )

        for i, (name, stats) in enumerate(sorted_kernels[:top_n], 1):
            pct = self.format_percentage(stats.total_time_ms, total_time_ms)
            display_name = name[:60] + "..." if len(name) > 63 else name
            print(
                f"{i:<4} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}  {display_name}"
            )

        print()

    def print_layer_summary(self, stage: Optional[Stage] = None) -> None:
        """Print layer-level summary."""
        if not self.result.layers:
            print("No layer analysis available.")
            return

        # Filter layers by stage if specified
        layers = self.result.layers
        if stage is not None:
            layers = [l for l in layers if l.stage == stage]

        if not layers:
            return

        stage_name = stage.value.upper() if stage else "ALL"
        print("-" * 100)
        print(f"LAYER ANALYSIS ({stage_name})")
        print("-" * 100)
        print()
        print(
            f"{'Layer':<8} {'Type':<10} {'Total':>12} {'Attn':>12} {'MoE/FC':>12} {'Comm':>12} {'Quant':>12} {'#Kern':>8}"
        )
        print("-" * 100)

        for layer in layers:
            moe_fc_time = (
                layer.moe_time_us
                if layer.layer_type in (LayerType.MLA_MOE, LayerType.MHA_MOE, LayerType.GDN_MOE, LayerType.MOE)
                else layer.linear_time_us
            )
            print(
                f"{layer.layer_idx:<8} {layer.layer_type.value:<10} "
                f"{self.format_time_us(layer.total_time_us):>12} "
                f"{self.format_time_us(layer.attention_time_us):>12} "
                f"{self.format_time_us(moe_fc_time):>12} "
                f"{self.format_time_us(layer.communication_time_us):>12} "
                f"{self.format_time_us(layer.quantization_time_us):>12} "
                f"{len(layer.kernels):>8}"
            )

        print()

    def print_layer_detail(self, layer_idx: int, stage: Optional[Stage] = None) -> None:
        """Print detailed kernel sequence for a specific layer."""
        if not self.result.layers:
            print("No layer analysis available.")
            return

        # Find the layer
        layer = None
        for l in self.result.layers:
            if l.layer_idx == layer_idx and (stage is None or l.stage == stage):
                layer = l
                break

        if layer is None:
            print(f"Layer {layer_idx} not found.")
            return

        # Filter out trailing ALLREDUCE (belongs to next layer)
        kernels = layer.kernels
        if kernels and kernels[-1].simplified_name == "ALLREDUCE":
            kernels = kernels[:-1]

        # Recalculate total time without trailing ALLREDUCE
        total_time = sum(k.duration_us for k in kernels)

        print("-" * 180)
        print(
            f"LAYER {layer.layer_idx} DETAIL ({layer.layer_type.value}, {layer.stage.value.upper()})"
        )
        print("-" * 180)
        print()
        print(
            f"{'#':<4} {'Duration (us)':>14} {'%':>7} {'Type':<14} {'Short Name':<18} {'Kernel Name'}"
        )
        print("-" * 180)

        for i, kernel in enumerate(kernels, 1):
            pct = self.format_percentage(kernel.duration_us, total_time)
            short_name = kernel.simplified_name if kernel.simplified_name else "-"
            # Truncate kernel name to 50 characters
            display_name = kernel.name[:50] + "..." if len(kernel.name) > 50 else kernel.name
            print(
                f"{i:<4} {kernel.duration_us:>14.3f} {pct:>7} "
                f"{kernel.kernel_type.value:<14} {short_name:<18} {display_name}"
            )

        print()
        print(f"Layer Total: {total_time:.3f} us ({total_time/1000:.3f} ms)")

        # Print breakdown (recalculated without trailing ALLREDUCE)
        breakdown = {}
        for k in kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us

        print()
        print("Breakdown:")
        for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
            pct = self.format_percentage(time_us, total_time)
            print(f"  - {kt.value:<14} {time_us:>14.3f} us ({pct})")

        print()

    def print_full_report(
        self, top_n: int = 20, stage_filter: Optional[str] = None
    ) -> None:
        """Print full analysis report."""
        self.print_summary()
        self.print_type_breakdown()

        if stage_filter is None or stage_filter == "all":
            for stage in [Stage.PREFILL, Stage.DECODE]:
                if stage in self.result.per_stage_stats:
                    self.print_stage_type_breakdown(stage)
                    self.print_top_kernels(stage=stage, top_n=top_n)
        elif stage_filter == "prefill":
            self.print_stage_type_breakdown(Stage.PREFILL)
            self.print_top_kernels(stage=Stage.PREFILL, top_n=top_n)
        elif stage_filter == "decode":
            self.print_stage_type_breakdown(Stage.DECODE)
            self.print_top_kernels(stage=Stage.DECODE, top_n=top_n)

    def print_layer_debug(self, layer_idx: int, top_n: int = 20) -> None:
        """Print debug details for a specific layer's stage decision."""
        if not self.result.layers:
            print("No layer analysis available.")
            return

        layer = None
        for l in self.result.layers:
            if l.layer_idx == layer_idx:
                layer = l
                break

        if layer is None:
            print(f"Layer {layer_idx} not found.")
            return

        prefill_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.PREFILL)
        decode_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.DECODE)
        unknown_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.UNKNOWN)

        print("-" * 120)
        print(f"LAYER DEBUG {layer.layer_idx} ({layer.layer_type.value})")
        print("-" * 120)
        print(f"Layer stage (after vote): {layer.stage.value}")
        print(
            f"Vote counts (initial stages): prefill={prefill_count}, decode={decode_count}, unknown={unknown_count}"
        )
        print()

        sorted_kernels = sorted(layer.kernels, key=lambda k: k.duration_us, reverse=True)
        print(
            f"{'#':<4} {'Duration (us)':>14} {'Type':<14} {'Init':<8} {'Final':<8} {'Kernel Name'}"
        )
        print("-" * 120)
        for i, k in enumerate(sorted_kernels[:top_n], 1):
            display_name = k.name[:80] + "..." if len(k.name) > 83 else k.name
            print(
                f"{i:<4} {k.duration_us:>14.3f} {k.kernel_type.value:<14} "
                f"{k.initial_stage.value:<8} {k.stage.value:<8} {display_name}"
            )
        print()

    def print_auto_layer_details(self, num_layers: int = 2) -> None:
        """Auto-detect and print details for first N prefill and decode layers."""
        if not self.result.layers:
            print("No layer analysis available.")
            return

        prefill_layers = [l for l in self.result.layers if l.stage == Stage.PREFILL]
        decode_layers = [l for l in self.result.layers if l.stage == Stage.DECODE]

        # Sort by layer_idx to get the first ones
        prefill_layers.sort(key=lambda l: l.layer_idx)
        decode_layers.sort(key=lambda l: l.layer_idx)

        selected_prefill = prefill_layers[:num_layers]
        selected_decode = decode_layers[:num_layers]

        print("=" * 80)
        print("AUTO-DETECTED LAYER DETAILS")
        print("=" * 80)
        print()
        print(
            f"Total layers detected: {len(self.result.layers)} "
            f"(prefill: {len(prefill_layers)}, decode: {len(decode_layers)})"
        )
        print()

        if selected_prefill:
            print(f"First prefill layer: layer {selected_prefill[0].layer_idx}")
        else:
            print("No prefill layers detected.")

        if selected_decode:
            print(f"First decode layer: layer {selected_decode[0].layer_idx}")
        else:
            print("No decode layers detected.")

        print()

        for layer in selected_prefill:
            self.print_layer_detail(layer.layer_idx)

        for layer in selected_decode:
            self.print_layer_detail(layer.layer_idx)


class CSVExporter:
    """Exports analysis results to CSV files."""

    def __init__(self, result: AnalysisResult):
        self.result = result

    def export_kernel_stats(self, path: str) -> None:
        """Export per-kernel statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "kernel_name",
                    "count",
                    "total_time_us",
                    "avg_time_us",
                    "min_time_us",
                    "max_time_us",
                ]
            )

            for name, stats in sorted(
                self.result.per_kernel_stats.items(),
                key=lambda x: x[1].total_time_us,
                reverse=True,
            ):
                writer.writerow(
                    [
                        name,
                        stats.count,
                        f"{stats.total_time_us:.3f}",
                        f"{stats.avg_time_us:.3f}",
                        (
                            f"{stats.min_time_us:.3f}"
                            if stats.min_time_us != float("inf")
                            else "0"
                        ),
                        f"{stats.max_time_us:.3f}",
                    ]
                )

        logger.info(f"Kernel stats exported to: {path}")

    def export_events(self, path: str) -> None:
        """Export all kernel events to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "timestamp_us",
                    "duration_us",
                    "stage",
                    "kernel_type",
                    "simplified_name",
                    "grid",
                    "block",
                    "kernel_name",
                ]
            )

            for event in self.result.events:
                writer.writerow(
                    [
                        f"{event.timestamp_us:.3f}",
                        f"{event.duration_us:.3f}",
                        event.stage.value,
                        event.kernel_type.value,
                        event.simplified_name,
                        str(event.grid) if event.grid else "",
                        str(event.block) if event.block else "",
                        event.name,
                    ]
                )

        logger.info(f"Events exported to: {path}")

    def export_layers(self, path: str) -> None:
        """Export layer-level statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "layer_idx",
                    "layer_type",
                    "stage",
                    "total_time_us",
                    "attention_time_us",
                    "moe_time_us",
                    "linear_time_us",
                    "communication_time_us",
                    "quantization_time_us",
                    "kernel_count",
                ]
            )

            for layer in self.result.layers:
                writer.writerow(
                    [
                        layer.layer_idx,
                        layer.layer_type.value,
                        layer.stage.value,
                        f"{layer.total_time_us:.3f}",
                        f"{layer.attention_time_us:.3f}",
                        f"{layer.moe_time_us:.3f}",
                        f"{layer.linear_time_us:.3f}",
                        f"{layer.communication_time_us:.3f}",
                        f"{layer.quantization_time_us:.3f}",
                        len(layer.kernels),
                    ]
                )

        logger.info(f"Layer stats exported to: {path}")

    def export_full_analysis(
        self,
        output_path: str,
        layer_indices: Optional[List[int]] = None,
        stage_filter: Optional[Stage] = None,
    ) -> str:
        """Export full analysis to Excel file with multiple sheets.

        Creates sheets:
        - Summary: Overall summary and breakdown
        - Layer_N: Detailed kernel list for each specified layer

        Args:
            output_path: Output Excel file path (.xlsx)
            layer_indices: List of layer indices to export (None = all layers)
            stage_filter: Filter layers by stage (None = all stages)

        Returns:
            Path to created Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        wb = Workbook()

        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary)

        # Filter layers
        layers = self.result.layers
        if stage_filter is not None:
            layers = [l for l in layers if l.stage == stage_filter]

        if layer_indices is not None:
            layers = [l for l in layers if l.layer_idx in layer_indices]

        # Create sheet for each layer
        for layer in layers:
            ws_layer = wb.create_sheet(title=f"Layer_{layer.layer_idx}")
            self._write_layer_sheet(ws_layer, layer)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Exported Excel file: {output_path}")
        return output_path

    def _write_summary_sheet(self, ws) -> None:
        """Write summary statistics to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        row = 1

        # Overall summary section
        ws.cell(row=row, column=1, value="=== OVERALL SUMMARY ===").font = Font(bold=True, size=12)
        row += 1

        headers = ["Metric", "Value (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        total_time_us = self.result.total_time_us
        ws.cell(row=row, column=1, value="Total Kernel Time")
        ws.cell(row=row, column=2, value=round(total_time_us, 3))
        ws.cell(row=row, column=3, value="100.0%")
        row += 1

        for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
            stats = self.result.per_stage_stats.get(stage, KernelStats())
            pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
            ws.cell(row=row, column=1, value=f"{stage.value.capitalize()} Time")
            ws.cell(row=row, column=2, value=round(stats.total_time_us, 3))
            ws.cell(row=row, column=3, value=pct)
            row += 1

        row += 1

        # Type breakdown section
        ws.cell(row=row, column=1, value="=== BREAKDOWN BY TYPE ===").font = Font(bold=True, size=12)
        row += 1

        headers = ["Type", "Count", "Total (us)", "Avg (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        sorted_types = sorted(
            self.result.per_type_stats.items(),
            key=lambda x: x[1].total_time_us,
            reverse=True,
        )
        for kernel_type, stats in sorted_types:
            pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
            ws.cell(row=row, column=1, value=kernel_type.value)
            ws.cell(row=row, column=2, value=stats.count)
            ws.cell(row=row, column=3, value=round(stats.total_time_us, 3))
            ws.cell(row=row, column=4, value=round(stats.avg_time_us, 3))
            ws.cell(row=row, column=5, value=pct)
            row += 1

        row += 1

        # Layer summary section
        if self.result.layers:
            ws.cell(row=row, column=1, value="=== LAYER SUMMARY ===").font = Font(bold=True, size=12)
            row += 1

            headers = ["Layer", "Type", "Stage", "Total (us)", "Attention (us)", "MoE (us)",
                       "Linear (us)", "Comm (us)", "Quant (us)", "Kernels"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1

            for layer in self.result.layers:
                ws.cell(row=row, column=1, value=layer.layer_idx)
                ws.cell(row=row, column=2, value=layer.layer_type.value)
                ws.cell(row=row, column=3, value=layer.stage.value)
                ws.cell(row=row, column=4, value=round(layer.total_time_us, 3))
                ws.cell(row=row, column=5, value=round(layer.attention_time_us, 3))
                ws.cell(row=row, column=6, value=round(layer.moe_time_us, 3))
                ws.cell(row=row, column=7, value=round(layer.linear_time_us, 3))
                ws.cell(row=row, column=8, value=round(layer.communication_time_us, 3))
                ws.cell(row=row, column=9, value=round(layer.quantization_time_us, 3))
                ws.cell(row=row, column=10, value=len(layer.kernels))
                row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    def _write_layer_sheet(self, ws, layer: LayerEvent) -> None:
        """Write layer detail to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Filter out trailing ALLREDUCE
        kernels = layer.kernels
        if kernels and kernels[-1].simplified_name == "ALLREDUCE":
            kernels = kernels[:-1]

        total_time = sum(k.duration_us for k in kernels)

        row = 1

        # Layer header
        ws.cell(row=row, column=1, value=f"Layer {layer.layer_idx} ({layer.layer_type.value}, {layer.stage.value})").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value=f"Total Time: {total_time:.3f} us ({total_time/1000:.3f} ms)")
        row += 2

        # Kernel details header
        headers = ["#", "Duration (us)", "%", "Type", "Short Name", "Kernel Name"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        # Kernel details
        for i, kernel in enumerate(kernels, 1):
            pct = f"{100.0 * kernel.duration_us / total_time:.1f}%" if total_time > 0 else "0.0%"
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=round(kernel.duration_us, 3))
            ws.cell(row=row, column=3, value=pct)
            ws.cell(row=row, column=4, value=kernel.kernel_type.value)
            ws.cell(row=row, column=5, value=kernel.simplified_name)
            ws.cell(row=row, column=6, value=kernel.name)
            row += 1

        row += 1

        # Breakdown section
        ws.cell(row=row, column=1, value="=== BREAKDOWN ===").font = Font(bold=True)
        row += 1

        headers = ["Type", "Time (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        breakdown = {}
        for k in kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us

        for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
            pct = f"{100.0 * time_us / total_time:.1f}%" if total_time > 0 else "0.0%"
            ws.cell(row=row, column=1, value=kt.value)
            ws.cell(row=row, column=2, value=round(time_us, 3))
            ws.cell(row=row, column=3, value=pct)
            row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Limit kernel name column width
            if column == 'F':
                ws.column_dimensions[column].width = 80
            else:
                ws.column_dimensions[column].width = min(max_length + 2, 30)

    def _export_summary_csv(self, path: str) -> None:
        """Export summary statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)

            # Overall summary
            writer.writerow(["=== OVERALL SUMMARY ==="])
            writer.writerow(["metric", "value", "percentage"])
            total_time_us = self.result.total_time_us
            writer.writerow(["total_kernel_time_us", f"{total_time_us:.3f}", "100.0%"])

            for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
                stats = self.result.per_stage_stats.get(stage, KernelStats())
                pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
                writer.writerow([f"{stage.value}_time_us", f"{stats.total_time_us:.3f}", pct])

            writer.writerow([])

            # Type breakdown
            writer.writerow(["=== BREAKDOWN BY TYPE ==="])
            writer.writerow(["type", "count", "total_time_us", "avg_time_us", "percentage"])

            sorted_types = sorted(
                self.result.per_type_stats.items(),
                key=lambda x: x[1].total_time_us,
                reverse=True,
            )
            for kernel_type, stats in sorted_types:
                pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
                writer.writerow([
                    kernel_type.value,
                    stats.count,
                    f"{stats.total_time_us:.3f}",
                    f"{stats.avg_time_us:.3f}",
                    pct,
                ])

            writer.writerow([])

            # Layer summary
            if self.result.layers:
                writer.writerow(["=== LAYER SUMMARY ==="])
                writer.writerow([
                    "layer_idx", "layer_type", "stage", "total_time_us",
                    "attention_time_us", "moe_time_us", "linear_time_us",
                    "communication_time_us", "quantization_time_us", "kernel_count"
                ])

                for layer in self.result.layers:
                    writer.writerow([
                        layer.layer_idx,
                        layer.layer_type.value,
                        layer.stage.value,
                        f"{layer.total_time_us:.3f}",
                        f"{layer.attention_time_us:.3f}",
                        f"{layer.moe_time_us:.3f}",
                        f"{layer.linear_time_us:.3f}",
                        f"{layer.communication_time_us:.3f}",
                        f"{layer.quantization_time_us:.3f}",
                        len(layer.kernels),
                    ])

        logger.info(f"Summary exported to: {path}")

    def _export_layer_detail_csv(self, layer: LayerEvent, path: str) -> None:
        """Export detailed kernel list for a single layer to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)

            # Layer header
            writer.writerow([f"=== LAYER {layer.layer_idx} ({layer.layer_type.value}, {layer.stage.value}) ==="])
            writer.writerow([])

            # Filter out trailing ALLREDUCE
            kernels = layer.kernels
            if kernels and kernels[-1].simplified_name == "ALLREDUCE":
                kernels = kernels[:-1]

            total_time = sum(k.duration_us for k in kernels)

            # Kernel details
            writer.writerow(["index", "duration_us", "percentage", "type", "short_name", "kernel_name"])

            for i, kernel in enumerate(kernels, 1):
                pct = f"{100.0 * kernel.duration_us / total_time:.1f}%" if total_time > 0 else "0.0%"
                writer.writerow([
                    i,
                    f"{kernel.duration_us:.3f}",
                    pct,
                    kernel.kernel_type.value,
                    kernel.simplified_name,
                    kernel.name,
                ])

            writer.writerow([])

            # Layer total and breakdown
            writer.writerow(["=== LAYER TOTAL ==="])
            writer.writerow(["total_time_us", f"{total_time:.3f}"])
            writer.writerow([])

            writer.writerow(["=== BREAKDOWN ==="])
            writer.writerow(["type", "time_us", "percentage"])

            breakdown = {}
            for k in kernels:
                if k.kernel_type not in breakdown:
                    breakdown[k.kernel_type] = 0.0
                breakdown[k.kernel_type] += k.duration_us

            for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
                pct = f"{100.0 * time_us / total_time:.1f}%" if total_time > 0 else "0.0%"
                writer.writerow([kt.value, f"{time_us:.3f}", pct])


def parse_layer_indices(layer_spec: str) -> Optional[List[int]]:
    """Parse layer specification string into list of indices.

    Supports formats:
    - "4,10,20" -> [4, 10, 20]
    - "4-10" -> [4, 5, 6, 7, 8, 9, 10]
    - "4-10,20,30-32" -> [4, 5, ..., 10, 20, 30, 31, 32]
    """
    if layer_spec is None or not layer_spec.strip():
        return []

    indices = []
    parts = layer_spec.split(",")
    for part in parts:
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            indices.extend(range(int(start), int(end) + 1))
        else:
            indices.append(int(part))

    return sorted(set(indices))


def main():
    parser = argparse.ArgumentParser(
        description="Analyze torch profiler trace files for kernel execution times.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python -m sglang.srt.utils.trace_analyzer /path/to/trace.json.gz

  # Show top 30 kernels, decode stage only
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --top-n 30 --stage decode

  # Layer-level analysis (terminal output)
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --show-layer-terminal

  # Show specific layer detail on terminal
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --layer-report 5 --show-layer-terminal

  # Export full analysis to Excel file (summary + layer details as sheets)
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 4,10,20

  # Export layers 0-10 for prefill stage
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 0-10 --stage prefill

  # Export to Excel without terminal output
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 4,10

  # Debug specific layers
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --debug-layers 4,10
""",
    )

    parser.add_argument("trace_file", help="Path to trace file (.json.gz or .json)")
    parser.add_argument(
        "--top-n",
        type=int,
        default=20,
        help="Number of top kernels to show (default: 20)",
    )
    parser.add_argument(
        "--stage",
        choices=["prefill", "decode", "all"],
        default="all",
        help="Filter by stage (default: all)",
    )
    parser.add_argument(
        "--layer-analysis",
        action="store_true",
        help="Deprecated: layer detection is enabled by default",
    )
    parser.add_argument(
        "--layer-report", type=int, metavar="N", help="Show detailed report for layer N on terminal"
    )
    parser.add_argument(
        "--show-layer-terminal",
        action="store_true",
        help="Show layer analysis on terminal (default: False)",
    )
    parser.add_argument(
        "--export-csv",
        metavar="FILENAME",
        help="Export analysis to Excel file (.xlsx) with multiple sheets (Summary + Layer_N sheets)",
    )
    parser.add_argument(
        "--export-layers",
        metavar="LAYERS",
        help="Layers to include in CSV export: '4,10,20' or '4-10' (required with --export-csv)",
        default=None,
    )
    parser.add_argument(
        "--debug-layers",
        metavar="LAYERS",
        help="Print debug info for specific layers: '4,10,20' or '4-10'",
        default=None,
    )
    parser.add_argument(
        "--debug-top-n",
        type=int,
        default=20,
        help="Number of kernels to show in layer debug (default: 20)",
    )
    parser.add_argument(
        "--csv-stats", metavar="FILE", help="Export kernel stats to CSV file"
    )
    parser.add_argument(
        "--csv-events", metavar="FILE", help="Export all events to CSV file"
    )
    parser.add_argument(
        "--csv-layers", metavar="FILE", help="Export layer stats to CSV file"
    )
    parser.add_argument(
        "--grid-threshold",
        type=int,
        default=10000,
        help="Grid[0] threshold for decode heuristic (default: 10000)",
    )
    parser.add_argument(
        "--mtp-qseqlen-decode",
        action="store_true",
        help="Treat attention qseqlen>=2 kernels as decode for layer stage inference",
    )
    parser.add_argument(
        "--split-half-layers",
        action="store_true",
        help="Keep attention and MoE/FC half-layers separate instead of "
        "merging them into single full layers (default: merged)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Enable verbose logging"
    )

    args = parser.parse_args()

    # Configure logging
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )

    # Run analysis
    analyzer = TraceAnalyzer(
        grid_threshold=args.grid_threshold,
        mtp_qseqlen_decode=args.mtp_qseqlen_decode,
        split_half_layers=args.split_half_layers,
    )

    try:
        trace_data = analyzer.load_trace(args.trace_file)
    except FileNotFoundError:
        logger.error(f"Trace file not found: {args.trace_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in trace file: {e}")
        sys.exit(1)

    # Detect layers by default for consistent stage stats
    detect_layers = True

    result = analyzer.analyze(trace_data, detect_layers=detect_layers)

    # Determine stage filter
    stage_enum = None
    if args.stage == "prefill":
        stage_enum = Stage.PREFILL
    elif args.stage == "decode":
        stage_enum = Stage.DECODE

    # Print report
    formatter = ReportFormatter(result)
    formatter.print_full_report(top_n=args.top_n, stage_filter=args.stage)

    # Auto-show first 2 prefill + first 2 decode layer details
    if detect_layers:
        formatter.print_auto_layer_details(num_layers=5)

    # Print layer analysis on terminal only if --show-layer-terminal is set
    if detect_layers and args.show_layer_terminal:
        formatter.print_layer_summary(stage=stage_enum)

        if args.layer_report is not None:
            formatter.print_layer_detail(args.layer_report, stage=stage_enum)

    if detect_layers and args.debug_layers:
        debug_layers = parse_layer_indices(args.debug_layers)
        if not debug_layers:
            logger.error("--debug-layers must include at least one layer number.")
            sys.exit(1)
        for layer_idx in debug_layers:
            formatter.print_layer_debug(layer_idx, top_n=args.debug_top_n)

    # Export CSVs if requested
    exporter = CSVExporter(result)

    # Export full analysis to Excel file
    if args.export_csv:
        layer_indices = parse_layer_indices(args.export_layers)
        if not layer_indices:
            logger.error("--export-layers is required and must include at least one layer number.")
            sys.exit(1)

        # Save Excel file in the same folder as the input trace file
        trace_dir = os.path.dirname(os.path.abspath(args.trace_file))
        output_filename = args.export_csv if args.export_csv.endswith('.xlsx') else f"{args.export_csv}.xlsx"
        output_path = os.path.join(trace_dir, output_filename)

        try:
            created_file = exporter.export_full_analysis(
                output_path=output_path,
                layer_indices=layer_indices,
                stage_filter=stage_enum,
            )
            print(f"\nExported Excel file: {created_file}")
        except ImportError as e:
            logger.error(str(e))
            sys.exit(1)

    # Legacy CSV exports
    if args.csv_stats:
        exporter.export_kernel_stats(args.csv_stats)
    if args.csv_events:
        exporter.export_events(args.csv_events)
    if args.csv_layers:
        exporter.export_layers(args.csv_layers)


if __name__ == "__main__":
    main()
