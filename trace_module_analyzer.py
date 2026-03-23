#!/usr/bin/env python3
"""Trace Module Analyzer — Correlation-based GPU kernel classification.

Uses nn.Module hierarchy from PyTorch traces to automatically classify GPU
kernels by module, eliminating the need for hardcoded regex patterns.

Supports two trace modes:
  - Full mode: kernel events present (diffusion traces) — correlates via cuda_runtime
  - CPU-only mode: no kernel events (LLM traces) — uses cpu_op time containment

Pipeline:
  ModuleTreeBuilder     — Parse python_function events into an nn.Module hierarchy tree.
  KernelCorrelator      — Map GPU kernels to modules via cuda_runtime correlation IDs.
  CudaGraphCorrelator   — Handle CUDA-graph-replayed kernels by mapping them to
                           synthetic layer modules.
  CpuOpCorrelator       — Map cpu_ops to modules via timestamp containment (CPU-only mode).
  ModuleAggregator      — Roll up per-module statistics (time, counts, breakdowns).
  PhaseDetector         — Detect prefill vs decode phases in LLM traces.
  ReportGenerator       — Produce console summaries and Excel reports.
  TraceModuleAnalyzer   — Top-level orchestrator that chains all the above.

PhaseDetector (prefill vs decode):
  Three-tier approach — use explicit function-call markers when available,
  fall back to CUDA graph replay detection, then to keyword heuristics.

  1. Phase markers: During trace loading, scan python_function events for
     SGLang's ModelRunner dispatch functions:
       - "model_runner.py(...): forward_extend"  → prefill (duration span)
       - "model_runner.py(...): forward_decode"  → decode  (duration span)
     Each marker records (start_ts, end_ts, phase, tid, pid).

     NOTE: We do NOT use forward_batch_info "is_extend"/"is_decode" events.
     Those are boolean property checks on ForwardMode that fire during both
     prefill and decode paths (e.g. inside init_new, attention backend init),
     making them unreliable as phase indicators.

  2. Tagging modules: For each module node, check whether it falls within
     any forward_extend or forward_decode time span. Children inherit their
     parent's phase since a single forward pass is entirely prefill or decode.

  3. CUDA graph replay: CudaGraphReplay roots are always decode (hardcoded
     at construction time in CudaGraphReplayHandler).

  4. Fallback (no markers): If the trace has no model_runner markers
     (e.g. non-SGLang traces), scan each module's cpu_op event names and
     majority-vote on "prefill"/"extend" vs "decode" keywords.

  5. Propagation: _propagate_phase pushes assigned phases down to any untagged
     child nodes, ensuring every node in the tree has a phase label.
"""

import argparse
import bisect
import csv
import gzip
import json
import logging
import math
import operator
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    from fix_rocm_trace_flow import fix_trace as _rocm_fix_trace
    _HAS_ROCM_FIX = True
except ImportError:
    _HAS_ROCM_FIX = False

# Maximum data rows per Excel tab (excluding header).  Keeps file size
# manageable for large traces (e.g. DeepSeek with 24K+ decoder instances).
MAX_ROWS_PER_TAB = 1000


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ModuleNode:
    """Tree node representing an nn.Module invocation."""
    name: str           # e.g. "WanTransformerBlock_5"
    module_type: str    # e.g. "WanTransformerBlock"
    instance_id: int    # e.g. 5
    ts: float           # start timestamp (us)
    end: float          # ts + dur
    tid: int
    pid: int
    children: List["ModuleNode"] = field(default_factory=list)
    kernels: List[Dict] = field(default_factory=list)
    cpu_ops: List[Dict] = field(default_factory=list)


class KernelDetail:
    """Individual kernel/op record for detail reporting.

    Uses __slots__ for fast instantiation (critical when creating millions
    of instances for large traces).
    """
    __slots__ = ("name", "duration", "category", "module_path",
                 "ts", "phase", "input_dims", "source_path")

    def __init__(self, name: str, duration: float, category: str,
                 module_path: str, ts: float = 0.0, phase: str = "",
                 input_dims: str = "", source_path: str = ""):
        self.name = name
        self.duration = duration
        self.category = category
        self.module_path = module_path
        self.ts = ts
        self.phase = phase
        self.input_dims = input_dims
        self.source_path = source_path


@dataclass
class ModuleStats:
    """Aggregated statistics for a module node."""
    name: str
    module_type: str
    instance_id: int
    depth: int
    total_kernel_time: float = 0.0
    total_cpu_op_time: float = 0.0
    self_kernel_time: float = 0.0
    self_cpu_op_time: float = 0.0
    kernel_count: int = 0
    cpu_op_count: int = 0
    kernel_breakdown: Dict[str, Tuple[float, int]] = field(default_factory=dict)  # category -> (dur, count)
    kernel_details: List[KernelDetail] = field(default_factory=list)  # all kernels/ops in time order
    children_stats: List["ModuleStats"] = field(default_factory=list)
    phase: str = ""  # "prefill" / "decode" / ""


# ---------------------------------------------------------------------------
# Module tree builder
# ---------------------------------------------------------------------------

class ModuleTreeBuilder:
    """Build nn.Module hierarchy tree from python_function events."""

    MODULE_PREFIX = "nn.Module: "

    def build(self, events: List[Dict]) -> List[ModuleNode]:
        """Build forest from all events (filters for nn.Module internally)."""
        module_events = [
            e for e in events
            if e.get("cat") == "python_function"
            and str(e.get("name", "")).startswith(self.MODULE_PREFIX)
            and e.get("dur") is not None
        ]
        return self.build_from_module_events(module_events)

    def build_from_module_events(self, module_events: List[Dict]) -> List[ModuleNode]:
        """Build forest from pre-filtered nn.Module events (faster for large traces)."""
        if not module_events:
            return []

        # Group by (pid, tid)
        by_thread: Dict[Tuple[int, int], List[Dict]] = defaultdict(list)
        for e in module_events:
            by_thread[(e["pid"], e["tid"])].append(e)

        roots = []
        for (pid, tid), thread_events in by_thread.items():
            thread_events.sort(key=lambda x: (x["ts"], -x.get("dur", 0)))
            thread_roots = self._build_thread(thread_events, pid, tid)
            roots.extend(thread_roots)

        return roots

    def _build_thread(self, events: List[Dict], pid: int, tid: int) -> List[ModuleNode]:
        """Stack-based nesting for events on a single thread."""
        roots = []
        stack: List[ModuleNode] = []

        for e in events:
            ts = e["ts"]
            dur = e.get("dur", 0)
            end = ts + dur
            raw_name = e["name"][len(self.MODULE_PREFIX):]
            module_type, instance_id = self._parse_name(raw_name)

            node = ModuleNode(
                name=raw_name,
                module_type=module_type,
                instance_id=instance_id,
                ts=ts, end=end,
                tid=tid, pid=pid,
            )

            # Pop finished ancestors
            while stack and stack[-1].end <= ts:
                stack.pop()

            if stack:
                stack[-1].children.append(node)
            else:
                roots.append(node)

            stack.append(node)

        return roots

    @staticmethod
    def _parse_name(raw_name: str) -> Tuple[str, int]:
        m = re.match(r"^(.+?)_(\d+)$", raw_name)
        if m:
            return m.group(1), int(m.group(2))
        return raw_name, 0


# ---------------------------------------------------------------------------
# Kernel correlator (full mode: kernel → cuda_runtime → module)
# ---------------------------------------------------------------------------

class _IntervalIndex:
    """Shared interval index for O(log n) module lookup by timestamp."""

    def __init__(self, roots: List[ModuleNode]):
        self._intervals: Dict[Tuple[int, int], List[Tuple[float, float, ModuleNode]]] = defaultdict(list)
        self._starts_cache: Dict[Tuple[int, int], List[float]] = {}
        self._collect_all_nodes(roots)
        for key in self._intervals:
            self._intervals[key].sort(key=lambda x: (x[0], -x[1]))
            self._starts_cache[key] = [iv[0] for iv in self._intervals[key]]

    def _collect_all_nodes(self, nodes: List[ModuleNode]):
        for node in nodes:
            self._intervals[(node.pid, node.tid)].append((node.ts, node.end, node))
            self._collect_all_nodes(node.children)

    def find_deepest(self, ts: float, pid: int, tid: int) -> Optional[ModuleNode]:
        key = (pid, tid)
        intervals = self._intervals.get(key)
        if not intervals:
            return None
        starts = self._starts_cache[key]
        idx = bisect.bisect_right(starts, ts) - 1
        best: Optional[ModuleNode] = None
        best_span = float("inf")
        for i in range(max(0, idx - 20), min(len(intervals), idx + 20)):
            s, e, node = intervals[i]
            if s <= ts <= e:
                span = e - s
                if span < best_span:
                    best_span = span
                    best = node
            elif s > ts:
                break
        return best


class CpuOpShapeIndex:
    """Map correlation IDs to cpu_op tensor shapes via timestamp containment.

    For each cuda_runtime/cuda_driver launch event, finds the enclosing cpu_op
    on the same thread and extracts its ``Input Dims`` argument.
    """

    def __init__(self, cpu_ops: List[Dict], runtime_events: List[Dict],
                 driver_events: Optional[List[Dict]] = None):
        # Build interval index of cpu_ops that carry Input Dims
        shaped_ops: Dict[Tuple[int, int], List[Tuple[float, float, str]]] = defaultdict(list)
        for op in cpu_ops:
            dims = op.get("args", {}).get("Input Dims")
            if not dims:
                continue
            ts = op["ts"]
            dur = op.get("dur", 0)
            tid = op["tid"]
            pid = op.get("pid", tid)
            dims_str = self._format_dims(dims)
            if dims_str:
                shaped_ops[(pid, tid)].append((ts, ts + dur, dims_str))

        for key in shaped_ops:
            shaped_ops[key].sort(key=lambda x: (x[0], -x[1]))
        self._shaped_ops = shaped_ops
        self._starts_cache: Dict[Tuple[int, int], List[float]] = {
            k: [iv[0] for iv in v] for k, v in shaped_ops.items()
        }

        # Build correlation → shape mapping
        self._corr_to_shape: Dict[int, str] = {}
        all_launches = list(runtime_events)
        if driver_events:
            all_launches.extend(driver_events)
        for e in all_launches:
            corr = e.get("args", {}).get("correlation")
            if corr is None or corr in self._corr_to_shape:
                continue
            ts = e["ts"]
            tid = e["tid"]
            pid = e.get("pid", tid)
            shape = self._find_shape(ts, pid, tid)
            if shape:
                self._corr_to_shape[corr] = shape

    def get_shape(self, correlation_id: int) -> str:
        """Return Input Dims string for a kernel's correlation ID, or ''."""
        return self._corr_to_shape.get(correlation_id, "")

    def _find_shape(self, ts: float, pid: int, tid: int) -> str:
        key = (pid, tid)
        intervals = self._shaped_ops.get(key)
        if not intervals:
            return ""
        starts = self._starts_cache[key]
        idx = bisect.bisect_right(starts, ts) - 1
        # Search nearby intervals for the deepest (narrowest) enclosing cpu_op
        best = ""
        best_span = float("inf")
        for i in range(max(0, idx - 5), min(len(intervals), idx + 5)):
            s, e, dims_str = intervals[i]
            if s <= ts <= e:
                span = e - s
                if span < best_span:
                    best_span = span
                    best = dims_str
            elif s > ts + 100:
                break
        return best

    @staticmethod
    def _format_dims(dims) -> str:
        """Format Input Dims into a concise string, omitting empty entries."""
        if not isinstance(dims, list):
            return ""
        non_empty = [d for d in dims if isinstance(d, list) and len(d) > 0]
        if not non_empty:
            return ""
        return str(non_empty)


class PythonSourceIndex:
    """Map correlation IDs to the Python source location of each kernel launch.

    For each cuda_runtime/cuda_driver launch event, finds the narrowest
    enclosing python_function event whose name contains a ``.py(`` source
    reference (e.g. ``sglang/srt/layers/attention/fla/chunk.py(26):
    chunk_gated_delta_rule_fwd``).

    Falls back to a cleaned-up ``<built-in method NAME ...>`` if no ``.py``
    source is found in the enclosing stack.
    """

    _BUILTIN_RE = re.compile(
        r"<built-in (?:method|function) (\S+)")
    _FRAMEWORK_PREFIXES = (
        "torch/", "threading.py", "multiprocessing/",
        "<string>", "tqdm/", "importlib/", "contextlib.py",
    )
    _WRAPPER_RE = re.compile(
        r":\s*(?:wrapper|custom_wrapper|outer_wrapper|wrapper_custom"
        r"|__call__|_call_impl|<lambda>|<module>|<genexpr>"
        r"|decorate_context|decorate_fwd|run)\s*$"
    )

    def __init__(self, pyfunc_events: List[Dict], runtime_events: List[Dict],
                 driver_events: Optional[List[Dict]] = None):
        pf_intervals: Dict[Tuple[int, int], List[Tuple[float, float, str]]] = defaultdict(list)
        for e in pyfunc_events:
            dur = e.get("dur")
            if dur is None or dur < 1.0:
                continue
            ts = e["ts"]
            tid = e["tid"]
            pid = e.get("pid", tid)
            pf_intervals[(pid, tid)].append((ts, ts + dur, e.get("name", "")))

        for key in pf_intervals:
            pf_intervals[key].sort(key=lambda x: (x[0], -x[1]))
        self._intervals = pf_intervals
        self._starts_cache: Dict[Tuple[int, int], List[float]] = {
            k: [iv[0] for iv in v] for k, v in pf_intervals.items()
        }

        self._corr_to_source: Dict[int, str] = {}
        all_launches = list(runtime_events)
        if driver_events:
            all_launches.extend(driver_events)
        for e in all_launches:
            corr = e.get("args", {}).get("correlation")
            if corr is None or corr in self._corr_to_source:
                continue
            ts = e["ts"]
            tid = e["tid"]
            pid = e.get("pid", tid)
            src = self._find_source(ts, pid, tid)
            if src:
                self._corr_to_source[corr] = src

    def get_source(self, correlation_id: int) -> str:
        return self._corr_to_source.get(correlation_id, "")

    def _find_source(self, ts: float, pid: int, tid: int) -> str:
        """Find the best Python source location enclosing *ts*.

        Four-tier preference (narrowest wins within each tier):
          1. App-level ``.py(`` frames — not framework, not wrapper/dispatch
          2. App-level ``.py(`` frames — not framework (wrapper allowed)
          3. Any ``.py(`` frame (including framework internals)
          4. ``<built-in method/function NAME ...>`` with pybind11 noise stripped
        """
        key = (pid, tid)
        intervals = self._intervals.get(key)
        if not intervals:
            return ""
        starts = self._starts_cache[key]
        idx = bisect.bisect_right(starts, ts) - 1

        best_app = ""
        best_app_span = float("inf")
        best_app_wrap = ""
        best_app_wrap_span = float("inf")
        best_any_py = ""
        best_any_py_span = float("inf")
        best_fallback = ""
        best_fb_span = float("inf")

        fw_prefixes = self._FRAMEWORK_PREFIXES
        wrapper_re = self._WRAPPER_RE

        lo = max(0, idx - 80)
        hi = min(len(intervals), idx + 10)
        for i in range(lo, hi):
            s, e, name = intervals[i]
            if s <= ts <= e:
                span = e - s
                if ".py(" in name:
                    is_framework = name.startswith(fw_prefixes)
                    is_wrapper = bool(wrapper_re.search(name))
                    if not is_framework:
                        if not is_wrapper and span < best_app_span:
                            best_app_span = span
                            best_app = name
                        if span < best_app_wrap_span:
                            best_app_wrap_span = span
                            best_app_wrap = name
                    if span < best_any_py_span:
                        best_any_py_span = span
                        best_any_py = name
                elif span < best_fb_span:
                    best_fb_span = span
                    best_fallback = name
            elif s > ts + 100:
                break

        if best_app:
            return best_app
        if best_app_wrap:
            return best_app_wrap
        if best_any_py:
            return best_any_py
        if best_fallback:
            m = self._BUILTIN_RE.match(best_fallback)
            if m:
                return f"<built-in {m.group(1)}>"
            return best_fallback
        return ""


class KernelCorrelator:
    """Map GPU kernels to modules via correlation ID chain."""

    def __init__(self, runtime_events: List[Dict], roots: List[ModuleNode],
                 driver_events: Optional[List[Dict]] = None):
        # Build correlation → launch event mapping from both cuda_runtime
        # and cuda_driver events.  On B200, many kernels are launched via
        # cuLaunchKernelEx (cuda_driver) instead of cuda_runtime, so both
        # sources are needed for complete correlation coverage.
        self._corr_to_rt: Dict[int, Dict] = {}
        for e in runtime_events:
            corr = e.get("args", {}).get("correlation")
            if corr is not None:
                self._corr_to_rt[corr] = e
        if driver_events:
            for e in driver_events:
                corr = e.get("args", {}).get("correlation")
                if corr is not None and corr not in self._corr_to_rt:
                    self._corr_to_rt[corr] = e

        self._index = _IntervalIndex(roots)

    def correlate(self, kernel_events: List[Dict], roots: List[ModuleNode],
                  shape_index: Optional["CpuOpShapeIndex"] = None,
                  source_index: Optional["PythonSourceIndex"] = None) -> int:
        """Assign each kernel to its deepest enclosing module. Returns count matched."""
        matched = 0
        for k in kernel_events:
            corr = k.get("args", {}).get("correlation")
            if corr is None:
                continue
            rt = self._corr_to_rt.get(corr)
            if rt is None:
                continue
            cpu_ts = rt["ts"]
            cpu_tid = rt["tid"]
            cpu_pid = rt.get("pid", cpu_tid)
            module = self._index.find_deepest(cpu_ts, cpu_pid, cpu_tid)
            if module is not None:
                if shape_index is not None:
                    k["_input_dims"] = shape_index.get_shape(corr)
                if source_index is not None:
                    k["_source_path"] = source_index.get_source(corr)
                k["_matched"] = True
                module.kernels.append(k)
                matched += 1
        return matched


# ---------------------------------------------------------------------------
# CUDA graph replay correlator
# ---------------------------------------------------------------------------

class CudaGraphCorrelator:
    """Correlate CUDA-graph-replayed kernels to synthetic layer modules."""

    _COMM_RE = re.compile(
        r"all_reduce|cross_device_reduce|nccl|rccl|broadcast|allgather"
        r"|reduce_scatter|quickreduce|all_to_all", re.IGNORECASE)
    _ATTN_RE = re.compile(
        r"aiter::mla_|mla_a8w8|decode_attention|flash_attn|attention|softmax"
        r"|fmha_|mla_reduce|kv_cache|paged_attention|chunk_gated_delta_rule"
        r"|fused_gdn|kn_get_mla_metadata|kn_mla_reduce|gating_delta_rule",
        re.IGNORECASE)
    _MOE_RE = re.compile(
        r"fused_moe|moe_align|topk|expert|MoeFlatmm|MoeSorting|kernel_moe_gemm"
        r"|kernel_moe_mxgemm|shared_experts|grouped_topk|fused_append_shared_experts",
        re.IGNORECASE)

    def __init__(self, runtime_events):
        self._graph_corrs = set()
        for e in runtime_events:
            name = e.get("name", "")
            if "hipGraphLaunch" in name or "cudaGraphLaunch" in name:
                corr = e.get("args", {}).get("correlation")
                if corr is not None:
                    self._graph_corrs.add(corr)

    @property
    def has_graph_replays(self):
        return bool(self._graph_corrs)

    def correlate(self, gpu_events, capture_roots):
        """Split graph-replayed events into synthetic layer modules.

        Returns (new_roots, matched_count).
        """
        # 1. Partition: graph-replay events grouped by correlation ID
        corr_to_events = defaultdict(list)
        for e in gpu_events:
            corr = e.get("args", {}).get("correlation")
            if corr in self._graph_corrs:
                corr_to_events[corr].append(e)

        if not corr_to_events:
            return [], 0

        # Sort each group by timestamp
        for corr in corr_to_events:
            corr_to_events[corr].sort(key=lambda e: e.get("ts", 0))

        # 2. Extract capture-iteration layer names for naming (needed before
        # template detection so we can skip half-layer merging when the
        # capture iteration already has distinct half-layer module types).
        capture_layer_names = self._extract_layer_names(capture_roots)
        has_distinct_half_layers = self._has_distinct_half_layer_types(
            capture_layer_names)

        # 3. Deduplicate into templates by kernel count (signature)
        templates = {}  # sig_len -> layer boundaries [(start, end, label), ...]
        for corr, evts in corr_to_events.items():
            sig_len = len(evts)
            if sig_len not in templates:
                names = [e.get("name", "") for e in evts]
                templates[sig_len] = self._detect_layers(
                    names, skip_merge=has_distinct_half_layers)

        # 3b. Classify each template as "target" or "draft" based on layer count.
        # The target model's CUDA graph has many layers (e.g. 61 for DeepSeek V3);
        # the MTP draft model's graph has very few (e.g. 1).
        max_layers = max(len(b) for b in templates.values()) if templates else 0
        template_is_target = {}
        for sig_len, bounds in templates.items():
            template_is_target[sig_len] = len(bounds) >= max(max_layers // 2, 2)

        has_both = (any(template_is_target.values())
                    and not all(template_is_target.values()))

        # 4. Build synthetic module trees for each replay
        new_roots = []
        matched = 0
        target_idx = 0
        draft_idx = 0
        for corr in sorted(corr_to_events,
                           key=lambda c: corr_to_events[c][0].get("ts", 0)):
            evts = corr_to_events[corr]
            sig_len = len(evts)
            layer_bounds = templates.get(sig_len, [(0, sig_len, "unknown")])
            is_target = template_is_target.get(sig_len, True)

            if has_both:
                if is_target:
                    root_name = f"CudaGraphReplay_Target_{target_idx}"
                    root_type = "CudaGraphReplay_Target"
                    root_id = target_idx
                    target_idx += 1
                else:
                    root_name = f"CudaGraphReplay_Draft_{draft_idx}"
                    root_type = "CudaGraphReplay_Draft"
                    root_id = draft_idx
                    draft_idx += 1
            else:
                root_name = f"CudaGraphReplay_{target_idx + draft_idx}"
                root_type = "CudaGraphReplay"
                root_id = target_idx + draft_idx
                target_idx += 1

            root = ModuleNode(
                name=root_name,
                module_type=root_type,
                instance_id=root_id,
                ts=evts[0].get("ts", 0),
                end=evts[-1].get("ts", 0) + evts[-1].get("dur", 0),
                tid=evts[0].get("tid", 0),
                pid=evts[0].get("pid", 0),
            )
            root._phase = "decode"

            for layer_i, (start, end, label) in enumerate(layer_bounds):
                if is_target and layer_i < len(capture_layer_names):
                    layer_name = capture_layer_names[layer_i]
                else:
                    layer_name = f"Layer_{layer_i}"
                mod_type, inst_id = ModuleTreeBuilder._parse_name(layer_name)

                layer_evts = evts[start:end]
                if not layer_evts:
                    continue

                layer_node = ModuleNode(
                    name=layer_name,
                    module_type=mod_type,
                    instance_id=inst_id,
                    ts=layer_evts[0].get("ts", 0),
                    end=layer_evts[-1].get("ts", 0) + layer_evts[-1].get("dur", 0),
                    tid=root.tid,
                    pid=root.pid,
                )
                layer_node._phase = "decode"
                layer_node.kernels = layer_evts
                matched += len(layer_evts)
                root.children.append(layer_node)

            new_roots.append(root)

        if has_both:
            print(f"  CUDA graph sub-types: {target_idx} target, {draft_idx} draft")

        return new_roots, matched

    def _detect_layers(self, names, skip_merge=False):
        """COMM-based segmentation + half-layer merging.

        Returns [(start_idx, end_idx, type_label), ...].

        skip_merge: when True, never merge half-layers.  Used when the
        capture iteration already has distinct module types for each half
        (e.g. Qwen3_5AttentionDecoderLayer + Qwen3_5LinearDecoderLayer).
        """
        cats = [self._quick_cat(n) for n in names]
        comm_pos = [i for i, c in enumerate(cats) if c == "COMM"]

        if len(comm_pos) < 2:
            return [(0, len(names), "unknown")]

        # Build segments ending at each COMM
        segments = []
        prev = 0
        for cp in comm_pos:
            seg = cats[prev:cp + 1]
            segments.append((prev, cp + 1, "ATTN" in seg, "MOE" in seg))
            prev = cp + 1
        if prev < len(cats):
            seg = cats[prev:]
            segments.append((prev, len(cats), "ATTN" in seg, "MOE" in seg))

        if skip_merge:
            return [(s[0], s[1], self._seg_label(s[2], s[3])) for s in segments]

        # Decide: half-layer vs full-layer model
        attn_only = sum(1 for s in segments if s[2] and not s[3])
        full_layer = sum(1 for s in segments if s[2] and s[3])

        if attn_only > full_layer:
            return self._merge_half_layers(segments)
        else:
            return [(s[0], s[1], self._seg_label(s[2], s[3])) for s in segments]

    @staticmethod
    def _merge_half_layers(segments):
        """Merge adjacent ATTN-only + non-ATTN pairs into full layers."""
        layers = []
        i = 0
        while i < len(segments):
            start, end, has_attn, has_moe = segments[i]
            if has_attn and not has_moe and i + 1 < len(segments):
                _, next_end, _, next_moe = segments[i + 1]
                layers.append((start, next_end, "MoE" if next_moe else "FC"))
                i += 2
            else:
                layers.append((start, end,
                               CudaGraphCorrelator._seg_label(has_attn, has_moe)))
                i += 1
        return layers

    @staticmethod
    def _seg_label(has_attn, has_moe):
        if has_attn and has_moe:
            return "MoE"
        if has_attn:
            return "Attn"
        if has_moe:
            return "MoE"
        return "other"

    def _quick_cat(self, name):
        if self._COMM_RE.search(name):
            return "COMM"
        if self._ATTN_RE.search(name):
            return "ATTN"
        if self._MOE_RE.search(name):
            return "MOE"
        return "X"

    @staticmethod
    def _extract_layer_names(roots):
        """Extract decoder layer names from capture-iteration module tree."""
        names = []
        for root in roots:
            for child in root.children:
                if ("DecoderLayer" in child.module_type
                        or "TransformerBlock" in child.module_type):
                    names.append(child.name)
        return names

    @staticmethod
    def _has_distinct_half_layer_types(layer_names):
        """Check if capture layer names contain multiple distinct DecoderLayer types.

        Models like Qwen3.5 split each transformer layer into two separate
        nn.Module types (AttentionDecoderLayer + LinearDecoderLayer).  When
        this is the case, the CUDA graph segmentation should NOT merge
        adjacent half-layers, because each half already has its own module
        identity from the capture iteration.
        """
        types = set()
        for name in layer_names:
            m = re.match(r"^(.+?)_(\d+)$", name)
            types.add(m.group(1) if m else name)
        return len(types) > 1


# ---------------------------------------------------------------------------
# CPU-op correlator (CPU-only mode: cpu_op → module via time containment)
# ---------------------------------------------------------------------------

class CpuOpCorrelator:
    """Map cpu_ops to modules via time containment on the same thread."""

    def __init__(self, roots: List[ModuleNode]):
        self._index = _IntervalIndex(roots)

    def correlate(self, cpu_ops: List[Dict], roots: List[ModuleNode]) -> int:
        """Assign cpu_ops to deepest enclosing module. Returns count matched."""
        matched = 0
        find = self._index.find_deepest  # avoid attribute lookup in hot loop
        for op in cpu_ops:
            dur = op.get("dur", 0)
            if dur < 1.0:  # skip trivial ops under 1 us
                continue
            ts = op["ts"]
            tid = op["tid"]
            pid = op.get("pid", tid)
            module = find(ts, pid, tid)
            if module is not None:
                module.cpu_ops.append(op)
                matched += 1
        return matched


# ---------------------------------------------------------------------------
# Module aggregator
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Kernel category classification — loaded from kernel_categories.csv
# ---------------------------------------------------------------------------

_DEFAULT_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kernel_categories.csv")


def _load_kernel_categories(csv_path: str = _DEFAULT_CSV) -> List[Tuple[str, re.Pattern]]:
    """Load (category, compiled_regex) list from a CSV file.

    The CSV must have columns: category, pattern
    Each pattern is a regex alternation (e.g. "nccl|rccl|all_reduce").
    Rows are matched top-to-bottom; first match wins.
    """
    if not os.path.isfile(csv_path):
        logger.warning("kernel_categories.csv not found at %s, no categories loaded", csv_path)
        return []
    categories = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cat = row["category"].strip()
            pat = row["pattern"].strip()
            categories.append((cat, re.compile(pat, re.IGNORECASE)))
    return categories


_KERNEL_CATEGORIES = _load_kernel_categories()


_categorize_cache: Dict[str, str] = {}


def _categorize_kernel(name: str) -> str:
    """Categorize a kernel name into a high-level category."""
    cached = _categorize_cache.get(name)
    if cached is not None:
        return cached
    for cat, pat in _KERNEL_CATEGORIES:
        if pat.search(name):
            _categorize_cache[name] = cat
            return cat
    _categorize_cache[name] = "other"
    return "other"


def _categorize_cpu_op(name: str) -> str:
    """Categorize a cpu_op name into a high-level category."""
    cached = _categorize_cache.get(name)
    if cached is not None:
        return cached
    for cat, pat in _KERNEL_CATEGORIES:
        if pat.search(name):
            _categorize_cache[name] = cat
            return cat
    _categorize_cache[name] = "other"
    return "other"


class ModuleAggregator:
    """Compute per-module statistics with hierarchical breakdown."""

    def aggregate(self, roots: List[ModuleNode], mode: str) -> List[ModuleStats]:
        """Recursively aggregate stats for all root nodes."""
        return [self._aggregate_node(r, 0, mode, "") for r in roots]

    def _aggregate_node(self, node: ModuleNode, depth: int, mode: str,
                        parent_path: str) -> ModuleStats:
        path = f"{parent_path}/{node.name}" if parent_path else node.name
        node_phase = getattr(node, "_phase", "")
        stats = ModuleStats(
            name=node.name,
            module_type=node.module_type,
            instance_id=node.instance_id,
            depth=depth,
        )

        # Direct kernel stats
        for k in node.kernels:
            dur = k.get("dur", 0)
            kname = k.get("name", "")
            stats.self_kernel_time += dur
            stats.kernel_count += 1
            cat = _categorize_kernel(kname)
            prev_dur, prev_cnt = stats.kernel_breakdown.get(cat, (0.0, 0))
            stats.kernel_breakdown[cat] = (prev_dur + dur, prev_cnt + 1)
            stats.kernel_details.append(KernelDetail(
                name=kname, duration=dur, category=cat,
                module_path=path, ts=k.get("ts", 0), phase=node_phase,
                input_dims=k.get("_input_dims", ""),
                source_path=k.get("_source_path", "")))

        # Direct cpu_op stats
        for op in node.cpu_ops:
            dur = op.get("dur", 0)
            opname = op.get("name", "")
            stats.self_cpu_op_time += dur
            stats.cpu_op_count += 1
            cat = _categorize_cpu_op(opname)
            prev_dur, prev_cnt = stats.kernel_breakdown.get(cat, (0.0, 0))
            stats.kernel_breakdown[cat] = (prev_dur + dur, prev_cnt + 1)
            dims = op.get("args", {}).get("Input Dims", "")
            if dims:
                dims = CpuOpShapeIndex._format_dims(dims)
            stats.kernel_details.append(KernelDetail(
                name=opname, duration=dur, category=cat,
                module_path=path, ts=op.get("ts", 0), phase=node_phase,
                input_dims=dims if dims else ""))

        # Sort own details by timestamp
        stats.kernel_details.sort(key=operator.attrgetter("ts"))

        # Recurse children
        stats.total_kernel_time = stats.self_kernel_time
        stats.total_cpu_op_time = stats.self_cpu_op_time
        for child in node.children:
            child_stats = self._aggregate_node(child, depth + 1, mode, path)
            stats.children_stats.append(child_stats)
            stats.total_kernel_time += child_stats.total_kernel_time
            stats.total_cpu_op_time += child_stats.total_cpu_op_time
            stats.kernel_count += child_stats.kernel_count
            stats.cpu_op_count += child_stats.cpu_op_count
            # Merge child breakdown into parent's total
            for cat, (dur, cnt) in child_stats.kernel_breakdown.items():
                prev_dur, prev_cnt = stats.kernel_breakdown.get(cat, (0.0, 0))
                stats.kernel_breakdown[cat] = (prev_dur + dur, prev_cnt + cnt)

        return stats


# ---------------------------------------------------------------------------
# Phase detector (prefill vs decode for LLM traces)
# ---------------------------------------------------------------------------

class PhaseDetector:
    """Detect prefill vs decode phase using ModelRunner dispatch function spans.

    Uses forward_extend / forward_decode time ranges from model_runner.py
    python_function events. Each marker is (start_ts, end_ts, phase, tid, pid).
    A module whose start time falls within a marker's [start, end] range is
    assigned that phase. CUDA graph replays are always decode (set elsewhere).
    """

    def detect_from_markers(self, roots: List[ModuleNode], phase_markers: List[Tuple]):
        """Tag modules using pre-extracted phase markers (start, end, phase, tid, pid)."""
        if not phase_markers:
            self._detect_from_cpu_ops(roots)
            return
        phase_markers_sorted = sorted(phase_markers, key=lambda m: m[0])
        self._tag_nodes_from_phases(roots, phase_markers_sorted)

    def _tag_nodes_from_phases(self, nodes: List[ModuleNode], phases: List[Tuple],
                               parent_phase: Optional[str] = None,
                               is_root: bool = True,
                               _phase_starts: Optional[List[float]] = None):
        if _phase_starts is None:
            _phase_starts = [p[0] for p in phases]
        for node in nodes:
            if parent_phase:
                node._phase = parent_phase  # type: ignore[attr-defined]
            elif not is_root:
                phase = self._find_enclosing_phase(node.ts, phases, _phase_starts)
                if phase:
                    node._phase = phase  # type: ignore[attr-defined]
            child_phase = getattr(node, "_phase", None) if not is_root else None
            self._tag_nodes_from_phases(
                node.children, phases,
                parent_phase=child_phase,
                is_root=False,
                _phase_starts=_phase_starts)

    @staticmethod
    def _find_enclosing_phase(ts: float, phases: List[Tuple],
                              phase_starts: List[float]) -> Optional[str]:
        """Find the phase marker whose [start, end] range contains *ts*."""
        idx = bisect.bisect_right(phase_starts, ts) - 1
        if idx < 0:
            return None
        start, end, phase, _, _ = phases[idx]
        if start <= ts <= end:
            return phase
        return None

    def _detect_from_cpu_ops(self, nodes: List[ModuleNode]):
        """Fall back to checking cpu_op names for prefill/decode keywords."""
        for node in nodes:
            prefill_count = sum(1 for op in node.cpu_ops if "prefill" in op.get("name", "").lower() or "extend" in op.get("name", "").lower())
            decode_count = sum(1 for op in node.cpu_ops if "decode" in op.get("name", "").lower())
            if prefill_count > decode_count:
                node._phase = "prefill"  # type: ignore[attr-defined]
            elif decode_count > prefill_count:
                node._phase = "decode"  # type: ignore[attr-defined]
            self._detect_from_cpu_ops(node.children)


# ---------------------------------------------------------------------------
# Model info text generator (from trace module roots, no re-load)
# ---------------------------------------------------------------------------

def _generate_model_info_html(xlsx_path: str,
                              output_html: Optional[str] = None,
                              serve: bool = False,
                              port: int = 8765
                              ) -> Optional[str]:
    """Generate interactive module tree HTML from an analysis.xlsx using visualize_module_tree.py.

    Returns the output HTML path on success, or None on failure.
    If serve=True, starts an HTTP server after generating the HTML.
    """
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return None

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        from visualize_module_tree import main as visualize_main
    except ImportError as e:
        logger.warning("visualize_module_tree not available: %s", e)
        return None

    if output_html is None:
        output_html = os.path.join(
            os.path.dirname(os.path.abspath(xlsx_path)) or ".",
            "module_tree.html")

    try:
        visualize_main(xlsx_path, output_html, serve=serve, port=port)
        if os.path.isfile(output_html):
            return output_html
    except Exception as e:
        logger.warning("Failed to generate module tree HTML: %s", e)

    return None


# ---------------------------------------------------------------------------
# Report generator
# ---------------------------------------------------------------------------

class ReportGenerator:
    """Generate console and Excel reports from module stats."""

    def print_layer_detail(self, stats_list: List[ModuleStats], mode: str,
                           detail_module: str, module_index: Optional[int] = None):
        """Print kernel-by-kernel detail for a specific module type (one representative instance).

        Shows the full sequence of kernels in time order — the repetitive pattern within a layer.
        """
        # Find matching module instances
        matches: List[ModuleStats] = []
        self._find_modules(stats_list, detail_module, matches)

        if not matches:
            print(f"\nNo modules matching '{detail_module}' found.")
            return

        # Pick the instance to show
        if module_index is not None:
            target = [m for m in matches if m.instance_id == module_index]
            if not target:
                print(f"\nInstance {module_index} not found for '{detail_module}'. "
                      f"Available: {sorted(set(m.instance_id for m in matches))}")
                return
            selected = target[0]
            pick_reason = "user-specified"
        else:
            # Pick instance closest to median (most representative)
            selected, pick_reason = self._pick_median_instance(matches, mode)

        # Collect all kernel details from this module and its descendants, in time order
        all_details = self._collect_all_details(selected)
        all_details.sort(key=lambda d: d.ts)

        time_val = selected.total_kernel_time if mode == "full" else selected.total_cpu_op_time

        # Compute stats across all instances for context
        all_times = [(m.total_kernel_time if mode == "full" else m.total_cpu_op_time)
                     for m in matches]
        mean_t, std_t, min_t, max_t, _ = ReportGenerator._time_stats(all_times)

        # Compute wall time (last end - first start) vs sum of durations
        sum_dur = sum(d.duration for d in all_details)
        if all_details:
            wall_start = min(d.ts for d in all_details)
            wall_end = max(d.ts + d.duration for d in all_details)
            wall_time = wall_end - wall_start
        else:
            wall_time = 0

        print(f"\n{'='*100}")
        print(f"  Layer Detail: {selected.name}  [{len(all_details)} items]")
        print(f"  Kernel sum: {sum_dur:,.0f} us  |  Wall time: {wall_time:,.0f} us  "
              f"(overlap: {sum_dur - wall_time:,.0f} us)")
        print(f"  Selected: {pick_reason}  |  "
              f"All {len(matches)} instances: mean={mean_t:,.0f} std={std_t:,.0f} "
              f"min={min_t:,.0f} max={max_t:,.0f}")
        phase = getattr(selected, 'phase', '')
        if phase:
            print(f"  Phase: {phase}")
        print(f"{'='*100}")
        has_dims = any(d.input_dims for d in all_details)
        if has_dims:
            print(f"  {'#':>4s}  {'Duration (us)':>13s}  {'% wall time':>6s}  {'Category':>15s}  {'Module':30s}  {'Kernel Name':50s}  Input Dims")
            print(f"  {'-'*4}  {'-'*13}  {'-'*6}  {'-'*15}  {'-'*30}  {'-'*50}  {'-'*40}")
        else:
            print(f"  {'#':>4s}  {'Duration (us)':>13s}  {'% wall time':>6s}  {'Category':>15s}  {'Module':30s}  Kernel Name")
            print(f"  {'-'*4}  {'-'*13}  {'-'*6}  {'-'*15}  {'-'*30}  {'-'*50}")

        for i, d in enumerate(all_details, 1):
            pct = d.duration / wall_time * 100 if wall_time > 0 else 0
            leaf = d.module_path.rsplit("/", 1)[-1] if "/" in d.module_path else d.module_path
            kname = d.name[:80]
            if has_dims:
                dims = d.input_dims[:60] if d.input_dims else ""
                print(f"  {i:4d}  {d.duration:13,.1f}  {pct:5.1f}%  {d.category:>15s}  {leaf:30s}  {kname:50s}  {dims}")
            else:
                print(f"  {i:4d}  {d.duration:13,.1f}  {pct:5.1f}%  {d.category:>15s}  {leaf:30s}  {kname}")

    def _pick_median_instance(self, matches: List[ModuleStats],
                              mode: str) -> Tuple[ModuleStats, str]:
        """Pick the instance closest to the median total time."""
        if len(matches) == 1:
            return matches[0], f"only instance (id={matches[0].instance_id})"

        timed = [(m.total_kernel_time if mode == "full" else m.total_cpu_op_time, m)
                 for m in matches]
        timed.sort(key=lambda x: x[0])
        mid = len(timed) // 2
        median_val = timed[mid][0]
        # Find closest to median
        best = min(timed, key=lambda x: abs(x[0] - median_val))
        return best[1], (f"closest to median (id={best[1].instance_id}, "
                         f"{best[0]:,.0f} us, median={median_val:,.0f} us)")

    def _find_modules(self, stats_list: List[ModuleStats], module_type: str,
                      out: List[ModuleStats]):
        for s in stats_list:
            if s.module_type == module_type:
                out.append(s)
            self._find_modules(s.children_stats, module_type, out)

    def _collect_all_details(self, stats: ModuleStats) -> List[KernelDetail]:
        """Collect kernel details from this node and all descendants."""
        result = list(stats.kernel_details)
        for child in stats.children_stats:
            result.extend(self._collect_all_details(child))
        return result

    def print_type_summary(self, stats_list: List[ModuleStats], mode: str,
                           max_detail_modules: int = 3,
                           detail_modules: Optional[List[str]] = None):
        """Print summary matching the Summary tab: metrics, categories, detail tabs."""
        grand_total = sum(
            (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
            for s in stats_list)

        # Build wrapper chains and type totals for hierarchy display
        wrapper_types: set = set()
        type_children_map: Dict[str, set] = {}
        top_type_names: set = set()
        root_chains: Dict[str, List[str]] = {}
        type_totals: Dict[str, float] = defaultdict(float)
        if not detail_modules:
            top_type_names, root_chains, type_totals, wrapper_types, type_children_map = \
                self._select_max_detail_modules(stats_list, mode, max_detail_modules)

        # --- Section 1: Kernel Time Summary (with hierarchy) ---
        print("\n" + "=" * 90)
        print("  Kernel Time Summary")
        print("=" * 90)
        print(f"  {'Metric':<45s} {'Value (us)':>14s} {'%':>7s}  {'Detail Tab'}")
        print("  " + "-" * 85)
        print(f"  {'Total Kernel Time':<45s} {grand_total:>14,.1f} {'100.0%':>7s}")

        root_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        for s in stats_list:
            root_by_type[s.module_type].append(s)
        sorted_root_types = sorted(
            root_by_type.items(),
            key=lambda x: -sum(
                (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                for s in x[1]))
        for rtype, rlist in sorted_root_types:
            rtype_total = sum(
                (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                for s in rlist)
            pct = rtype_total / grand_total * 100 if grand_total > 0 else 0
            chain = root_chains.get(rtype, [rtype])
            root_phase = self._infer_root_phase(rlist)
            print(f"  {rtype:<45s} {rtype_total:>14,.1f} {pct:>6.1f}%")
            if len(chain) > 1:
                children = chain[1:]
                for i, ctype in enumerate(children):
                    is_last = (i == len(children) - 1)
                    connector = "└── " if is_last else "├── "
                    ct = type_totals.get(ctype, 0)
                    ct_pct = ct / grand_total * 100 if grand_total > 0 else 0
                    label = f"    {connector}{ctype}"
                    if ctype in top_type_names:
                        if root_phase:
                            suffix = f" ({root_phase[:3]})"
                            tab = f"{ctype[:28 - len(suffix)]}{suffix}"
                        else:
                            tab = ctype[:28]
                    else:
                        tab = ""
                    print(f"  {label:<45s} {ct:>14,.1f} {ct_pct:>6.1f}%  {tab}")

        # --- Section 2: Category breakdown ---
        global_cat_agg: Dict[str, List] = {}
        self._collect_global_category_agg(stats_list, global_cat_agg, mode)
        sorted_cats = sorted(global_cat_agg.items(), key=lambda x: -x[1][0])

        print("\n" + "=" * 90)
        print("  Kernel Category Breakdown")
        print("=" * 90)
        print(f"  {'Category':<20s} {'Count':>8s} {'Total (us)':>14s} {'Avg (us)':>12s} {'%':>8s}")
        print("  " + "-" * 64)
        for cat, (dur, cnt) in sorted_cats:
            pct = dur / grand_total * 100 if grand_total > 0 else 0
            avg = dur / cnt if cnt > 0 else 0
            print(f"  {cat:<20s} {cnt:>8,d} {dur:>14,.1f} {avg:>12,.1f} {pct:>7.1f}%")
        print()

    def _select_max_detail_modules(self, stats_list: List[ModuleStats], mode: str,
                             max_detail_modules: int) -> Tuple[set, Dict[str, List[str]],
                                                         Dict[str, float], set,
                                                         Dict[str, set]]:
        """Select detail tab types, prioritizing root chain leaves.

        Returns (top_type_names, root_chains, type_totals,
                 wrapper_types, type_children_map).
        """
        type_agg_flat: Dict[str, List[float]] = defaultdict(list)
        self._collect_by_type(stats_list, type_agg_flat, mode)
        sorted_types_flat = sorted(type_agg_flat.items(), key=lambda x: -sum(x[1]))
        wrapper_types = self._find_wrapper_types(stats_list, mode)
        type_children_map = self._build_type_children(stats_list)
        type_totals = {mtype: sum(times) for mtype, times in sorted_types_flat}

        # Build wrapper chains for each root module type
        root_types = {s.module_type for s in stats_list}
        root_chains: Dict[str, List[str]] = {}
        for rtype in root_types:
            chain = self._trace_wrapper_chain_full(
                rtype, wrapper_types, type_children_map, stats_list, mode)
            root_chains[rtype] = chain

        # Step 1: Reserve slots for each root chain's leaves.
        # A chain may have multiple leaves when a wrapper has several
        # significant child types (e.g. LinearDecoderLayer + AttentionDecoderLayer).
        # Chain leaves are the last element(s) — they get detail tabs even if
        # they are technically wrappers themselves.
        top_type_names: set = set()
        skip_types: set = set(wrapper_types)
        for rtype in root_types:
            chain = root_chains[rtype]
            if len(chain) <= 1:
                continue
            # All chain elements after the root are candidate leaves
            leaves = chain[1:]
            for leaf in leaves:
                if leaf not in top_type_names:
                    top_type_names.add(leaf)
                    skip_types.discard(leaf)
                    desc = self._all_descendant_types(leaf, type_children_map)
                    skip_types.update(desc)

        # Step 2: Fill remaining slots from sorted list
        for mtype, _ in sorted_types_flat:
            if len(top_type_names) >= max_detail_modules:
                break
            if mtype in skip_types:
                continue
            top_type_names.add(mtype)
            desc = self._all_descendant_types(mtype, type_children_map)
            skip_types.update(desc)

        return top_type_names, root_chains, type_totals, wrapper_types, type_children_map

    def _collect_by_type(self, stats_list: List[ModuleStats],
                         agg: Dict[str, List[float]], mode: str):
        for s in stats_list:
            time_val = s.total_kernel_time if mode == "full" else s.total_cpu_op_time
            agg[s.module_type].append(time_val)
            self._collect_by_type(s.children_stats, agg, mode)

    def _collect_by_type_total(self, stats_list: List[ModuleStats],
                               agg: Dict[str, float], mode: str):
        """Collect total time per module type (summed, not list)."""
        for s in stats_list:
            time_val = s.total_kernel_time if mode == "full" else s.total_cpu_op_time
            agg[s.module_type] += time_val
            self._collect_by_type_total(s.children_stats, agg, mode)

    @staticmethod
    def _infer_root_phase(rlist: List[ModuleStats]) -> str:
        """Return the dominant phase for a group of root-level stats.

        Checks the root stats first; if they have no phase (common for
        wrapper modules like DeepseekV2Model), falls back to the majority
        phase of their immediate children.
        """
        for s in rlist:
            p = getattr(s, "phase", "")
            if p:
                return p
        from collections import Counter
        child_phases = Counter()
        for s in rlist:
            for c in s.children_stats:
                p = getattr(c, "phase", "")
                if p:
                    child_phases[p] += 1
        if child_phases:
            return child_phases.most_common(1)[0][0]
        return ""

    def export_excel(self, stats_list: List[ModuleStats], mode: str,
                     output_path: str, max_detail_modules: int = 3,
                     detail_modules: Optional[List[str]] = None,
):
        """Export analysis to Excel workbook."""
        try:
            import openpyxl
            import openpyxl.worksheet.properties
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return

        self.max_detail_modules = max_detail_modules
        wb = Workbook()
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        title_font = Font(bold=True, size=12)

        # Pre-compute grand total and category breakdown (used by Summary + Overview)
        grand_total = sum(
            (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
            for s in stats_list)
        total_kernel_count = sum(
            s.kernel_count if mode == "full" else s.cpu_op_count
            for s in stats_list)
        # Aggregate kernel categories across entire tree
        global_cat_agg: Dict[str, List] = {}
        self._collect_global_category_agg(stats_list, global_cat_agg, mode)

        # Pre-compute detail tab selection (needed by Summary tab and detail sheets)
        if detail_modules:
            top_type_names = set(detail_modules)
        else:
            top_type_names, _, _, _, _ = \
                self._select_max_detail_modules(stats_list, mode, max_detail_modules)

        # --- Summary tab (one-pager overview) ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        summary_last_row = self._write_summary_tab(
            ws_summary, stats_list, mode, grand_total,
            total_kernel_count, global_cat_agg,
            header_font, header_fill, title_font,
            top_type_names)

        # --- Overview sheet (hierarchical type tree, with % of Parent + stats) ---
        ws_ov = wb.create_sheet(title="Overview")
        type_headers = ["Module Type", "Depth", "Count",
                        "Mean (us)", "Std (us)", "Min (us)", "Max (us)",
                        "Total (us)", "% of Parent", "Top Kernel"]
        for col, h in enumerate(type_headers, 1):
            cell = ws_ov.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Build hierarchical type summary
        root_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        for s in stats_list:
            root_by_type[s.module_type].append(s)
        sorted_root_types = sorted(
            root_by_type.items(),
            key=lambda x: -sum(
                (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                for s in x[1]))
        type_row = 2
        for rtype, rlist in sorted_root_types:
            root_times = [(s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                          for s in rlist]
            rep = next((s for s in rlist if s.instance_id == 1), rlist[0])
            rep_time = rep.total_kernel_time if mode == "full" else rep.total_cpu_op_time
            root_rep_kernel = self._get_rep_kernel(rep, mode)
            self._write_type_row(ws_ov, type_row, rtype, 0, len(rlist),
                                 root_times, grand_total, bold=True,
                                 rep_kernel=root_rep_kernel)
            type_row += 1
            if rep.children_stats:
                type_row, _ = self._write_children_hierarchy(
                    ws_ov, rep, mode, type_row, rep_time, rlist)
        ws_ov.column_dimensions["A"].width = 40
        ws_ov.column_dimensions["J"].width = 80

        # --- Module Tree tab (full tree with collapsible outline groups) ---
        ws_tree = wb.create_sheet(title="Module Tree")
        # Outline settings: collapse buttons at top of group (summaryBelow=False)
        ws_tree.sheet_properties.outlinePr = openpyxl.worksheet.properties.Outline(
            summaryBelow=False, summaryRight=False)
        tree_headers = ["Module", "Total Time (us)", "Kernels/Ops",
                        "Breakdown", "Phase"]
        for col, h in enumerate(tree_headers, 1):
            cell = ws_tree.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        tree_row = 2
        seen_tree_roots: set = set()
        for s in stats_list:
            # Show one representative instance per root module type
            if s.module_type in seen_tree_roots:
                continue
            seen_tree_roots.add(s.module_type)
            tree_row = self._write_tree_rows(ws_tree, s, mode, tree_row, 0)
        ws_tree.column_dimensions["A"].width = 50
        ws_tree.column_dimensions["D"].width = 60

        # --- GPU Kernels sheet: global kernel name aggregation ---
        ws3 = wb.create_sheet(title="GPU Kernels")
        bd_headers = ["Kernel Name", "Category", "Total Duration (us)",
                       "Count", "Avg (us)", "% of Total", "Module Types"]
        for col, h in enumerate(bd_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        bd_row = 2
        bd_row = self._write_kernel_name_breakdown(ws3, stats_list, mode, bd_row,
                                                   grand_total)
        ws3.column_dimensions["A"].width = 80
        ws3.column_dimensions["G"].width = 50

        # --- Per-module-type detail sheets ---
        type_agg_flat: Dict[str, List[float]] = defaultdict(list)
        self._collect_by_type(stats_list, type_agg_flat, mode)
        type_total_time = {mtype: sum(times) for mtype, times in type_agg_flat.items()}
        # Seed detail-sheet representatives from the Module Tree instances so
        # that the detail tab numbers match the tree.  _find_median_instance
        # only fills types not already present.
        seen_types: Dict[Tuple[str, str], ModuleStats] = {}
        for s in stats_list:
            if s.module_type in seen_tree_roots:
                self._collect_tree_instances(s, seen_types)
                break  # only the first root was written to the tree
        self._find_median_instance_per_type(stats_list, seen_types, mode,
                                            force_types=top_type_names)
        # Sort detail tabs by total kernel time descending (highest % first)
        sorted_detail_items = sorted(
            seen_types.items(),
            key=lambda item: -type_total_time.get(item[0][0], 0))
        for (mtype, phase), rep_stats in sorted_detail_items:
            if mtype not in top_type_names:
                continue
            # Include phase suffix in sheet name for LLM traces
            if phase:
                suffix = f" ({phase[:3]})"  # "prefill" -> "(pre)", "decode" -> "(dec)"
                sheet_name = f"{mtype[:28 - len(suffix)]}{suffix}"
            else:
                sheet_name = mtype[:28]  # Excel sheet name limit = 31 chars
            ws_det = wb.create_sheet(title=sheet_name)
            all_details = self._collect_all_details(rep_stats)
            all_details.sort(key=lambda d: d.ts)

            # Compute wall time and sum of durations
            sum_dur = sum(d.duration for d in all_details)
            if all_details:
                wall_start = min(d.ts for d in all_details)
                wall_end = max(d.ts + d.duration for d in all_details)
                wall_time = wall_end - wall_start
            else:
                wall_time = 0

            # Row 1: title
            phase_label = f" [{phase}]" if phase else ""
            ws_det.cell(row=1, column=1,
                        value=f"{rep_stats.name}{phase_label} — {len(all_details)} kernels").font = Font(bold=True, size=12)
            ws_det.cell(row=2, column=1,
                        value=f"Kernel sum: {sum_dur:,.0f} us  |  "
                              f"Wall time: {wall_time:,.0f} us  |  "
                              f"Overlap: {sum_dur - wall_time:,.0f} us  "
                              f"(% uses wall time)")

            # Category summary table (rows 3+)
            cat_agg: Dict[str, float] = defaultdict(float)
            for d in all_details:
                cat_agg[d.category] += d.duration
            sorted_cats = sorted(cat_agg.items(), key=lambda x: -x[1])
            cat_total = sum(cat_agg.values())

            cur_row = 3
            for lbl, col_idx in [("Category", 1), ("Percentage (%)", 2),
                                 ("Total Duration (us)", 3)]:
                cell = ws_det.cell(row=cur_row, column=col_idx, value=lbl)
                cell.font = header_font
                cell.fill = header_fill
            cur_row += 1
            for cat, dur in sorted_cats:
                pct_cat = dur / cat_total * 100 if cat_total > 0 else 0
                ws_det.cell(row=cur_row, column=1, value=cat)
                ws_det.cell(row=cur_row, column=2, value=f"{pct_cat:.0f}%")
                ws_det.cell(row=cur_row, column=3, value=round(dur, 1))
                cur_row += 1
            ws_det.cell(row=cur_row, column=1, value="Total").font = Font(bold=True)
            ws_det.cell(row=cur_row, column=2, value="100%").font = Font(bold=True)
            ws_det.cell(row=cur_row, column=3, value=round(cat_total, 1)).font = Font(bold=True)
            cur_row += 2  # blank row before kernel list

            # Kernel detail headers
            det_headers = ["Module", "Input Dims", "Kernel Name",
                          "Duration (us)", "% of wall time", "Category",
                          "Path"]
            for col, h in enumerate(det_headers, 1):
                cell = ws_det.cell(row=cur_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            cur_row += 1
            detail_truncated = len(all_details) > MAX_ROWS_PER_TAB
            for i, d in enumerate(all_details[:MAX_ROWS_PER_TAB], 1):
                pct = d.duration / wall_time * 100 if wall_time > 0 else 0
                leaf = d.module_path.rsplit("/", 1)[-1] if "/" in d.module_path else d.module_path
                ws_det.cell(row=cur_row, column=1, value=leaf)
                ws_det.cell(row=cur_row, column=2, value=d.input_dims)
                ws_det.cell(row=cur_row, column=3, value=d.name)
                ws_det.cell(row=cur_row, column=4, value=round(d.duration, 1))
                ws_det.cell(row=cur_row, column=5, value=round(pct, 1))
                ws_det.cell(row=cur_row, column=6, value=d.category)
                ws_det.cell(row=cur_row, column=7, value=d.source_path)
                cur_row += 1
            if detail_truncated:
                ws_det.cell(row=cur_row, column=1,
                            value=f"... truncated at {MAX_ROWS_PER_TAB} rows")
            ws_det.column_dimensions["A"].width = 35
            ws_det.column_dimensions["B"].width = 60
            ws_det.column_dimensions["C"].width = 80
            ws_det.column_dimensions["G"].width = 80

        wb.save(output_path)
        print(f"\nExcel report saved to: {output_path}")

    @staticmethod
    def _time_stats(times: List[float]) -> Tuple[float, float, float, float, float]:
        """Compute mean, std, min, max, total from a list of times."""
        n = len(times)
        total = sum(times)
        if n == 0:
            return 0, 0, 0, 0, 0
        mean = total / n
        if n > 1:
            variance = sum((t - mean) ** 2 for t in times) / (n - 1)
            std = math.sqrt(variance)
        else:
            std = 0
        return mean, std, min(times), max(times), total

    @staticmethod
    def _top_kernel_name(stats: ModuleStats) -> Tuple[str, float]:
        """Find the kernel name with the highest total duration from self kernels.

        Returns (kernel_name, total_duration).  Aggregates across all instances
        of the same kernel name within this module's own kernel_details.
        """
        if not stats.kernel_details:
            return "", 0.0
        by_name: Dict[str, float] = {}
        for d in stats.kernel_details:
            by_name[d.name] = by_name.get(d.name, 0.0) + d.duration
        if not by_name:
            return "", 0.0
        best_name = max(by_name, key=by_name.get)  # type: ignore[arg-type]
        return best_name, by_name[best_name]

    def _collect_global_category_agg(self, stats_list: List[ModuleStats],
                                     agg: Dict[str, List], mode: str):
        """Aggregate kernel categories globally: cat -> [total_dur, count]."""
        for s in stats_list:
            for d in s.kernel_details:
                entry = agg.get(d.category)
                if entry:
                    entry[0] += d.duration
                    entry[1] += 1
                else:
                    agg[d.category] = [d.duration, 1]
            self._collect_global_category_agg(s.children_stats, agg, mode)

    def _write_summary_tab(self, ws, stats_list: List[ModuleStats], mode: str,
                           grand_total: float, total_kernel_count: int,
                           global_cat_agg: Dict[str, List],
                           header_font, header_fill, title_font,
                           top_type_names: set):
        """Write the Summary tab: kernel time summary with detail tabs, category breakdown."""
        from openpyxl.styles import Font, PatternFill

        # Build wrapper chains and type totals
        root_chains: Dict[str, List[str]] = {}
        type_totals: Dict[str, float] = defaultdict(float)
        self._collect_by_type_total(stats_list, type_totals, mode)
        wrapper_types = self._find_wrapper_types(stats_list, mode)
        type_children_map = self._build_type_children(stats_list)
        root_types = {s.module_type for s in stats_list}
        for rtype in root_types:
            chain = self._trace_wrapper_chain_full(
                rtype, wrapper_types, type_children_map, stats_list, mode)
            root_chains[rtype] = chain
        top_type_set = top_type_names

        row = 1

        # --- Section 1: Kernel Time Summary ---
        ws.cell(row=row, column=1, value="Metric").font = header_font
        ws.cell(row=row, column=2, value="Value (us)").font = header_font
        ws.cell(row=row, column=3, value="Percentage").font = header_font
        ws.cell(row=row, column=4, value="Detail Tab").font = header_font
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = header_fill
        row += 1

        ws.cell(row=row, column=1, value="Total Kernel Time")
        ws.cell(row=row, column=2, value=round(grand_total, 1))
        ws.cell(row=row, column=3, value="100.0%")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        root_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        for s in stats_list:
            root_by_type[s.module_type].append(s)
        sorted_root_types = sorted(
            root_by_type.items(),
            key=lambda x: -sum(
                (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                for s in x[1]))
        for rtype, rlist in sorted_root_types:
            rtype_total = sum(
                (s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                for s in rlist)
            pct = rtype_total / grand_total * 100 if grand_total > 0 else 0
            chain = root_chains.get(rtype, [rtype])
            root_phase = self._infer_root_phase(rlist)
            ws.cell(row=row, column=1, value=rtype)
            ws.cell(row=row, column=2, value=round(rtype_total, 1))
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
            row += 1
            if len(chain) > 1:
                children = chain[1:]
                for i, ctype in enumerate(children):
                    is_last = (i == len(children) - 1)
                    connector = "└── " if is_last else "├── "
                    ct = type_totals.get(ctype, 0)
                    ct_pct = ct / grand_total * 100 if grand_total > 0 else 0
                    ws.cell(row=row, column=1, value=f"    {connector}{ctype}")
                    ws.cell(row=row, column=2, value=round(ct, 1))
                    ws.cell(row=row, column=3, value=f"{ct_pct:.1f}%")
                    if ctype in top_type_set:
                        if root_phase:
                            suffix = f" ({root_phase[:3]})"
                            detail_ref = f"{ctype[:28 - len(suffix)]}{suffix}"
                        else:
                            detail_ref = ctype[:28]
                        ws.cell(row=row, column=4, value=detail_ref)
                        for c in range(1, 5):
                            ws.cell(row=row, column=c).font = Font(bold=True)
                    row += 1

        row += 1  # blank row

        # --- Section 2: Category breakdown ---
        ws.cell(row=row, column=1, value="Category").font = header_font
        ws.cell(row=row, column=2, value="Count").font = header_font
        ws.cell(row=row, column=3, value="Total (us)").font = header_font
        ws.cell(row=row, column=4, value="Avg (us)").font = header_font
        ws.cell(row=row, column=5, value="Percentage").font = header_font
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = header_fill
        row += 1

        sorted_cats = sorted(global_cat_agg.items(), key=lambda x: -x[1][0])
        for cat, (dur, cnt) in sorted_cats:
            pct = dur / grand_total * 100 if grand_total > 0 else 0
            avg = dur / cnt if cnt > 0 else 0
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=cnt)
            ws.cell(row=row, column=3, value=round(dur, 1))
            ws.cell(row=row, column=4, value=round(avg, 1))
            ws.cell(row=row, column=5, value=f"{pct:.1f}%")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15

        return row

    def _write_type_row(self, ws, row: int, label: str, depth: int, count: int,
                        times: List[float], parent_time: float,
                        bold: bool = False, pct_total: Optional[float] = None,
                        rep_kernel: str = ""):
        """Write one row of the Overview hierarchy with stats columns.

        pct_total: if given, use this as the numerator for % of Parent
                   (for children: sum within one parent, not global total).
        rep_kernel: representative kernel name for this row.
        """
        from openpyxl.styles import Font
        mean, std, tmin, tmax, total = self._time_stats(times)
        pct_num = pct_total if pct_total is not None else total
        pct = pct_num / parent_time * 100 if parent_time > 0 else 0
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=depth)
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=round(mean, 1))
        ws.cell(row=row, column=5, value=round(std, 1))
        ws.cell(row=row, column=6, value=round(tmin, 1))
        ws.cell(row=row, column=7, value=round(tmax, 1))
        ws.cell(row=row, column=8, value=round(total, 1))
        ws.cell(row=row, column=9, value=round(pct, 1))
        ws.cell(row=row, column=10, value=rep_kernel)
        if bold:
            for c in range(1, 11):
                ws.cell(row=row, column=c).font = Font(bold=True)

    def _collect_all_of_type(self, stats_list: List[ModuleStats],
                            module_type: str) -> List[ModuleStats]:
        """Collect ALL instances of a module type across the entire tree."""
        result = []
        for s in stats_list:
            if s.module_type == module_type:
                result.append(s)
            result.extend(self._collect_all_of_type(s.children_stats, module_type))
        return result

    def _write_children_hierarchy(self, ws, rep: ModuleStats, mode: str,
                                  row: int, parent_time: float,
                                  all_stats: List[ModuleStats],
                                  max_depth: int = 5) -> Tuple[int, str]:
        """Recursively write children rows up to max_depth, with cross-instance stats.

        Returns (next_row, rep_kernel) where rep_kernel is the representative
        kernel name from the highest-time child (bubbled up for parent rows).
        """
        if not rep.children_stats:
            kname, _ = self._top_kernel_name(rep)
            return row, kname

        child_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        for child in rep.children_stats:
            child_by_type[child.module_type].append(child)

        sorted_children = sorted(
            child_by_type.items(),
            key=lambda x: -sum(
                (c.total_kernel_time if mode == "full" else c.total_cpu_op_time)
                for c in x[1]))

        rep_self_time = (rep.self_kernel_time if mode == "full"
                         else rep.self_cpu_op_time)
        child_depth = rep.depth + 1

        # Track the highest-time child's representative kernel for bubbling up
        best_child_kernel = ""
        best_child_time = -1.0

        for ctype, clist_in_rep in sorted_children:
            cdepth = clist_in_rep[0].depth
            if cdepth > max_depth:
                continue
            all_of_type = self._collect_all_of_type(all_stats, ctype)
            if all_of_type:
                child_times = [(c.total_kernel_time if mode == "full" else c.total_cpu_op_time)
                               for c in all_of_type]
            else:
                child_times = [(c.total_kernel_time if mode == "full" else c.total_cpu_op_time)
                               for c in clist_in_rep]
            child_indent = "    " * cdepth + "├── "
            rep_child_total = sum(
                (c.total_kernel_time if mode == "full" else c.total_cpu_op_time)
                for c in clist_in_rep)

            # Recurse first to get child's rep_kernel before writing this row
            child_rep = next((c for c in clist_in_rep if c.instance_id == 1), clist_in_rep[0])
            child_rep_time = child_rep.total_kernel_time if mode == "full" else child_rep.total_cpu_op_time

            # For leaf children (no sub-children), use their own top kernel
            # For parent children, recurse to get bubbled-up kernel
            if not child_rep.children_stats:
                child_kernel, _ = self._top_kernel_name(child_rep)
            else:
                # Peek: the recursion below will write sub-rows and return the kernel
                # We need to write THIS row first, then recurse.
                # So compute the kernel from the highest-time grandchild.
                child_kernel = self._get_rep_kernel(child_rep, mode)

            self._write_type_row(ws, row, f"{child_indent}{ctype}",
                                 cdepth, len(child_times), child_times,
                                 parent_time, pct_total=rep_child_total,
                                 rep_kernel=child_kernel)
            row += 1

            row, _ = self._write_children_hierarchy(
                ws, child_rep, mode, row, child_rep_time, all_stats, max_depth)

            if rep_child_total > best_child_time:
                best_child_time = rep_child_total
                best_child_kernel = child_kernel

        # Add "(self)" row if self_time is significant (>0.5% of parent)
        if rep_self_time > 0 and parent_time > 0:
            self_pct = rep_self_time / parent_time * 100
            if self_pct >= 0.5:
                self_indent = "    " * child_depth + "└── "
                from openpyxl.styles import Font
                self_kernel, _ = self._top_kernel_name(rep)
                self._write_type_row(
                    ws, row, f"{self_indent}(self)", child_depth, 1,
                    [rep_self_time], parent_time, pct_total=rep_self_time,
                    rep_kernel=self_kernel)
                for c in range(1, 11):
                    ws.cell(row=row, column=c).font = Font(italic=True, color="888888")
                row += 1
                if rep_self_time > best_child_time:
                    best_child_kernel = self_kernel

        return row, best_child_kernel

    def _get_rep_kernel(self, stats: ModuleStats, mode: str) -> str:
        """Get representative kernel name by following the highest-time child type
        down the tree until a leaf is reached.

        Groups children by module_type (matching the Summary sheet's grouping)
        so that e.g. 4 WanUpBlock instances are summed together before comparing
        against self_time.
        """
        if not stats.children_stats:
            kname, _ = self._top_kernel_name(stats)
            return kname

        # Group children by type and sum their times
        child_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        for c in stats.children_stats:
            child_by_type[c.module_type].append(c)

        best_type = ""
        best_type_time = 0.0
        best_type_rep: Optional[ModuleStats] = None
        for ctype, clist in child_by_type.items():
            type_total = sum(
                c.total_kernel_time if mode == "full" else c.total_cpu_op_time
                for c in clist)
            if type_total > best_type_time:
                best_type_time = type_total
                best_type = ctype
                best_type_rep = next(
                    (c for c in clist if c.instance_id == 1), clist[0])

        self_time = stats.self_kernel_time if mode == "full" else stats.self_cpu_op_time

        if self_time > best_type_time or best_type_rep is None:
            kname, _ = self._top_kernel_name(stats)
            return kname

        return self._get_rep_kernel(best_type_rep, mode)

    def _write_tree_rows(self, ws, stats: ModuleStats, mode: str,
                         row: int, depth: int, max_depth: int = 3) -> int:
        """Write tree-structured rows with Excel outline grouping for collapse/expand."""
        if depth > max_depth:
            return row
        if row > MAX_ROWS_PER_TAB + 1:
            ws.cell(row=row, column=1, value=f"... truncated at {MAX_ROWS_PER_TAB} rows")
            return row + 1
        time_val = stats.total_kernel_time if mode == "full" else stats.total_cpu_op_time
        count = stats.kernel_count if mode == "full" else stats.cpu_op_count
        phase = getattr(stats, "phase", "")

        indent = ("    " * depth + "├── ") if depth > 0 else ""
        ws.cell(row=row, column=1, value=f"{indent}{stats.name}")
        ws.cell(row=row, column=2, value=round(time_val, 1))
        ws.cell(row=row, column=3, value=count)

        # Build breakdown string
        if time_val > 0 and stats.kernel_breakdown:
            parts = []
            sorted_cats = sorted(stats.kernel_breakdown.items(), key=lambda x: -x[1][0])
            for cat, (dur, cnt) in sorted_cats:
                if dur <= 0:
                    continue
                pct = dur / time_val * 100
                parts.append(f"{cat}: {dur:,.0f} us ({pct:.1f}%)")
            ws.cell(row=row, column=4, value=", ".join(parts))

        ws.cell(row=row, column=5, value=phase)

        # Set outline level for collapsible grouping (depth > 0 can be collapsed)
        if depth > 0:
            ws.row_dimensions[row].outline_level = min(depth, 8)  # Excel max = 8

        row += 1

        for child in stats.children_stats:
            row = self._write_tree_rows(ws, child, mode, row, depth + 1)

        return row

    @staticmethod
    def _find_wrapper_types(stats_list: List[ModuleStats], mode: str,
                            threshold: float = 0.8) -> set:
        """Find module types that are thin wrappers around a dominant child.

        A type is a wrapper if, in its representative instance, a single child
        type accounts for >threshold of its total time.  These are skipped in
        auto-selection to prefer the smaller repetitive unit.
        Applies transitively: if A wraps B wraps C, both A and B are skipped.

        Also treats a type as a wrapper if it has very few instances (<=2) and
        its children collectively account for >threshold of its time with at
        least one child type having multiple instances (the repetitive unit).
        This handles models like Qwen3_5MoeForCausalLM which has two child
        decoder layer types (Linear + Attention) that together account for
        nearly all the time.
        """
        # Collect one representative per type (first instance seen)
        reps: Dict[str, ModuleStats] = {}
        ReportGenerator._collect_first_instance(stats_list, reps)

        # Count instances per type across the tree
        type_counts: Dict[str, int] = defaultdict(int)
        ReportGenerator._count_instances(stats_list, type_counts)

        skip: set = set()
        for mtype, rep in reps.items():
            parent_time = (rep.total_kernel_time if mode == "full"
                           else rep.total_cpu_op_time)
            if parent_time <= 0 or not rep.children_stats:
                continue
            # Group children by type within this one instance
            child_by_type: Dict[str, float] = defaultdict(float)
            for c in rep.children_stats:
                ct = c.total_kernel_time if mode == "full" else c.total_cpu_op_time
                child_by_type[c.module_type] += ct

            # Original check: single dominant child
            for ctype, ctime in child_by_type.items():
                if ctime / parent_time > threshold:
                    skip.add(mtype)
                    break

            if mtype in skip:
                continue

            # Additional check: non-repetitive parent (<=2 instances) whose
            # children collectively dominate and include repetitive types.
            if type_counts.get(mtype, 0) <= 2:
                total_child_time = sum(child_by_type.values())
                has_repetitive_child = any(
                    type_counts.get(ct, 0) > 2 for ct in child_by_type)
                if (total_child_time / parent_time > threshold
                        and has_repetitive_child):
                    skip.add(mtype)

        return skip

    @staticmethod
    def _collect_first_instance(stats_list: List[ModuleStats],
                                out: Dict[str, ModuleStats]):
        """Collect first instance of each module type from the tree."""
        for s in stats_list:
            if s.module_type not in out:
                out[s.module_type] = s
            ReportGenerator._collect_first_instance(s.children_stats, out)

    @staticmethod
    def _count_instances(stats_list: List[ModuleStats],
                         out: Dict[str, int]):
        """Count total instances of each module type across the tree."""
        for s in stats_list:
            out[s.module_type] += 1
            ReportGenerator._count_instances(s.children_stats, out)

    @staticmethod
    def _build_type_children(stats_list: List[ModuleStats]) -> Dict[str, set]:
        """Build parent_type -> set of direct child types from the stats tree."""
        result: Dict[str, set] = defaultdict(set)
        for s in stats_list:
            for child in s.children_stats:
                result[s.module_type].add(child.module_type)
            child_map = ReportGenerator._build_type_children(s.children_stats)
            for ptype, ctypes in child_map.items():
                result[ptype].update(ctypes)
        return result

    @staticmethod
    def _all_descendant_types(mtype: str, type_children: Dict[str, set]) -> set:
        """Get all transitive descendant types of a given module type."""
        result: set = set()
        stack = list(type_children.get(mtype, set()))
        while stack:
            ct = stack.pop()
            if ct not in result:
                result.add(ct)
                stack.extend(type_children.get(ct, set()))
        return result

    @staticmethod
    def _trace_wrapper_chain(mtype: str, wrapper_types: set,
                             type_children: Dict[str, set],
                             stats_list: List[ModuleStats],
                             mode: str) -> str:
        """Follow wrapper chain from mtype to the first non-wrapper descendant."""
        chain = ReportGenerator._trace_wrapper_chain_full(
            mtype, wrapper_types, type_children, stats_list, mode)
        return chain[-1] if chain else mtype

    @staticmethod
    def _trace_wrapper_chain_full(mtype: str, wrapper_types: set,
                                  type_children: Dict[str, set],
                                  stats_list: List[ModuleStats],
                                  mode: str) -> List[str]:
        """Follow wrapper chain, returning the full list [mtype, child, ..., leaf].

        When a wrapper has multiple significant child types (e.g.
        Qwen3_5MoeForCausalLM with LinearDecoderLayer + AttentionDecoderLayer),
        all significant children are included in the chain.
        """
        reps: Dict[str, ModuleStats] = {}
        ReportGenerator._collect_first_instance(stats_list, reps)
        chain = [mtype]
        current = mtype
        visited: set = set()
        while current in wrapper_types and current not in visited:
            visited.add(current)
            rep = reps.get(current)
            if not rep or not rep.children_stats:
                break
            child_by_type: Dict[str, float] = defaultdict(float)
            for c in rep.children_stats:
                ct = c.total_kernel_time if mode == "full" else c.total_cpu_op_time
                child_by_type[c.module_type] += ct
            parent_time = (rep.total_kernel_time if mode == "full"
                           else rep.total_cpu_op_time)
            # Include all child types that account for >10% of parent time
            significant = sorted(
                ((ct, cname) for cname, ct in child_by_type.items()
                 if parent_time > 0 and ct / parent_time > 0.10),
                reverse=True)
            if not significant:
                break
            if len(significant) > 1:
                for _, cname in significant:
                    chain.append(cname)
                break  # multi-child wrapper: don't recurse further
            chain.append(significant[0][1])
            current = significant[0][1]
        return chain

    @staticmethod
    def _collect_tree_instances(root_stats: ModuleStats,
                                out: Dict[Tuple[str, str], ModuleStats]):
        """Collect first instance of each (module_type, phase) from a tree.

        Walks the same tree that _write_tree_rows renders in the Module Tree
        tab, recording the first instance per (type, phase) pair.  These are
        used to seed the detail-sheet representatives so that detail tabs
        show the exact same instances visible in the tree.
        """
        key = (root_stats.module_type, getattr(root_stats, "phase", ""))
        if key not in out and root_stats.kernel_count > 0:
            out[key] = root_stats
        for child in root_stats.children_stats:
            ReportGenerator._collect_tree_instances(child, out)

    def _find_median_instance_per_type(self, stats_list: List[ModuleStats],
                                       seen: Dict[Tuple[str, str], ModuleStats],
                                       mode: str,
                                       force_types: Optional[set] = None):
        """Find median instance of each (module_type, phase) pair.

        Groups by (module_type, phase) so that prefill and decode get separate
        representative instances, producing separate detail sheets.
        Skips (type, phase) pairs already present in *seen* (e.g. seeded
        from the Module Tree) so that detail tabs stay consistent with
        the tree.

        force_types: if given, always include these types even if they are leaf modules.
        """
        # First collect all instances per type
        all_by_type: Dict[str, List[ModuleStats]] = defaultdict(list)
        self._collect_instances_by_type(stats_list, all_by_type)
        # Pick median for each (type, phase) pair
        for mtype, instances in all_by_type.items():
            if not any(s.children_stats for s in instances):
                if not (force_types and mtype in force_types):
                    continue
            # Sub-group by phase
            by_phase: Dict[str, List[ModuleStats]] = defaultdict(list)
            for s in instances:
                by_phase[getattr(s, "phase", "")].append(s)
            for phase, phase_instances in by_phase.items():
                if (mtype, phase) in seen:
                    continue
                timed = sorted(phase_instances,
                               key=lambda s: s.total_kernel_time if mode == "full" else s.total_cpu_op_time)
                mid = len(timed) // 2
                seen[(mtype, phase)] = timed[mid]

    def _collect_instances_by_type(self, stats_list: List[ModuleStats],
                                   out: Dict[str, List[ModuleStats]]):
        for s in stats_list:
            if s.kernel_count > 0:
                out[s.module_type].append(s)
            self._collect_instances_by_type(s.children_stats, out)

    def _write_kernel_name_breakdown(self, ws, stats_list: List[ModuleStats],
                                     mode: str, row: int,
                                     grand_total: float) -> int:
        """Write kernel name breakdown aggregated globally (one row per unique kernel name)."""
        # Aggregate across entire tree: kernel_name -> (total_dur, count, category, module_types)
        global_agg: Dict[str, List] = {}  # name -> [dur, count, category, set(module_types)]
        self._collect_global_kernel_agg(stats_list, global_agg)

        sorted_kernels = sorted(global_agg.items(), key=lambda x: -x[1][0])
        truncated = len(sorted_kernels) > MAX_ROWS_PER_TAB
        sorted_kernels = sorted_kernels[:MAX_ROWS_PER_TAB]
        for kname, (dur, cnt, cat, mtypes) in sorted_kernels:
            pct = dur / grand_total * 100 if grand_total > 0 else 0
            ws.cell(row=row, column=1, value=kname)
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=round(dur, 1))
            ws.cell(row=row, column=4, value=cnt)
            ws.cell(row=row, column=5, value=round(dur / cnt, 1) if cnt > 0 else 0)
            ws.cell(row=row, column=6, value=round(pct, 1))
            ws.cell(row=row, column=7, value=", ".join(sorted(mtypes)))
            row += 1
        if truncated:
            ws.cell(row=row, column=1, value=f"... truncated at {MAX_ROWS_PER_TAB} rows")
            row += 1
        return row

    def _collect_global_kernel_agg(self, stats_list: List[ModuleStats],
                                   agg: Dict[str, List]):
        """Recursively aggregate kernel durations globally by kernel name."""
        for s in stats_list:
            for d in s.kernel_details:
                entry = agg.get(d.name)
                if entry:
                    entry[0] += d.duration
                    entry[1] += 1
                    entry[3].add(s.module_type)
                else:
                    agg[d.name] = [d.duration, 1, d.category, {s.module_type}]
            self._collect_global_kernel_agg(s.children_stats, agg)


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

class TraceModuleAnalyzer:
    """Main orchestrator: load trace, build module tree, correlate, aggregate, report."""

    def __init__(self, trace_path: str,
                 output_path: Optional[str] = None,
                 detail_modules: Optional[List[str]] = None,
                 module_index: Optional[int] = None,
                 max_detail_modules: int = 3,
                 auto_fix_rocm: bool = True,
                 model_info: bool = False,
                 port: int = 8765):
        self.trace_path = trace_path
        self.output_path = output_path
        self.detail_modules = detail_modules or []
        self.module_index = module_index
        self.max_detail_modules = max_detail_modules
        self.auto_fix_rocm = auto_fix_rocm
        self.model_info = model_info
        self.port = port

    def run(self):
        import time as _time
        _t0 = _time.monotonic()
        def _elapsed():
            return f"[{_time.monotonic() - _t0:.1f}s]"

        # Step 1: Load trace with single-pass categorization
        print(f"Loading trace: {self.trace_path}")
        data = self._load_trace(self.trace_path)
        events = data.get("traceEvents", [])
        print(f"  Total events: {len(events):,}")

        # Auto-detect and fix ROCm traces (missing hipGraphLaunch flow events)
        if self.auto_fix_rocm and _HAS_ROCM_FIX:
            data, n_launches, n_injected = _rocm_fix_trace(data)
            events = data.get("traceEvents", [])
            if n_launches > 0:
                print(f"  ROCm trace detected: {n_launches:,} hipGraphLaunch events, "
                      f"{n_injected:,} flow events injected")
        elif self.auto_fix_rocm and not _HAS_ROCM_FIX:
            logger.debug("fix_rocm_trace_flow not available — skipping ROCm auto-fix")

        # Single-pass categorization: extract everything we need
        kernel_events = []
        runtime_events = []
        driver_events = []
        cpu_ops = []
        module_events = []  # pre-filtered nn.Module events
        pyfunc_events = []  # non-module python_function events (for source path)
        gpu_memcpy = []
        gpu_memset = []
        phase_markers = []  # (ts, phase, tid, pid) for prefill/decode detection

        MODULE_PREFIX = "nn.Module: "
        for e in events:
            cat = e.get("cat", "")
            if cat == "kernel":
                kernel_events.append(e)
            elif cat == "cuda_runtime":
                runtime_events.append(e)
            elif cat == "cuda_driver":
                driver_events.append(e)
            elif cat == "cpu_op":
                cpu_ops.append(e)
            elif cat == "gpu_memcpy":
                gpu_memcpy.append(e)
            elif cat == "gpu_memset":
                gpu_memset.append(e)
            elif cat == "python_function":
                name = e.get("name", "")
                if name.startswith(MODULE_PREFIX) and e.get("dur") is not None:
                    module_events.append(e)
                else:
                    if e.get("dur") is not None:
                        pyfunc_events.append(e)
                    if "model_runner" in name and e.get("dur") is not None:
                        if ": forward_extend" in name:
                            phase_markers.append((e["ts"], e["ts"] + e["dur"],
                                                  "prefill", e["tid"], e.get("pid")))
                        elif ": forward_decode" in name:
                            phase_markers.append((e["ts"], e["ts"] + e["dur"],
                                                  "decode", e["tid"], e.get("pid")))

        # Free the raw events list — we no longer need it
        del events
        del data

        # Detect mode
        mode = "full" if kernel_events else "cpu_only"
        print(f"  kernel events: {len(kernel_events):,}")
        print(f"  cuda_runtime events: {len(runtime_events):,}")
        print(f"  cuda_driver events: {len(driver_events):,}")
        print(f"  cpu_op events: {len(cpu_ops):,}")
        print(f"  nn.Module events: {len(module_events):,}")
        print(f"  python_function events: {len(pyfunc_events):,}")
        print(f"  gpu_memcpy events: {len(gpu_memcpy):,}")
        print(f"  gpu_memset events: {len(gpu_memset):,}")
        print(f"  phase markers: {len(phase_markers):,}")
        print(f"  Mode: {mode}")

        # Step 2: Build module hierarchy tree
        print(f"  {_elapsed()} event categorization done")
        if not module_events:
            print("\nWARNING: No nn.Module events found. Trace may not have been captured "
                  "with `with_modules=True`.")
            print("Falling back to basic kernel-only analysis.")
            self._fallback_kernel_only(kernel_events + gpu_memcpy + gpu_memset)
            return

        print("\nBuilding module hierarchy...")
        builder = ModuleTreeBuilder()
        roots = builder.build_from_module_events(module_events)
        del module_events  # free memory
        print(f"  Root modules: {len(roots)}")
        if not roots:
            print("ERROR: Failed to build module tree.")
            return

        # Print root module names
        for r in roots[:10]:
            print(f"    {r.name} (tid={r.tid}, {r.end - r.ts:,.0f} us)")
        if len(roots) > 10:
            print(f"    ... and {len(roots) - 10} more")

        # Step 3: Correlate compute events → modules
        print(f"  {_elapsed()} module hierarchy built")
        if mode == "full":
            print("\nBuilding cpu_op shape index for Input Dims...")
            shape_index = CpuOpShapeIndex(cpu_ops, runtime_events, driver_events)
            print(f"  Indexed {len(shape_index._corr_to_shape):,} correlation→shape mappings")

            print("Building python source index for kernel source paths...")
            source_index = PythonSourceIndex(pyfunc_events, runtime_events, driver_events)
            print(f"  Indexed {len(source_index._corr_to_source):,} correlation→source mappings")
            del pyfunc_events  # free memory

            print("Correlating kernels to modules via cuda_runtime + cuda_driver...")
            all_gpu_events = kernel_events + gpu_memcpy + gpu_memset
            correlator = KernelCorrelator(runtime_events, roots,
                                          driver_events=driver_events)
            graph_correlator = CudaGraphCorrelator(runtime_events)
            del runtime_events, driver_events  # free memory
            matched = correlator.correlate(all_gpu_events, roots,
                                           shape_index=shape_index,
                                           source_index=source_index)
            del shape_index, source_index  # free memory
            print(f"  Matched {matched:,} / {len(all_gpu_events):,} GPU events to modules")

            # Step 3b: CUDA graph replay correlation
            if graph_correlator.has_graph_replays:
                unmatched = [e for e in all_gpu_events if not e.get("_matched")]
                graph_roots, graph_matched = graph_correlator.correlate(
                    unmatched, roots)
                if graph_roots:
                    roots.extend(graph_roots)
                    matched += graph_matched
                    print(f"  CUDA graph: {graph_matched:,} kernels in "
                          f"{len(graph_roots)} replay(s)")
                    print(f"  Total matched: {matched:,} / "
                          f"{len(all_gpu_events):,} GPU events")
        else:
            print("\nCorrelating cpu_ops to modules via time containment...")
            del runtime_events, driver_events, pyfunc_events
            correlator = CpuOpCorrelator(roots)
            matched = correlator.correlate(cpu_ops, roots)
            print(f"  Matched {matched:,} / {len(cpu_ops):,} cpu_ops to modules")
            del cpu_ops

        print(f"  {_elapsed()} correlation done")
        # Step 4: Detect prefill vs decode phase
        phase_detector = PhaseDetector()
        phase_detector.detect_from_markers(roots, phase_markers)
        self._propagate_phase(roots)

        # Step 5: Aggregate stats
        print(f"  {_elapsed()} phase detection done")
        print("\nAggregating module statistics...")
        aggregator = ModuleAggregator()
        stats_list = aggregator.aggregate(roots, mode)
        # Propagate phase from nodes
        self._copy_phase_to_stats(roots, stats_list)

        print(f"  {_elapsed()} aggregation done")
        # Step 6: Output
        reporter = ReportGenerator()
        reporter.print_type_summary(stats_list, mode,
                                    max_detail_modules=self.max_detail_modules,
                                    detail_modules=self.detail_modules)

        for dm in self.detail_modules:
            reporter.print_layer_detail(stats_list, mode, dm,
                                        self.module_index)

        print(f"  {_elapsed()} console output done")

        # Determine xlsx path: explicit -o, or auto-generate for --model-info
        xlsx_path = self.output_path
        tmp_xlsx = None
        if not xlsx_path and self.model_info:
            # Generate a temporary xlsx so visualize_module_tree can read it
            trace_dir = os.path.dirname(os.path.abspath(self.trace_path))
            trace_base = os.path.splitext(os.path.basename(self.trace_path))[0]
            if trace_base.endswith(".json"):
                trace_base = trace_base[:-5]
            xlsx_path = os.path.join(trace_dir, f"{trace_base}_analysis.xlsx")
            tmp_xlsx = xlsx_path

        if xlsx_path:
            reporter.export_excel(stats_list, mode, xlsx_path,
                                  max_detail_modules=self.max_detail_modules,
                                  detail_modules=self.detail_modules)
            print(f"  {_elapsed()} excel export done")

        if self.model_info and xlsx_path:
            html_path = os.path.splitext(xlsx_path)[0] + "_module_tree.html"
            result = _generate_model_info_html(xlsx_path, html_path,
                                               serve=True, port=self.port)
            if result:
                print(f"\n  Module tree visualization: {result}")
            else:
                print("\n  WARNING: Could not generate module tree visualization")

            # Clean up temporary xlsx if we created one
            if tmp_xlsx and os.path.isfile(tmp_xlsx):
                try:
                    os.unlink(tmp_xlsx)
                except OSError:
                    pass

    def _propagate_phase(self, nodes: List[ModuleNode]):
        """Propagate phase from parent to children if not set."""
        for node in nodes:
            phase = getattr(node, "_phase", None)
            if phase:
                self._set_phase_recursive(node.children, phase)
            self._propagate_phase(node.children)

    def _set_phase_recursive(self, nodes: List[ModuleNode], phase: str):
        for node in nodes:
            if not getattr(node, "_phase", None):
                node._phase = phase  # type: ignore[attr-defined]
            self._set_phase_recursive(node.children, phase)

    def _copy_phase_to_stats(self, nodes: List[ModuleNode], stats_list: List[ModuleStats]):
        for node, stats in zip(nodes, stats_list):
            stats.phase = getattr(node, "_phase", "")
            self._copy_phase_to_stats(node.children, stats.children_stats)


    def _fallback_kernel_only(self, kernel_events: List[Dict]):
        """Basic kernel-only analysis when no module events are present."""
        if not kernel_events:
            print("No kernel events found either. Nothing to analyze.")
            return

        total_dur = sum(k.get("dur", 0) for k in kernel_events)
        breakdown: Dict[str, Tuple[float, int]] = {}
        for k in kernel_events:
            cat = _categorize_kernel(k.get("name", ""))
            dur = k.get("dur", 0)
            prev_dur, prev_cnt = breakdown.get(cat, (0.0, 0))
            breakdown[cat] = (prev_dur + dur, prev_cnt + 1)

        print(f"\n{'='*70}")
        print("  Kernel-Only Analysis (no module hierarchy)")
        print(f"{'='*70}")
        print(f"  Total kernels: {len(kernel_events):,}")
        print(f"  Total duration: {total_dur:,.0f} us ({total_dur/1000:,.1f} ms)")
        print(f"\n  {'Category':<20s} {'Duration (us)':>14s} {'Count':>8s} {'%':>7s}")
        print("  " + "-" * 51)
        for cat, (dur, cnt) in sorted(breakdown.items(), key=lambda x: -x[1][0]):
            pct = dur / total_dur * 100 if total_dur > 0 else 0
            print(f"  {cat:<20s} {dur:>14,.0f} {cnt:>8,d} {pct:>6.1f}%")

    @staticmethod
    def _load_trace(path: str) -> Dict[str, Any]:
        if path.endswith(".gz"):
            with gzip.open(path, "rt", encoding="utf-8") as f:
                return json.load(f)
        else:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Trace Module Analyzer — correlation-based GPU kernel classification",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  # Generate an Excel report
  python trace_module_analyzer.py trace.json.gz -o report.xlsx

  # Include detail sheets for up to 10 module types
  python trace_module_analyzer.py trace.json.gz -o report.xlsx --max-detail-modules 10

  # Show kernel-by-kernel detail for a specific module type
  python trace_module_analyzer.py trace.json.gz --detail-module WanTransformerBlock

  # Pick the 5th occurrence instead of the median
  python trace_module_analyzer.py trace.json.gz --detail-module WanTransformerBlock --module-index 5

  # Generate interactive module tree HTML and serve it (auto-generates xlsx if no -o)
  python trace_module_analyzer.py trace.json.gz --model-info
  python trace_module_analyzer.py trace.json.gz -o report.xlsx --model-info --port 9000
""",
    )
    parser.add_argument("trace_file", help="Path to trace file (.json.gz or .json)")
    parser.add_argument("-o", "--output", dest="output", default=None,
                        help="Output Excel report path (.xlsx)")
    parser.add_argument("--max-detail-modules", type=int, default=3,
                        help="Number of module types to generate detail sheets for "
                             "(default: 3, 0=all)")
    parser.add_argument("--detail-module", nargs="+", default=None,
                        help="Specify module types for kernel-by-kernel detail "
                             "(e.g. --detail-module WanTransformerBlock FSDPT5Block)")
    parser.add_argument("--module-index", type=int, default=None,
                        help="Which occurrence of the module to show detail for "
                             "(default: the instance closest to the median)")
    parser.add_argument("--model-info", action="store_true",
                        help="Generate interactive module tree HTML visualization "
                             "and start HTTP server (requires visualize_module_tree.py)")
    parser.add_argument("--port", type=int, default=8765,
                        help="HTTP server port for --model-info (default: 8765)")
    parser.add_argument("--no-rocm-fix", action="store_true",
                        help="Disable automatic ROCm trace fix (hipGraphLaunch flow events)")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Enable debug logging")

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    # Resolve output path: if relative, save to same folder as trace file
    output_path = args.output
    if output_path and not os.path.isabs(output_path):
        trace_dir = os.path.dirname(os.path.abspath(args.trace_file))
        output_path = os.path.join(trace_dir, output_path)

    try:
        analyzer = TraceModuleAnalyzer(
            trace_path=args.trace_file,
            output_path=output_path,
            detail_modules=args.detail_module,
            module_index=args.module_index,
            max_detail_modules=args.max_detail_modules if args.max_detail_modules > 0 else 999,
            auto_fix_rocm=not args.no_rocm_fix,
            model_info=args.model_info,
            port=args.port,
        )
        analyzer.run()
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in trace file: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
