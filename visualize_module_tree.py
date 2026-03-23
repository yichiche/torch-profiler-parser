#!/usr/bin/env python3
"""Interactive Module Tree Visualizer with Architecture Diagram.

Reads an analysis.xlsx (from trace_module_analyzer.py) and generates
a self-contained interactive HTML with two views:
  1. Module Tree — expandable tree-table with timing/breakdown
  2. Architecture — InferenceX-style flowchart with folded layers
"""

import argparse
import json
import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class TreeNode:
    id: int
    name: str
    module_type: str
    color_category: str
    depth: int
    time_us: float
    kernel_count: int
    breakdown: List[Tuple[str, float, float]]
    phase: Optional[str]
    children: List["TreeNode"] = field(default_factory=list)
    pct_of_parent: float = 100.0


# ---------------------------------------------------------------------------
# Module category classification — regex-based, no hardcoded module names
# ---------------------------------------------------------------------------

_CATEGORY_PATTERNS = [
    ("lm_head",       re.compile(r"LogitsProcessor|LMHead|lm_head", re.I)),
    ("sampler",       re.compile(r"^Sampler$", re.I)),
    ("embedding",     re.compile(r"Embed|Embedding|RotaryEmb|VocabParallel", re.I)),
    ("attention",     re.compile(r"Attention|Attn|MLA|MultiHead|DeltaNet|RadixLinear", re.I)),
    ("moe_router",    re.compile(r"Router|Gate(?!Proj)|TopK|MoEGate", re.I)),
    ("expert",        re.compile(r"Expert|MoE(?!Gate)|FusedMoE|SparseMoe", re.I)),
    ("norm",          re.compile(r"Norm|RMSNorm|LayerNorm|BatchNorm", re.I)),
    ("ffn",           re.compile(r"MLP|FFN|FeedForward|SiluAndMul|Activation", re.I)),
    ("projection",    re.compile(
        r"Linear|Proj|ColumnParallel|RowParallel|Merged|QKVParallel|Replicated", re.I)),
    ("infrastructure", re.compile(r"CudaGraph|Replay", re.I)),
]

_CAT_COLORS = {
    "attention": "#f0a500", "projection": "#06d6a0", "expert": "#9b5de5",
    "norm": "#4ea8de", "embedding": "#00c9db", "ffn": "#2ec4b6",
    "moe_router": "#e76f51", "lm_head": "#7c3aed", "sampler": "#6c757d",
    "infrastructure": "#6c757d", "decoder_layer": "#64748b",
    "model_root": "#e0e0e0", "default": "#888",
}

_BD_COLORS = {
    "communication": "#e76f51", "quantization": "#e9c46a", "attention": "#f0a500",
    "gemm": "#76c893", "normalization": "#4ea8de", "elementwise": "#b8b8d0",
    "embedding": "#00c9db", "moe": "#9b5de5", "other": "#6c757d",
}


def classify_module(module_type: str) -> str:
    for cat, pat in _CATEGORY_PATTERNS:
        if pat.search(module_type):
            return cat
    lower = module_type.lower()
    if "decoder" in lower or "layer" in lower:
        return "decoder_layer"
    if "causal" in lower or "model" in lower or "transformer" in lower:
        return "model_root"
    return "default"


def extract_module_type(name: str) -> str:
    m = re.match(r"^(.+?)_(\d+)$", name)
    return m.group(1) if m else name


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

_BREAKDOWN_RE = re.compile(r"(\w+):\s+([\d,]+(?:\.\d+)?)\s+us\s+\(([\d.]+)%\)")


def parse_breakdown(raw: Optional[str]) -> List[Tuple[str, float, float]]:
    if not raw:
        return []
    return [(m.group(1), float(m.group(2).replace(",", "")), float(m.group(3)))
            for m in _BREAKDOWN_RE.finditer(raw)]


def parse_module_entry(raw: str) -> Tuple[int, str]:
    stripped = raw.lstrip()
    depth = (len(raw) - len(stripped)) // 4
    name = stripped
    for prefix in ["├── ", "└── "]:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    return depth, name.strip()


# ---------------------------------------------------------------------------
# Reading Excel tabs
# ---------------------------------------------------------------------------


def read_module_tree(xlsx_path: str) -> List[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Module Tree"]
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        module_raw = row[0]
        if module_raw is None or "truncated" in str(module_raw).lower():
            continue
        rows.append({
            "module_raw": str(module_raw),
            "time_us": float(row[1]) if row[1] is not None else 0.0,
            "kernel_count": int(row[2]) if row[2] is not None else 0,
            "breakdown_raw": row[3],
            "phase": row[4] if len(row) > 4 and row[4] and str(row[4]).strip() else None,
        })
    wb.close()
    return rows


def read_overview(xlsx_path: str) -> Dict[str, dict]:
    """Read Overview tab → {module_type: {count, mean, std, pct_parent, ...}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {}
    try:
        ws = wb["Overview"]
    except KeyError:
        wb.close()
        return result
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row[0] is None:
            continue
        raw_name = str(row[0])
        _, name = parse_module_entry(raw_name)
        mt = extract_module_type(name) if name != "(self)" else None
        if mt is None:
            continue
        result[mt] = {
            "depth": int(row[1]) if row[1] is not None else 0,
            "count": int(row[2]) if row[2] is not None else 0,
            "mean": float(row[3]) if row[3] is not None else 0,
            "std": float(row[4]) if row[4] is not None else 0,
            "min": float(row[5]) if row[5] is not None else 0,
            "max": float(row[6]) if row[6] is not None else 0,
            "total": float(row[7]) if row[7] is not None else 0,
            "pct_parent": float(row[8]) if row[8] is not None else 0,
        }
    wb.close()
    return result


def read_summary_tab(xlsx_path: str) -> Dict[str, dict]:
    """Read Summary tab category section → {category: {count, total, pct}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {}
    try:
        ws = wb["Summary"]
    except KeyError:
        wb.close()
        return result
    in_cats = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Category":
            in_cats = True
            continue
        if in_cats and row[0] is not None:
            result[str(row[0])] = {
                "count": int(row[1]) if row[1] is not None else 0,
                "total": float(row[2]) if row[2] is not None else 0,
                "pct": str(row[4]) if row[4] is not None else "",
            }
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Tree construction
# ---------------------------------------------------------------------------

_node_counter = 0


def _make_node(depth, name, time_us, kernel_count, breakdown_raw, phase):
    global _node_counter
    mt = extract_module_type(name)
    node = TreeNode(id=_node_counter, name=name, module_type=mt,
                    color_category=classify_module(mt), depth=depth,
                    time_us=time_us, kernel_count=kernel_count,
                    breakdown=parse_breakdown(breakdown_raw), phase=phase)
    _node_counter += 1
    return node


def build_tree(rows):
    global _node_counter
    _node_counter = 0
    roots, stack = [], []
    for row in rows:
        depth, name = parse_module_entry(row["module_raw"])
        node = _make_node(depth, name, row["time_us"], row["kernel_count"],
                          row["breakdown_raw"], row["phase"])
        while stack and stack[-1][0] >= depth:
            stack.pop()
        if stack:
            stack[-1][1].children.append(node)
        else:
            roots.append(node)
        stack.append((depth, node))
    return roots


def compute_metrics(roots):
    max_time = 0.0
    def walk(node, parent_time):
        nonlocal max_time
        if node.time_us > max_time:
            max_time = node.time_us
        node.pct_of_parent = round(node.time_us / parent_time * 100, 1) if parent_time > 0 else 0.0
        for c in node.children:
            walk(c, node.time_us)
    for r in roots:
        r.pct_of_parent = 100.0
        if r.time_us > max_time:
            max_time = r.time_us
        for c in r.children:
            walk(c, r.time_us)
    return max_time


# ---------------------------------------------------------------------------
# Architecture tree — fold repeated patterns
# ---------------------------------------------------------------------------


def _detect_repeating_pattern(type_seq: List[str]):
    """Find the best repeating pattern that covers the most elements.

    Tries starting from each possible offset to handle header elements
    (e.g., VocabParallelEmbedding before the decoder layers).

    Returns (pattern, repeats, start_idx, end_idx) or None.
    """
    n = len(type_seq)
    best = None  # (coverage, pattern, repeats, start_idx, end_idx)
    # Try different start offsets (header items before the pattern)
    for start in range(min(n // 2, 5)):  # max 5 header items
        sub = type_seq[start:]
        sub_n = len(sub)
        for plen in range(1, sub_n // 2 + 1):
            pattern = sub[:plen]
            repeats = 0
            for i in range(0, sub_n, plen):
                if sub[i:i + plen] == pattern:
                    repeats += 1
                else:
                    break
            if repeats >= 2:
                coverage = repeats * plen
                if best is None or coverage > best[0]:
                    best = (coverage, pattern, repeats, start, start + coverage)
    if best:
        return best[1], best[2], best[3], best[4]
    return None


def _make_block(child: TreeNode, overview: dict, count: int = 1) -> dict:
    mt = child.module_type
    return {
        "type": "block", "moduleType": mt,
        "category": classify_module(mt), "countInUnit": count,
        "timeUs": child.time_us, "kernelCount": child.kernel_count,
        "overview": overview.get(mt, {}),
        "children": _fold_children(child.children, overview),
    }


def _fold_children(children: List[TreeNode], overview: dict) -> List[dict]:
    """Recursively fold children into arch blocks."""
    if not children:
        return []

    types = [c.module_type for c in children]
    pat_result = _detect_repeating_pattern(types)

    blocks = []
    if pat_result:
        pattern, repeats, start_idx, end_idx = pat_result
        plen = len(pattern)

        # Header blocks (before the repeating pattern)
        for child in children[:start_idx]:
            blocks.append(_make_block(child, overview))

        # Build the pattern unit items (deduplicated)
        pattern_items = []
        seen = set()
        for offset in range(plen):
            mt = pattern[offset]
            if mt not in seen:
                seen.add(mt)
                count_in_unit = sum(1 for t in pattern if t == mt)
                rep_node = children[start_idx + offset]
                pattern_items.append({
                    "type": "block", "moduleType": mt,
                    "category": classify_module(mt),
                    "countInUnit": count_in_unit,
                    "timeUs": rep_node.time_us,
                    "kernelCount": rep_node.kernel_count,
                    "overview": overview.get(mt, {}),
                    "children": _fold_children(rep_node.children, overview),
                })

        blocks.append({
            "type": "group",
            "repeats": repeats,
            "patternTypes": pattern,
            "items": pattern_items,
        })

        # Remainder blocks (after the repeating pattern)
        for child in children[end_idx:]:
            blocks.append(_make_block(child, overview))
    else:
        # No repeating pattern — group consecutive same-type runs
        i = 0
        while i < len(children):
            j = i + 1
            while j < len(children) and types[j] == types[i]:
                j += 1
            blocks.append(_make_block(children[i], overview, count=j - i))
            i = j

    return blocks


def build_arch_data(roots: List[TreeNode], overview: dict) -> dict:
    """Build the architecture diagram data from the tree roots."""
    # Find the main model root: prefer model_root category, then most children
    main_root = None
    header_blocks = []
    footer_blocks = []

    # First pass: find the main model root
    for r in roots:
        cat = classify_module(r.module_type)
        if cat == "model_root" and r.children:
            if main_root is None or len(r.children) > len(main_root.children):
                main_root = r
    # Fallback: largest root by child count (excluding infrastructure)
    if main_root is None:
        for r in roots:
            cat = classify_module(r.module_type)
            if r.children and cat not in ("infrastructure", "embedding", "lm_head", "sampler"):
                if main_root is None or len(r.children) > len(main_root.children):
                    main_root = r

    # Second pass: categorize remaining roots
    for r in roots:
        if r is main_root:
            continue
        cat = classify_module(r.module_type)
        ov = overview.get(r.module_type, {})
        block = {
            "type": "block", "moduleType": r.module_type,
            "category": cat, "countInUnit": 1,
            "timeUs": r.time_us, "kernelCount": r.kernel_count,
            "overview": ov,
            "children": _fold_children(r.children, overview) if r.children else [],
        }
        if cat == "embedding":
            header_blocks.append(block)
        else:
            footer_blocks.append(block)

    model_blocks = []
    model_name = ""
    if main_root:
        model_name = main_root.module_type
        model_blocks = _fold_children(main_root.children, overview)

    return {
        "modelName": model_name,
        "modelTimeUs": main_root.time_us if main_root else 0,
        "header": header_blocks,
        "body": model_blocks,
        "footer": footer_blocks,
        "catColors": _CAT_COLORS,
    }


# ---------------------------------------------------------------------------
# JSON serialization
# ---------------------------------------------------------------------------


def nodes_to_json(roots):
    def serialize(node):
        return {
            "id": node.id, "name": node.name,
            "moduleType": node.module_type,
            "colorCategory": node.color_category,
            "depth": node.depth, "timeUs": node.time_us,
            "kernelCount": node.kernel_count,
            "breakdown": [{"cat": c, "timeUs": t, "pct": p} for c, t, p in node.breakdown],
            "phase": node.phase, "pctOfParent": node.pct_of_parent,
            "children": [serialize(c) for c in node.children],
        }
    return [serialize(r) for r in roots]


def compute_summary(roots):
    total_time = sum(r.time_us for r in roots)
    total_kernels = sum(r.kernel_count for r in roots)
    count = [0]
    def walk(n):
        count[0] += 1
        for c in n.children:
            walk(c)
    for r in roots:
        walk(r)
    return {"totalTime": round(total_time, 1), "totalKernels": total_kernels,
            "moduleCount": count[0]}


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------


def generate_html(roots, max_time_us, basename, overview, summary_tab):
    tree_json = json.dumps(nodes_to_json(roots), separators=(",", ":"))
    summary = json.dumps(compute_summary(roots), separators=(",", ":"))
    arch_data = json.dumps(build_arch_data(roots, overview), separators=(",", ":"))
    cat_colors_json = json.dumps(_CAT_COLORS, separators=(",", ":"))
    bd_colors_json = json.dumps(_BD_COLORS, separators=(",", ":"))

    return (HTML_TEMPLATE
            .replace("__TREE_DATA__", tree_json)
            .replace("__MAX_TIME__", str(max_time_us))
            .replace("__BASENAME__", basename)
            .replace("__SUMMARY__", summary)
            .replace("__ARCH_DATA__", arch_data)
            .replace("__CAT_COLORS__", cat_colors_json)
            .replace("__BD_COLORS__", bd_colors_json))


# ---------------------------------------------------------------------------
# HTML Template
# ---------------------------------------------------------------------------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Module Tree - __BASENAME__</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg-primary:#1a1a2e;--bg-secondary:#16213e;--bg-toolbar:#0f3460;
  --bg-row-hover:#1e2a4a;--text-primary:#e0e0e0;--text-secondary:#a0a0b0;
  --text-dim:#666;--border-color:#2a2a4a;
}
body{background:var(--bg-primary);color:var(--text-primary);
  font-family:'Inter','Segoe UI',system-ui,-apple-system,sans-serif;
  font-size:13px;line-height:1.4;}
.app{display:flex;flex-direction:column;height:100vh}

/* Tab bar */
.tab-bar{display:flex;background:var(--bg-secondary);border-bottom:2px solid var(--border-color)}
.tab-btn{padding:10px 24px;cursor:pointer;font-size:13px;font-weight:600;
  color:var(--text-secondary);border:none;background:none;
  border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .15s}
.tab-btn:hover{color:var(--text-primary);background:rgba(255,255,255,0.03)}
.tab-btn.active{color:#4ea8de;border-bottom-color:#4ea8de}
.tab-page{display:none;flex:1;overflow:hidden;flex-direction:column}
.tab-page.active{display:flex}

/* Toolbar */
.toolbar{background:var(--bg-toolbar);padding:10px 16px;
  display:flex;flex-wrap:wrap;align-items:center;gap:10px;
  border-bottom:1px solid var(--border-color)}
.toolbar h1{font-size:14px;font-weight:600;margin-right:12px;white-space:nowrap}
.toolbar button{background:rgba(255,255,255,0.1);color:var(--text-primary);
  border:1px solid var(--border-color);border-radius:6px;
  padding:4px 10px;cursor:pointer;font-size:12px;transition:background .15s}
.toolbar button:hover{background:rgba(255,255,255,0.2)}
.toolbar button.active{background:rgba(78,168,222,0.3);border-color:#4ea8de}
.toolbar input[type=text]{background:rgba(255,255,255,0.08);color:var(--text-primary);
  border:1px solid var(--border-color);border-radius:6px;
  padding:4px 10px;font-size:12px;width:180px;outline:none}
.toolbar input[type=text]:focus{border-color:#4ea8de}
.toolbar select{background:rgba(255,255,255,0.08);color:var(--text-primary);
  border:1px solid var(--border-color);border-radius:6px;padding:4px 8px;font-size:12px;outline:none}
.toolbar label{font-size:12px;display:flex;align-items:center;gap:4px;cursor:pointer}
.toolbar .sep{width:1px;height:20px;background:var(--border-color)}
.ctrl-group{display:flex;align-items:center;gap:6px}
.ctrl-label{font-size:11px;color:var(--text-secondary);text-transform:uppercase;letter-spacing:.5px}

/* Summary */
.summary-bar{background:var(--bg-secondary);padding:8px 16px;font-size:12px;
  color:var(--text-secondary);border-bottom:1px solid var(--border-color);display:flex;gap:24px}
.summary-bar span{display:flex;align-items:center;gap:4px}
.summary-bar .val{color:var(--text-primary);font-weight:600}

/* Column headers & tree rows */
.col-headers{display:grid;grid-template-columns:1fr 100px 80px 70px 220px;
  padding:6px 16px 6px 40px;background:var(--bg-secondary);
  border-bottom:2px solid var(--border-color);font-size:11px;font-weight:600;
  text-transform:uppercase;letter-spacing:.5px;color:var(--text-secondary)}
.col-headers span:nth-child(n+2){text-align:right}
.col-headers span:last-child{text-align:left;padding-left:8px}
.tree-container{flex:1;overflow-y:auto;overflow-x:hidden}
.tree-row{position:relative;display:grid;grid-template-columns:1fr 100px 80px 70px 220px;
  align-items:center;min-height:32px;padding:0 16px;
  border-bottom:1px solid rgba(42,42,74,0.5);cursor:pointer;transition:background .12s}
.tree-row:hover{background:var(--bg-row-hover)}
.time-bar{position:absolute;left:0;top:0;bottom:0;opacity:.10;border-radius:0 4px 4px 0;
  pointer-events:none;z-index:0;min-width:2px}
.cell-module{display:flex;align-items:center;gap:6px;overflow:hidden;position:relative;z-index:1}
.toggle{width:16px;height:16px;display:inline-flex;align-items:center;justify-content:center;
  font-size:10px;flex-shrink:0;color:var(--text-secondary);transition:transform .15s;user-select:none}
.toggle.expanded{transform:rotate(90deg)}.toggle.leaf{visibility:hidden}
.color-dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.module-name{overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:13px}
.phase-badge{font-size:9px;padding:1px 6px;border-radius:8px;font-weight:600;
  text-transform:uppercase;flex-shrink:0;margin-left:4px}
.phase-prefill{background:rgba(46,196,182,0.25);color:#2ec4b6}
.phase-decode{background:rgba(240,165,0,0.25);color:#f0a500}
.cell-time,.cell-pct,.cell-kernels{text-align:right;font-variant-numeric:tabular-nums;
  position:relative;z-index:1;font-size:12px}
.cell-pct{color:var(--text-secondary)}
.cell-breakdown{position:relative;z-index:1;padding-left:8px}
.mini-bar{display:flex;height:14px;border-radius:3px;overflow:hidden;width:200px;
  background:rgba(255,255,255,0.04)}
.mini-seg{height:100%;min-width:1px}

/* Tooltip */
.tooltip{display:none;position:fixed;z-index:1000;background:#0d1b2a;
  border:1px solid var(--border-color);border-radius:8px;padding:12px;max-width:380px;
  font-size:12px;pointer-events:none;box-shadow:0 8px 32px rgba(0,0,0,0.5)}
.tooltip .tt-name{font-weight:600;font-size:13px;margin-bottom:4px}
.tooltip .tt-phase{font-size:11px;margin-bottom:4px}
.tooltip .tt-metrics{margin-bottom:8px}
.tooltip .tt-sep{border-top:1px solid var(--border-color);margin:6px 0}
.tooltip .tt-bd-row{display:flex;align-items:center;gap:6px;margin:2px 0}
.tooltip .tt-bd-dot{width:8px;height:8px;border-radius:2px;flex-shrink:0}
.tooltip .tt-bd-cat{flex:1}
.tooltip .tt-bd-val{color:var(--text-secondary);text-align:right}

/* Legend */
.legend{background:var(--bg-secondary);padding:8px 16px;border-top:1px solid var(--border-color);
  display:flex;flex-wrap:wrap;gap:12px;font-size:11px}
.legend-item{display:flex;align-items:center;gap:4px}
.legend-dot{width:8px;height:8px;border-radius:50%}
.no-results{padding:40px;text-align:center;color:var(--text-secondary);font-size:14px}

/* ===== Architecture Diagram ===== */
.arch-container{flex:1;overflow-y:auto;padding:32px 16px;
  display:flex;flex-direction:column;align-items:center}
.arch-flow{display:flex;flex-direction:column;align-items:center;gap:0;max-width:700px;width:100%}

/* Arrow connector */
.arch-arrow{width:2px;height:24px;background:var(--border-color);position:relative}
.arch-arrow::after{content:'';position:absolute;bottom:-4px;left:-4px;
  border-left:5px solid transparent;border-right:5px solid transparent;
  border-top:6px solid var(--border-color)}

/* Block */
.arch-block{width:100%;border-radius:10px;padding:14px 20px;text-align:center;
  border:2px solid;cursor:pointer;transition:transform .12s,box-shadow .12s;position:relative}
.arch-block:hover{transform:translateY(-1px);box-shadow:0 4px 20px rgba(0,0,0,0.3)}
.arch-block-name{font-weight:600;font-size:14px;color:#fff}
.arch-block-sub{font-size:11px;color:rgba(255,255,255,0.6);margin-top:2px}
.arch-block-count{position:absolute;top:-8px;right:12px;font-size:10px;font-weight:700;
  background:var(--bg-primary);color:var(--text-secondary);padding:1px 8px;
  border-radius:10px;border:1px solid var(--border-color)}

/* Group container */
.arch-group{width:100%;border:2px dashed var(--border-color);border-radius:14px;
  padding:20px 16px 16px;position:relative}
.arch-group-label{position:absolute;top:-12px;left:16px;font-size:12px;font-weight:600;
  background:var(--bg-primary);color:var(--text-secondary);padding:2px 10px;border-radius:8px;
  border:1px solid var(--border-color)}
.arch-group-inner{display:flex;flex-direction:column;align-items:center;gap:0}
.arch-group-toggle{position:absolute;top:-12px;right:16px;width:24px;height:24px;
  border-radius:50%;background:var(--bg-secondary);border:1px solid var(--border-color);
  color:var(--text-primary);font-size:14px;cursor:pointer;
  display:flex;align-items:center;justify-content:center;transition:background .15s}
.arch-group-toggle:hover{background:rgba(78,168,222,0.3)}
.arch-group.collapsed .arch-group-inner{display:none}
.arch-group.collapsed{padding-bottom:12px}

/* Sub-blocks inside groups */
.arch-sub-block{width:90%;border-radius:8px;padding:10px 16px;text-align:center;
  border:1.5px solid;cursor:pointer;transition:transform .1s}
.arch-sub-block:hover{transform:translateY(-1px)}
.arch-sub-name{font-weight:600;font-size:13px;color:#fff}
.arch-sub-info{font-size:10px;color:rgba(255,255,255,0.5);margin-top:1px}
.arch-sub-count{position:absolute;top:-6px;right:8px;font-size:9px;font-weight:700;
  background:var(--bg-primary);color:var(--text-secondary);padding:0 6px;
  border-radius:8px;border:1px solid var(--border-color)}
.arch-sub-block{position:relative}

/* Arch tooltip */
.arch-tooltip{display:none;position:fixed;z-index:1000;background:#0d1b2a;
  border:1px solid var(--border-color);border-radius:8px;padding:12px;
  max-width:350px;font-size:12px;pointer-events:none;box-shadow:0 8px 32px rgba(0,0,0,0.5)}
</style>
</head>
<body>
<div class="app">
  <!-- Tab bar -->
  <div class="tab-bar">
    <button class="tab-btn active" data-tab="tree">Module Tree</button>
    <button class="tab-btn" data-tab="arch">Architecture</button>
  </div>

  <!-- ===== Tree Page ===== -->
  <div class="tab-page active" id="page-tree">
    <div class="toolbar">
      <h1>__BASENAME__</h1>
      <div class="sep"></div>
      <div class="ctrl-group">
        <button id="btn-expand-all">Expand All</button>
        <button id="btn-collapse-all">Collapse All</button>
      </div>
      <div class="sep"></div>
      <div class="ctrl-group">
        <span class="ctrl-label">Depth</span>
        <button class="depth-btn" data-depth="0">0</button>
        <button class="depth-btn" data-depth="1">1</button>
        <button class="depth-btn" data-depth="2">2</button>
        <button class="depth-btn" data-depth="3">3</button>
        <button class="depth-btn" data-depth="4">4</button>
      </div>
      <div class="sep"></div>
      <div class="ctrl-group">
        <span class="ctrl-label">Search</span>
        <input type="text" id="search-input" placeholder="Filter modules...">
      </div>
      <div class="sep"></div>
      <div class="ctrl-group">
        <span class="ctrl-label">Sort</span>
        <select id="sort-select">
          <option value="original">Original</option>
          <option value="time-desc">Time ↓</option>
          <option value="time-asc">Time ↑</option>
          <option value="kernels-desc">Kernels ↓</option>
        </select>
      </div>
      <div class="sep"></div>
      <div class="ctrl-group">
        <label><input type="checkbox" class="phase-cb" value="prefill" checked> Prefill</label>
        <label><input type="checkbox" class="phase-cb" value="decode" checked> Decode</label>
        <label><input type="checkbox" class="phase-cb" value="" checked> Other</label>
      </div>
    </div>
    <div class="summary-bar" id="summary-bar"></div>
    <div class="col-headers">
      <span>Module</span><span>Time (μs)</span><span>% Parent</span>
      <span>Kernels</span><span style="padding-left:8px">Breakdown</span>
    </div>
    <div class="tree-container" id="tree-container"></div>
    <div class="legend" id="legend"></div>
  </div>

  <!-- ===== Architecture Page ===== -->
  <div class="tab-page" id="page-arch">
    <div class="arch-container">
      <div class="arch-flow" id="arch-flow"></div>
    </div>
  </div>
</div>
<div class="tooltip" id="tooltip"></div>

<script>
const TREE_DATA = __TREE_DATA__;
const MAX_TIME = __MAX_TIME__;
const SUMMARY = __SUMMARY__;
const ARCH_DATA = __ARCH_DATA__;
const CAT_COLORS = __CAT_COLORS__;
const BD_COLORS = __BD_COLORS__;

function getCatColor(cat) { return CAT_COLORS[cat] || CAT_COLORS['default'] || '#888'; }
function getBdColor(cat) { return BD_COLORS[cat] || BD_COLORS['other'] || '#6c757d'; }
function fmtNum(n) {
  if (n == null) return '-';
  return n.toLocaleString('en-US', {maximumFractionDigits:1});
}
function esc(s) { const d=document.createElement('div'); d.textContent=s; return d.innerHTML; }

// ===== Tab switching =====
document.querySelectorAll('.tab-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.tab-page').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById('page-' + btn.dataset.tab).classList.add('active');
    if (btn.dataset.tab === 'arch' && !archRendered) { renderArch(); archRendered = true; }
  });
});
// Hash routing
if (location.hash === '#arch') {
  document.querySelector('[data-tab="arch"]').click();
}

// ===== TREE VIEW (unchanged logic) =====
let expandedNodes = new Set();
let allNodes = [], nodeMap = {};
let sortMode = 'original', searchFilter = '';
let phaseFilter = {prefill:true,decode:true,'':true};
let activeDepth = -1;

function flattenTree(roots, parentId) {
  for (const node of sortChildren(roots)) {
    node.parentId = parentId;
    allNodes.push(node); nodeMap[node.id] = node;
    if (node.children && node.children.length) flattenTree(node.children, node.id);
  }
}
function sortChildren(ch) {
  if (sortMode === 'original') return ch;
  const s = [...ch];
  if (sortMode === 'time-desc') s.sort((a,b) => b.timeUs - a.timeUs);
  if (sortMode === 'time-asc') s.sort((a,b) => a.timeUs - b.timeUs);
  if (sortMode === 'kernels-desc') s.sort((a,b) => b.kernelCount - a.kernelCount);
  return s;
}
function rebuildFlat() { allNodes=[]; nodeMap={}; flattenTree(TREE_DATA, null); }
function isVisible(node) {
  let cur = node;
  while (cur.parentId != null) { if (!expandedNodes.has(cur.parentId)) return false; cur = nodeMap[cur.parentId]; }
  return true;
}
function getVisibleSet() {
  if (!searchFilter) return null;
  const vis = new Set();
  for (const n of allNodes) {
    if (n.name.toLowerCase().includes(searchFilter)) {
      vis.add(n.id);
      let cur = n;
      while (cur.parentId != null) { vis.add(cur.parentId); expandedNodes.add(cur.parentId); cur = nodeMap[cur.parentId]; }
    }
  }
  return vis;
}
function matchesPhase(node) { return phaseFilter[node.phase || ''] !== false; }

function createRow(node) {
  const row = document.createElement('div');
  row.className = 'tree-row'; row.dataset.id = node.id;
  const barW = MAX_TIME > 0 ? (node.timeUs / MAX_TIME) * 100 : 0;
  const bar = document.createElement('div');
  bar.className = 'time-bar';
  bar.style.width = Math.max(barW, node.timeUs > 0 ? 0.15 : 0) + '%';
  bar.style.backgroundColor = getCatColor(node.colorCategory);
  row.appendChild(bar);
  const modCell = document.createElement('div');
  modCell.className = 'cell-module';
  modCell.style.paddingLeft = (node.depth * 24) + 'px';
  const hasKids = node.children && node.children.length > 0;
  const toggle = document.createElement('span');
  toggle.className = 'toggle' + (hasKids ? (expandedNodes.has(node.id) ? ' expanded' : '') : ' leaf');
  toggle.textContent = '\u25b6'; modCell.appendChild(toggle);
  const dot = document.createElement('span');
  dot.className = 'color-dot'; dot.style.backgroundColor = getCatColor(node.colorCategory);
  modCell.appendChild(dot);
  const nm = document.createElement('span'); nm.className = 'module-name'; nm.textContent = node.name;
  modCell.appendChild(nm);
  if (node.phase) {
    const badge = document.createElement('span');
    badge.className = 'phase-badge phase-' + node.phase; badge.textContent = node.phase;
    modCell.appendChild(badge);
  }
  row.appendChild(modCell);
  const tc = document.createElement('div'); tc.className = 'cell-time'; tc.textContent = fmtNum(node.timeUs);
  row.appendChild(tc);
  const pc = document.createElement('div'); pc.className = 'cell-pct';
  pc.textContent = node.pctOfParent < 100 ? node.pctOfParent.toFixed(1) + '%' : '';
  row.appendChild(pc);
  const kc = document.createElement('div'); kc.className = 'cell-kernels'; kc.textContent = node.kernelCount;
  row.appendChild(kc);
  const bdCell = document.createElement('div'); bdCell.className = 'cell-breakdown';
  if (node.breakdown && node.breakdown.length) {
    const mb = document.createElement('div'); mb.className = 'mini-bar';
    for (const seg of node.breakdown) {
      const s = document.createElement('div'); s.className = 'mini-seg';
      s.style.width = seg.pct + '%'; s.style.backgroundColor = getBdColor(seg.cat);
      s.title = seg.cat + ': ' + fmtNum(seg.timeUs) + ' us (' + seg.pct + '%)';
      mb.appendChild(s);
    }
    bdCell.appendChild(mb);
  }
  row.appendChild(bdCell);
  return row;
}

function renderTree() {
  const container = document.getElementById('tree-container');
  const searchSet = getVisibleSet();
  const frag = document.createDocumentFragment();
  let count = 0;
  for (const node of allNodes) {
    if (!matchesPhase(node)) continue;
    if (searchSet !== null) { if (!searchSet.has(node.id)) continue; }
    else { if (!isVisible(node)) continue; }
    frag.appendChild(createRow(node)); count++;
  }
  container.innerHTML = '';
  if (count === 0) {
    const nr = document.createElement('div'); nr.className = 'no-results';
    nr.textContent = 'No modules match the current filters.'; container.appendChild(nr);
  } else container.appendChild(frag);
}

function toggleNode(id) {
  if (expandedNodes.has(id)) { expandedNodes.delete(id); collapseDesc(id); }
  else expandedNodes.add(id);
  activeDepth = -1; updateDepthBtns(); renderTree();
}
function collapseDesc(id) {
  const n = nodeMap[id]; if (!n||!n.children) return;
  for (const c of n.children) { expandedNodes.delete(c.id); collapseDesc(c.id); }
}
function expandToDepth(d) {
  expandedNodes.clear();
  for (const n of allNodes) { if (n.depth < d && n.children && n.children.length) expandedNodes.add(n.id); }
  activeDepth = d; updateDepthBtns(); renderTree();
}
function updateDepthBtns() {
  document.querySelectorAll('.depth-btn').forEach(b => b.classList.toggle('active', parseInt(b.dataset.depth) === activeDepth));
}

// Tooltip
const tooltip = document.getElementById('tooltip');
function showTooltip(e, node) {
  let html = '<div class="tt-name">' + esc(node.name) + '</div>';
  if (node.phase) html += '<div class="tt-phase">Phase: <strong>' + node.phase + '</strong></div>';
  html += '<div class="tt-metrics">Time: <strong>' + fmtNum(node.timeUs) + ' \u03bcs</strong>';
  if (node.pctOfParent < 100) html += ' (' + node.pctOfParent.toFixed(1) + '% of parent)';
  html += '<br>Kernels: <strong>' + node.kernelCount + '</strong></div>';
  if (node.breakdown && node.breakdown.length) {
    html += '<div class="tt-sep"></div>';
    for (const seg of node.breakdown) {
      html += '<div class="tt-bd-row"><span class="tt-bd-dot" style="background:' + getBdColor(seg.cat) + '"></span>' +
        '<span class="tt-bd-cat">' + seg.cat + '</span><span class="tt-bd-val">' +
        fmtNum(seg.timeUs) + ' \u03bcs (' + seg.pct + '%)</span></div>';
    }
  }
  tooltip.innerHTML = html; tooltip.style.display = 'block'; posTooltip(e);
}
function posTooltip(e) {
  const r = tooltip.getBoundingClientRect();
  let x = e.clientX+16, y = e.clientY-8;
  if (x+r.width > window.innerWidth-8) x = e.clientX-r.width-16;
  if (y+r.height > window.innerHeight-8) y = window.innerHeight-r.height-8;
  if (y < 8) y = 8;
  tooltip.style.left = x+'px'; tooltip.style.top = y+'px';
}
function hideTooltip() { tooltip.style.display = 'none'; }

function renderSummary() {
  document.getElementById('summary-bar').innerHTML =
    '<span>Total: <span class="val">' + fmtNum(SUMMARY.totalTime) + ' \u03bcs</span></span>' +
    '<span>Modules: <span class="val">' + SUMMARY.moduleCount + '</span></span>' +
    '<span>Kernels: <span class="val">' + fmtNum(SUMMARY.totalKernels) + '</span></span>';
}
function renderLegend() {
  const lg = document.getElementById('legend');
  const cats = Object.entries(CAT_COLORS).filter(([k]) => k !== 'default');
  lg.innerHTML = cats.map(([k,v]) =>
    '<span class="legend-item"><span class="legend-dot" style="background:'+v+'"></span>'+k+'</span>'
  ).join('');
}
function debounce(fn, ms) { let t; return (...a) => { clearTimeout(t); t = setTimeout(() => fn(...a), ms); }; }

// ===== ARCHITECTURE VIEW =====
let archRendered = false;

function renderArch() {
  const flow = document.getElementById('arch-flow');
  flow.innerHTML = '';
  // Header blocks
  for (const b of ARCH_DATA.header) { appendArchBlock(flow, b, false); addArrow(flow); }
  // Model name label
  if (ARCH_DATA.modelName) {
    const label = document.createElement('div');
    label.style.cssText = 'font-size:16px;font-weight:700;color:var(--text-primary);margin:8px 0;text-align:center';
    label.textContent = ARCH_DATA.modelName;
    const sub = document.createElement('div');
    sub.style.cssText = 'font-size:12px;color:var(--text-secondary);margin-bottom:8px';
    sub.textContent = fmtNum(ARCH_DATA.modelTimeUs) + ' \u03bcs total';
    flow.appendChild(label); flow.appendChild(sub);
  }
  // Body
  for (let i = 0; i < ARCH_DATA.body.length; i++) {
    const item = ARCH_DATA.body[i];
    if (item.type === 'group') renderArchGroup(flow, item);
    else appendArchBlock(flow, item, false);
    if (i < ARCH_DATA.body.length - 1) addArrow(flow);
  }
  // Footer
  for (const b of ARCH_DATA.footer) { addArrow(flow); appendArchBlock(flow, b, false); }
}

function addArrow(parent) {
  const a = document.createElement('div'); a.className = 'arch-arrow'; parent.appendChild(a);
}

function appendArchBlock(parent, block, isSub) {
  const el = document.createElement('div');
  el.className = isSub ? 'arch-sub-block' : 'arch-block';
  const color = getCatColor(block.category);
  el.style.borderColor = color;
  el.style.background = hexToRgba(color, 0.15);
  const nm = document.createElement('div');
  nm.className = isSub ? 'arch-sub-name' : 'arch-block-name';
  nm.textContent = block.moduleType;
  el.appendChild(nm);
  // Subtitle: timing info
  const ov = block.overview || {};
  let subText = '';
  if (ov.mean) subText = fmtNum(ov.mean) + ' \u03bcs' + (ov.std ? ' \u00b1 ' + fmtNum(ov.std) : '');
  else if (block.timeUs) subText = fmtNum(block.timeUs) + ' \u03bcs';
  if (ov.pct_parent) subText += ' (' + ov.pct_parent + '%)';
  if (subText) {
    const sub = document.createElement('div');
    sub.className = isSub ? 'arch-sub-info' : 'arch-block-sub';
    sub.textContent = subText; el.appendChild(sub);
  }
  // Count badge
  if (block.countInUnit > 1) {
    const badge = document.createElement('span');
    badge.className = isSub ? 'arch-sub-count' : 'arch-block-count';
    badge.textContent = '\u00d7' + block.countInUnit;
    el.appendChild(badge);
  }
  // Expandable children
  if (block.children && block.children.length && !isSub) {
    el.style.cursor = 'pointer';
    el.addEventListener('click', (e) => {
      e.stopPropagation();
      const existing = el.querySelector('.arch-children');
      if (existing) { existing.remove(); return; }
      const ch = document.createElement('div');
      ch.className = 'arch-children';
      ch.style.cssText = 'margin-top:8px;display:flex;flex-direction:column;align-items:center;gap:0';
      for (let i = 0; i < block.children.length; i++) {
        appendArchBlock(ch, block.children[i], true);
        if (i < block.children.length - 1) { const a = document.createElement('div'); a.className='arch-arrow'; a.style.height='16px'; ch.appendChild(a); }
      }
      el.appendChild(ch);
    });
  }
  // Tooltip on hover
  el.addEventListener('mouseenter', (e) => {
    let html = '<div class="tt-name">' + esc(block.moduleType) + '</div>';
    if (block.countInUnit > 1) html += '<div style="color:var(--text-secondary);">\u00d7' + block.countInUnit + ' instances</div>';
    html += '<div class="tt-metrics">';
    if (ov.count) html += 'Total instances: ' + ov.count + '<br>';
    if (ov.mean) html += 'Mean: ' + fmtNum(ov.mean) + ' \u03bcs \u00b1 ' + fmtNum(ov.std || 0) + '<br>';
    if (ov.min != null && ov.max != null) html += 'Range: ' + fmtNum(ov.min) + ' \u2013 ' + fmtNum(ov.max) + ' \u03bcs<br>';
    if (ov.total) html += 'Total: ' + fmtNum(ov.total) + ' \u03bcs<br>';
    if (ov.pct_parent) html += '% of parent: ' + ov.pct_parent + '%';
    html += '</div>';
    tooltip.innerHTML = html; tooltip.style.display = 'block'; posTooltip(e);
  });
  el.addEventListener('mouseleave', hideTooltip);
  el.addEventListener('mousemove', posTooltip);
  parent.appendChild(el);
}

function renderArchGroup(parent, group) {
  const container = document.createElement('div');
  container.className = 'arch-group';
  // Label
  const patDesc = summarizePattern(group.patternTypes);
  const label = document.createElement('div');
  label.className = 'arch-group-label';
  label.textContent = '\u00d7' + group.repeats + ' ' + patDesc;
  container.appendChild(label);
  // Toggle button
  const toggleBtn = document.createElement('button');
  toggleBtn.className = 'arch-group-toggle';
  toggleBtn.textContent = '\u2212';
  toggleBtn.addEventListener('click', (e) => {
    e.stopPropagation();
    container.classList.toggle('collapsed');
    toggleBtn.textContent = container.classList.contains('collapsed') ? '+' : '\u2212';
  });
  container.appendChild(toggleBtn);
  // Inner blocks
  const inner = document.createElement('div');
  inner.className = 'arch-group-inner';
  for (let i = 0; i < group.items.length; i++) {
    appendArchBlock(inner, group.items[i], false);
    if (i < group.items.length - 1) addArrow(inner);
  }
  container.appendChild(inner);
  parent.appendChild(container);
}

function summarizePattern(types) {
  // Count consecutive runs: [L,L,L,A] → "3×L + 1×A"
  const runs = [];
  let i = 0;
  while (i < types.length) {
    let j = i + 1;
    while (j < types.length && types[j] === types[i]) j++;
    const shortName = types[i].replace(/^.*?(\w+)$/, '$1');
    runs.push((j - i > 1 ? (j - i) + '\u00d7' : '') + shortName);
    i = j;
  }
  return runs.join(' + ');
}

function hexToRgba(hex, alpha) {
  const r = parseInt(hex.slice(1,3),16), g = parseInt(hex.slice(3,5),16), b = parseInt(hex.slice(5,7),16);
  return 'rgba('+r+','+g+','+b+','+alpha+')';
}

// ===== Init =====
function init() {
  rebuildFlat();
  for (const n of allNodes) { if (n.depth === 0 && n.children && n.children.length) expandedNodes.add(n.id); }
  activeDepth = 1; updateDepthBtns(); renderSummary(); renderLegend(); renderTree();

  document.getElementById('tree-container').addEventListener('click', e => {
    const row = e.target.closest('.tree-row'); if (!row) return;
    const id = parseInt(row.dataset.id), node = nodeMap[id];
    if (node && node.children && node.children.length) toggleNode(id);
  });
  document.getElementById('tree-container').addEventListener('mouseover', e => {
    const row = e.target.closest('.tree-row');
    if (row) showTooltip(e, nodeMap[parseInt(row.dataset.id)]);
  });
  document.getElementById('tree-container').addEventListener('mousemove', e => {
    if (tooltip.style.display === 'block') posTooltip(e);
  });
  document.getElementById('tree-container').addEventListener('mouseout', e => {
    const row = e.target.closest('.tree-row');
    if (row && !row.contains(e.relatedTarget)) hideTooltip();
  });
  document.getElementById('btn-expand-all').addEventListener('click', () => {
    for (const n of allNodes) { if (n.children && n.children.length) expandedNodes.add(n.id); }
    activeDepth = -1; updateDepthBtns(); renderTree();
  });
  document.getElementById('btn-collapse-all').addEventListener('click', () => {
    expandedNodes.clear(); activeDepth = 0; updateDepthBtns(); renderTree();
  });
  document.querySelectorAll('.depth-btn').forEach(btn => {
    btn.addEventListener('click', () => expandToDepth(parseInt(btn.dataset.depth)));
  });
  const debouncedSearch = debounce(() => {
    searchFilter = document.getElementById('search-input').value.toLowerCase().trim(); renderTree();
  }, 300);
  document.getElementById('search-input').addEventListener('input', debouncedSearch);
  document.getElementById('sort-select').addEventListener('change', e => {
    sortMode = e.target.value; rebuildFlat(); renderTree();
  });
  document.querySelectorAll('.phase-cb').forEach(cb => {
    cb.addEventListener('change', () => { phaseFilter[cb.value] = cb.checked; renderTree(); });
  });
}
document.addEventListener('DOMContentLoaded', init);
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Serve
# ---------------------------------------------------------------------------


def serve_html(html_path: str, port: int):
    import socket
    try:
        from fastapi import FastAPI
        from fastapi.responses import HTMLResponse
        import uvicorn
    except ImportError:
        print("\n  pip install fastapi uvicorn for --serve\n")
        _serve_fallback(html_path, port)
        return
    html_content = open(html_path, "r", encoding="utf-8").read()
    app = FastAPI(title="Module Tree Viewer")

    @app.get("/", response_class=HTMLResponse)
    async def index():
        return HTMLResponse(content=html_content)

    hostname = socket.gethostname()
    print(f"\n  Serving at:\n    Local:  http://localhost:{port}\n    Remote: http://{hostname}:{port}\n")
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="warning")


def _serve_fallback(html_path, port):
    import http.server, socket
    directory = os.path.dirname(os.path.abspath(html_path))
    filename = os.path.basename(html_path)

    class H(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *a, **kw): super().__init__(*a, directory=directory, **kw)
        def do_GET(self):
            if self.path in ("/", "/index.html"):
                self.send_response(302); self.send_header("Location", f"/{filename}"); self.end_headers(); return
            super().do_GET()
        def log_message(self, *a): pass

    hostname = socket.gethostname()
    print(f"\n  Serving at:\n    Local:  http://localhost:{port}\n    Remote: http://{hostname}:{port}\n")
    s = http.server.HTTPServer(("0.0.0.0", port), H)
    try: s.serve_forever()
    except KeyboardInterrupt: s.server_close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(xlsx_path, output_path, serve=False, port=8765):
    print(f"Reading {xlsx_path}...")
    rows = read_module_tree(xlsx_path)
    print(f"  {len(rows)} rows from 'Module Tree'")

    overview = read_overview(xlsx_path)
    print(f"  {len(overview)} module types from 'Overview'")

    summary_tab = read_summary_tab(xlsx_path)
    print(f"  {len(summary_tab)} categories from 'Summary'")

    roots = build_tree(rows)
    print(f"  {len(roots)} root modules")

    max_time = compute_metrics(roots)
    print(f"  Max time: {max_time:.1f} us")

    basename = os.path.basename(xlsx_path)
    html = generate_html(roots, max_time, basename, overview, summary_tab)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Output written to {output_path}")

    if serve:
        serve_html(output_path, port)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Module tree visualization")
    parser.add_argument("xlsx_path", help="Path to analysis.xlsx")
    parser.add_argument("-o", "--output", help="Output HTML path")
    parser.add_argument("--serve", action="store_true", help="Start HTTP server")
    parser.add_argument("--port", type=int, default=8765, help="Server port")
    args = parser.parse_args()
    output = args.output or os.path.join(os.path.dirname(args.xlsx_path) or ".", "module_tree.html")
    main(args.xlsx_path, output, serve=args.serve, port=args.port)
