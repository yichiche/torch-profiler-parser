#!/usr/bin/env python3
"""
Compare Analysis — Action-oriented diff of two trace_module_analyzer Excel reports.

Reads two analysis.xlsx files (baseline vs target) and produces an actionable
report that directly answers: "where does the performance difference come from?"

How it works:
  1. Parses both xlsx files, reading four sources from each:
     - "Summary" sheet: module-level and category-level totals.
     - "GPU Kernels" sheet: every individual kernel with name, category, time, count.
     - "Overview" sheet: module type hierarchy with timing statistics.
     - All other sheets (e.g. "Attention (pre)", "MLP (dec)"): treated as detail
       blocks with per-block kernel breakdowns, wall time, and overlap.

  2. Aggregates kernels globally by normalized short name (e.g. raw kernel names
     like "ck_tile_fmha_fwd_..." become "ck_tile::FmhaFwd"). This normalization
     is what allows matching the same logical kernel across two different runs.

  3. Detects kernel replacements via set difference on the aggregated short names:
     - GONE: kernels present in baseline but absent in target.
     - NEW:  kernels present in target but absent in baseline.
     Within the same category, GONE+NEW kernels are paired as a replacement.
     Across categories, unpaired gone-only and new-only groups are matched by
     total-time similarity (within 5x ratio), producing cross-category labels
     like "attention → other".

  4. Produces a multi-section report:
     - Executive Summary: total kernel time delta.
     - Phase Breakdown: per-tab (prefill/decode) comparison using detail blocks.
     - Tab-Level Comparison: matches detail tabs by name across both files and
       runs per-tab replacement detection + kernel diffs within each tab.
     - Kernel Replacements: what got swapped globally (from step 3).
     - Category Drill-Down: per-category delta with top contributing kernels.
     - Kernel Time Changes: same-name kernels that changed in duration.
     - Actionable Summary: condensed table of replacements and top movers.

Two use cases:
  1. Same-platform diff (e.g. MI355 aiter vs triton backend):
     → Identifies which kernels changed, got replaced, or fused.
  2. Cross-platform diff (e.g. MI355 vs B200):
     → Category-first drill-down to find biggest gaps, then top kernels within.

Output: terminal report + optional Excel diff report.

Usage:
    python compare_analysis.py baseline.xlsx target.xlsx [-o diff.xlsx] [--labels BF16 FP8]
"""

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────

_KNOWN_SHEETS = {"Summary", "Overview", "Module Tree", "GPU Kernels", "Model Info (WIP)"}

_BOLD = "\033[1m"
_DIM = "\033[2m"
_GREEN = "\033[32m"
_YELLOW = "\033[33m"
_CYAN = "\033[36m"
_RED = "\033[31m"
_RESET = "\033[0m"
_MAGENTA = "\033[35m"
_WHITE = "\033[97m"
_BG_RED = "\033[41m"
_BG_GREEN = "\033[42m"

# ── Data structures ──────────────────────────────────────────────────────────


@dataclass
class ModuleSummary:
    name: str
    total_us: float
    percentage: str
    detail_tab: str = ""


@dataclass
class CategorySummary:
    name: str
    count: int
    total_us: float
    avg_us: float
    percentage: str


@dataclass
class GPUKernel:
    name: str
    category: str
    total_us: float
    count: int
    avg_us: float
    pct_of_total: float
    module_types: str


@dataclass
class OverviewRow:
    module_type: str
    depth: int
    count: int
    mean_us: float
    std_us: float
    min_us: float
    max_us: float
    total_us: float
    pct_of_parent: float
    top_kernel: str


@dataclass
class DetailKernel:
    short_name: str
    category: str
    total_us: float = 0.0
    count: int = 0


@dataclass
class DetailBlockSummary:
    module_type: str
    kernel_count: int
    kernel_sum_us: float
    wall_time_us: float
    overlap_us: float
    categories: Dict[str, float] = field(default_factory=dict)
    kernels: Dict[str, DetailKernel] = field(default_factory=dict)


@dataclass
class AnalysisData:
    modules: List[ModuleSummary] = field(default_factory=list)
    categories: List[CategorySummary] = field(default_factory=list)
    gpu_kernels: List[GPUKernel] = field(default_factory=list)
    overview: List[OverviewRow] = field(default_factory=list)
    detail_blocks: Dict[str, DetailBlockSummary] = field(default_factory=dict)
    sheet_names: List[str] = field(default_factory=list)


# ── Parsing ──────────────────────────────────────────────────────────────────


def _safe_float(val: Any, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(str(val).replace(",", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    return int(_safe_float(val, float(default)))


def parse_analysis(filepath: str) -> AnalysisData:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    data = AnalysisData(sheet_names=wb.sheetnames)

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        in_categories = False

        for row in rows:
            cell0 = str(row[0]).strip() if row[0] is not None else ""
            if not cell0 or cell0 == "nan":
                continue

            if cell0 in ("Metric", "Category"):
                if cell0 == "Category":
                    in_categories = True
                continue

            if in_categories:
                data.categories.append(CategorySummary(
                    name=cell0,
                    count=_safe_int(row[1]),
                    total_us=_safe_float(row[2]),
                    avg_us=_safe_float(row[3]),
                    percentage=str(row[4]) if row[4] else "",
                ))
            else:
                data.modules.append(ModuleSummary(
                    name=cell0,
                    total_us=_safe_float(row[1]),
                    percentage=str(row[2]) if row[2] else "",
                    detail_tab=str(row[3]) if len(row) > 3 and row[3] else "",
                ))

    if "GPU Kernels" in wb.sheetnames:
        ws = wb["GPU Kernels"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            name = str(row[0]) if row[0] else ""
            if not name:
                continue
            data.gpu_kernels.append(GPUKernel(
                name=name,
                category=str(row[1]) if row[1] else "",
                total_us=_safe_float(row[2]),
                count=_safe_int(row[3]),
                avg_us=_safe_float(row[4]),
                pct_of_total=_safe_float(row[5]),
                module_types=str(row[6]) if len(row) > 6 and row[6] else "",
            ))

    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            mtype = str(row[0]) if row[0] else ""
            if not mtype:
                continue
            data.overview.append(OverviewRow(
                module_type=mtype,
                depth=_safe_int(row[1]),
                count=_safe_int(row[2]),
                mean_us=_safe_float(row[3]),
                std_us=_safe_float(row[4]),
                min_us=_safe_float(row[5]),
                max_us=_safe_float(row[6]),
                total_us=_safe_float(row[7]),
                pct_of_parent=_safe_float(row[8]),
                top_kernel=str(row[9]) if len(row) > 9 and row[9] else "",
            ))

    for sheet_name in wb.sheetnames:
        if sheet_name in _KNOWN_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        title_str = str(rows[0][0]) if rows[0][0] else ""
        kernel_count_match = re.search(r"(\d+)\s+kernels?", title_str)
        kernel_count = int(kernel_count_match.group(1)) if kernel_count_match else 0

        kernel_sum_us = 0.0
        wall_time_us = 0.0
        overlap_us = 0.0
        if len(rows) > 1 and rows[1][0]:
            info_str = str(rows[1][0])
            m = re.search(r"Kernel sum:\s*([\d,.]+)\s*us", info_str)
            if m:
                kernel_sum_us = float(m.group(1).replace(",", ""))
            m = re.search(r"Wall time:\s*([\d,.]+)\s*us", info_str)
            if m:
                wall_time_us = float(m.group(1).replace(",", ""))
            m = re.search(r"Overlap:\s*([\d,.]+)\s*us", info_str)
            if m:
                overlap_us = float(m.group(1).replace(",", ""))

        categories = {}
        kernels: Dict[str, DetailKernel] = {}
        in_kernels = False

        for row in rows[2:]:
            cell0 = str(row[0]).strip() if row[0] is not None else ""

            if not in_kernels:
                if cell0 == "Category":
                    continue
                if cell0 == "Module":
                    in_kernels = True
                    continue
                if cell0 and cell0 != "Total" and cell0 != "nan":
                    try:
                        cat_total = _safe_float(row[2])
                        if cat_total > 0:
                            categories[cell0] = cat_total
                    except (IndexError, ValueError):
                        pass
            else:
                if not cell0 or cell0.startswith("... truncated"):
                    if cell0.startswith("... truncated"):
                        break
                    continue
                kernel_name = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                duration = _safe_float(row[3]) if len(row) > 3 else 0.0
                category = str(row[5]) if len(row) > 5 and row[5] is not None else ""
                if kernel_name:
                    sn = _short_kernel_name(kernel_name)
                    if sn not in kernels:
                        kernels[sn] = DetailKernel(short_name=sn, category=category)
                    kernels[sn].total_us += duration
                    kernels[sn].count += 1

        data.detail_blocks[sheet_name] = DetailBlockSummary(
            module_type=sheet_name,
            kernel_count=kernel_count,
            kernel_sum_us=kernel_sum_us,
            wall_time_us=wall_time_us,
            overlap_us=overlap_us,
            categories=categories,
            kernels=kernels,
        )

    wb.close()
    return data


# ── Kernel short-name normalization ──────────────────────────────────────────


def _short_kernel_name(name: str) -> str:
    if "fmha_fwd_hd128_bf16" in name:
        return "aiter::fmha_fwd_hd128_bf16"
    if "ck_tile" in name and "FmhaFwd" in name:
        return "ck_tile::FmhaFwd"
    if "Memcpy HtoD" in name:
        return "Memcpy HtoD"
    if "Memcpy DtoH" in name:
        return "Memcpy DtoH"
    if "Custom_Cijk" in name:
        m = re.search(r"MT(\d+x\d+x\d+)", name)
        return f"hipblaslt GEMM ({m.group(1)})" if m else "hipblaslt GEMM"
    if "rcclGenericKernel" in name:
        return "RCCL AllReduce"
    if "kernel_grouped_conv" in name:
        return "CK GroupedConv"
    if "_rotary_embedding" in name:
        return "rotary_embedding"
    if "Cijk_Ailk" in name:
        m = re.search(r"MT(\d+x\d+x\d+)", name)
        return f"hipblaslt GEMM ({m.group(1)})" if m else "hipblaslt GEMM (small)"
    if "attn_fwd" == name.strip():
        return "attn_fwd (dist)"

    m = re.match(r"triton_\w+_(fused_[\w]+?)_\d+$", name)
    if m:
        return f"triton::{m.group(1)}"

    if len(name) > 60:
        return name[:57] + "..."
    return name


# ── Aggregate kernels by short name ──────────────────────────────────────────


@dataclass
class AggKernel:
    short_name: str
    category: str
    total_us: float = 0.0
    count: int = 0


def _aggregate_kernels(kernels: List[GPUKernel]) -> Dict[str, AggKernel]:
    result: Dict[str, AggKernel] = {}
    for k in kernels:
        sn = _short_kernel_name(k.name)
        if sn not in result:
            result[sn] = AggKernel(short_name=sn, category=k.category)
        result[sn].total_us += k.total_us
        result[sn].count += k.count
    return result


# ── Formatting helpers ───────────────────────────────────────────────────────


def _pct_change(base: float, target: float) -> float:
    if base > 0:
        return (target - base) / base * 100
    return 0.0 if target == base else float("inf")


def _fmt_us(us: float) -> str:
    if abs(us) >= 1e6:
        return f"{us/1e6:.2f}s"
    if abs(us) >= 1e3:
        return f"{us/1e3:.1f}ms"
    return f"{us:.0f}us"


def _fmt_delta(delta_us: float, base_us: float) -> str:
    pct = _pct_change(base_us, base_us + delta_us)
    sign = "+" if delta_us >= 0 else ""
    if abs(delta_us) >= 1e6:
        return f"{sign}{delta_us/1e6:.3f}s ({pct:+.1f}%)"
    if abs(delta_us) >= 1e3:
        return f"{sign}{delta_us/1e3:.1f}ms ({pct:+.1f}%)"
    return f"{sign}{delta_us:.0f}us ({pct:+.1f}%)"


def _color_delta(delta_us: float) -> str:
    if abs(delta_us) < 0.5:
        return _DIM
    return _RED if delta_us > 0 else _GREEN


def _bar(value: float, max_abs: float, width: int = 30) -> str:
    if max_abs == 0:
        return ""
    bar_len = int(abs(value) / max_abs * width)
    bar_len = max(bar_len, 1) if abs(value) > 0 else 0
    color = _RED if value > 0 else _GREEN
    return f"{color}{'█' * bar_len}{_RESET}"


def _trunc(s: str, width: int) -> str:
    """Truncate string to exactly `width` chars, padding or adding '...' as needed."""
    if len(s) <= width:
        return f"{s:<{width}s}"
    return s[:width - 3] + "..."


# ── Kernel replacement detection ─────────────────────────────────────────────


@dataclass
class KernelReplacement:
    """A detected kernel replacement: gone_kernels replaced by new_kernels."""
    category: str
    gone: List[Tuple[str, float]]  # (name, base_us)
    new: List[Tuple[str, float]]   # (name, target_us)
    net_delta_us: float

    @property
    def gone_total(self) -> float:
        return sum(us for _, us in self.gone)

    @property
    def new_total(self) -> float:
        return sum(us for _, us in self.new)


def _detect_replacements(base_agg: Dict[str, AggKernel],
                         tgt_agg: Dict[str, AggKernel]) -> List[KernelReplacement]:
    gone_by_cat: Dict[str, List[Tuple[str, float]]] = defaultdict(list)
    new_by_cat: Dict[str, List[Tuple[str, float]]] = defaultdict(list)

    for name, k in base_agg.items():
        if name not in tgt_agg:
            gone_by_cat[k.category].append((name, k.total_us))
    for name, k in tgt_agg.items():
        if name not in base_agg:
            new_by_cat[k.category].append((name, k.total_us))

    replacements = []
    all_cats = set(gone_by_cat) | set(new_by_cat)
    matched_cats = set()

    for cat in all_cats:
        gone = gone_by_cat.get(cat, [])
        new = new_by_cat.get(cat, [])
        if not gone and not new:
            continue
        gone_total = sum(us for _, us in gone)
        new_total = sum(us for _, us in new)
        if gone_total < 100 and new_total < 100:
            continue
        if gone and new:
            matched_cats.add(cat)
        replacements.append(KernelReplacement(
            category=cat,
            gone=sorted(gone, key=lambda x: x[1], reverse=True),
            new=sorted(new, key=lambda x: x[1], reverse=True),
            net_delta_us=new_total - gone_total,
        ))

    # Cross-category matching: categories with only GONE or only NEW kernels
    # may be paired if they have similar total time (likely a recategorization).
    gone_only = [r for r in replacements if r.gone and not r.new and r.category not in matched_cats]
    new_only = [r for r in replacements if r.new and not r.gone and r.category not in matched_cats]

    cross_matches = []
    used_gone = set()
    used_new = set()
    for gr in gone_only:
        best_match = None
        best_ratio = float("inf")
        for nr in new_only:
            if id(nr) in used_new:
                continue
            ratio = max(gr.gone_total, nr.new_total) / max(min(gr.gone_total, nr.new_total), 1)
            if ratio < best_ratio and ratio < 5.0:
                best_ratio = ratio
                best_match = nr
        if best_match:
            used_gone.add(id(gr))
            used_new.add(id(best_match))
            cross_matches.append(KernelReplacement(
                category=f"{gr.category} → {best_match.category}",
                gone=gr.gone,
                new=best_match.new,
                net_delta_us=best_match.new_total - gr.gone_total,
            ))

    final = [r for r in replacements if id(r) not in used_gone and id(r) not in used_new]
    final.extend(cross_matches)
    final.sort(key=lambda r: abs(r.net_delta_us), reverse=True)
    return final


def _detail_kernels_to_agg(kernels: Dict[str, DetailKernel]) -> Dict[str, AggKernel]:
    """Convert a DetailBlockSummary's kernel dict into AggKernel dict for reuse
    with _detect_replacements."""
    return {
        name: AggKernel(short_name=name, category=dk.category,
                        total_us=dk.total_us, count=dk.count)
        for name, dk in kernels.items()
    }


@dataclass
class TabComparison:
    """Comparison result for a single matched detail tab."""
    tab_name: str
    base_kernel_sum_us: float
    tgt_kernel_sum_us: float
    delta_us: float
    category_deltas: List[Tuple[str, float, float, float]]  # (cat, base, tgt, delta)
    kernel_deltas: List[Tuple[str, str, float, float, float]]  # (name, cat, base, tgt, delta)
    replacements: List[KernelReplacement]
    only_in_base: bool = False
    only_in_target: bool = False
    base_tab_name: str = ""
    tgt_tab_name: str = ""


def _extract_tab_module(tab_name: str) -> Tuple[str, str, str]:
    """Extract (base_module, phase, variant) from a detail tab name.

    Examples:
        'DeepseekV4DecoderLayer (dec,A)' -> ('DeepseekV4DecoderLayer', 'dec', 'A')
        'DeepseekV4DecoderLayer (A)'     -> ('DeepseekV4DecoderLayer', '',    'A')
        'Layer (dec)'                    -> ('Layer',                  'dec', '')
        'Sampler'                        -> ('Sampler',                '',    '')
    """
    m = re.match(r'^(.+?)\s*\(([^)]+)\)\s*$', tab_name)
    if not m:
        return (tab_name.strip(), '', '')
    base = m.group(1).strip()
    parts = [p.strip() for p in m.group(2).split(',')]
    phase = ''
    variant = ''
    for p in parts:
        if p in ('pre', 'dec', 'prefill', 'decode'):
            phase = p
        elif re.match(r'^[A-Z]$', p):
            variant = p
        else:
            phase = p
    return (base, phase, variant)


def _build_tab_comparison(base_tab: str, tgt_tab: str, display_name: str,
                          base_data: AnalysisData,
                          target_data: AnalysisData) -> TabComparison:
    bb = base_data.detail_blocks[base_tab]
    tb = target_data.detail_blocks[tgt_tab]

    cat_deltas = []
    all_cats = sorted(set(bb.categories) | set(tb.categories))
    for cat in all_cats:
        bv = bb.categories.get(cat, 0)
        tv = tb.categories.get(cat, 0)
        if abs(tv - bv) >= 10:
            cat_deltas.append((cat, bv, tv, tv - bv))
    cat_deltas.sort(key=lambda x: abs(x[3]), reverse=True)

    all_knames = set(bb.kernels) | set(tb.kernels)
    k_deltas = []
    for kn in all_knames:
        bk = bb.kernels.get(kn)
        tk = tb.kernels.get(kn)
        kbt = bk.total_us if bk else 0
        ktt = tk.total_us if tk else 0
        kcat = (bk or tk).category
        k_deltas.append((kn, kcat, kbt, ktt, ktt - kbt))
    k_deltas.sort(key=lambda x: abs(x[4]), reverse=True)

    base_agg = _detail_kernels_to_agg(bb.kernels)
    tgt_agg = _detail_kernels_to_agg(tb.kernels)
    reps = _detect_replacements(base_agg, tgt_agg)

    return TabComparison(
        tab_name=display_name,
        base_kernel_sum_us=bb.kernel_sum_us,
        tgt_kernel_sum_us=tb.kernel_sum_us,
        delta_us=tb.kernel_sum_us - bb.kernel_sum_us,
        category_deltas=cat_deltas,
        kernel_deltas=k_deltas,
        replacements=reps,
        base_tab_name=base_tab,
        tgt_tab_name=tgt_tab,
    )


def _compare_tabs(base: AnalysisData, target: AnalysisData) -> List[TabComparison]:
    """Compare detail blocks (tabs) across both files, matching by module type.

    Matching strategy (in priority order):
      1. Exact tab name match.
      2. Same base module + same variant letter (ignoring phase suffix differences).
      3. Same base module, 1:1 on each side (no variant letters).
    """
    base_tabs = set(base.detail_blocks)
    tgt_tabs = set(target.detail_blocks)

    matched_base: set = set()
    matched_tgt: set = set()
    results: List[TabComparison] = []

    # Pass 1: exact name match
    for tab in sorted(base_tabs & tgt_tabs):
        tc = _build_tab_comparison(tab, tab, tab, base, target)
        results.append(tc)
        matched_base.add(tab)
        matched_tgt.add(tab)

    # Pass 2: match by (base_module, variant), ignoring phase differences
    remaining_base = sorted(base_tabs - matched_base)
    remaining_tgt = sorted(tgt_tabs - matched_tgt)

    tgt_by_mod: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    for tab in remaining_tgt:
        bmod, _, variant = _extract_tab_module(tab)
        tgt_by_mod[(bmod, variant)].append(tab)

    still_unmatched_base = []
    for btab in remaining_base:
        bmod, _, bvariant = _extract_tab_module(btab)
        candidates = tgt_by_mod.get((bmod, bvariant), [])
        candidates = [c for c in candidates if c not in matched_tgt]
        if candidates:
            ttab = candidates[0]
            display = btab if btab == ttab else f"{bmod}" + (f" ({bvariant})" if bvariant else "")
            tc = _build_tab_comparison(btab, ttab, display, base, target)
            results.append(tc)
            matched_base.add(btab)
            matched_tgt.add(ttab)
        else:
            still_unmatched_base.append(btab)

    # Pass 3: match by base_module alone when both sides have exactly one
    # unmatched tab for that module
    remaining_tgt2 = sorted(tgt_tabs - matched_tgt)
    base_by_mod: Dict[str, List[str]] = defaultdict(list)
    tgt_by_mod2: Dict[str, List[str]] = defaultdict(list)
    for tab in still_unmatched_base:
        bmod, _, _ = _extract_tab_module(tab)
        base_by_mod[bmod].append(tab)
    for tab in remaining_tgt2:
        bmod, _, _ = _extract_tab_module(tab)
        tgt_by_mod2[bmod].append(tab)

    for bmod in sorted(base_by_mod):
        btabs = [t for t in base_by_mod[bmod] if t not in matched_base]
        ttabs = [t for t in tgt_by_mod2.get(bmod, []) if t not in matched_tgt]
        if not btabs or not ttabs:
            continue
        # Sort target tabs so variant A (dominant) comes first
        ttabs.sort(key=lambda t: _extract_tab_module(t)[2] or '')
        if len(btabs) == 1 and len(ttabs) >= 1:
            btab, ttab = btabs[0], ttabs[0]
            tc = _build_tab_comparison(btab, ttab, bmod, base, target)
            results.append(tc)
            matched_base.add(btab)
            matched_tgt.add(ttab)
        elif len(btabs) >= 1 and len(ttabs) == 1:
            btabs.sort(key=lambda t: _extract_tab_module(t)[2] or '')
            btab, ttab = btabs[0], ttabs[0]
            tc = _build_tab_comparison(btab, ttab, bmod, base, target)
            results.append(tc)
            matched_base.add(btab)
            matched_tgt.add(ttab)

    # Unmatched remainders
    for tab in sorted(base_tabs - matched_base):
        bb = base.detail_blocks[tab]
        results.append(TabComparison(
            tab_name=tab,
            base_kernel_sum_us=bb.kernel_sum_us, tgt_kernel_sum_us=0,
            delta_us=-bb.kernel_sum_us,
            category_deltas=[], kernel_deltas=[], replacements=[],
            only_in_base=True,
            base_tab_name=tab,
        ))

    for tab in sorted(tgt_tabs - matched_tgt):
        tb = target.detail_blocks[tab]
        results.append(TabComparison(
            tab_name=tab,
            base_kernel_sum_us=0, tgt_kernel_sum_us=tb.kernel_sum_us,
            delta_us=tb.kernel_sum_us,
            category_deltas=[], kernel_deltas=[], replacements=[],
            only_in_target=True,
            tgt_tab_name=tab,
        ))

    results.sort(key=lambda t: abs(t.delta_us), reverse=True)
    return results


# ── Terminal report ──────────────────────────────────────────────────────────


def _print_section(title: str, num: int):
    print(f"\n  {_BOLD}{_CYAN}{'━' * 70}{_RESET}")
    print(f"  {_BOLD}{_CYAN}[{num}] {title}{_RESET}")
    print(f"  {_BOLD}{_CYAN}{'━' * 70}{_RESET}\n")


def print_report(base: AnalysisData, target: AnalysisData,
                 label_a: str = "Baseline", label_b: str = "Target"):

    base_total = base.modules[0].total_us if base.modules else 0
    tgt_total = target.modules[0].total_us if target.modules else 0
    total_delta = tgt_total - base_total
    pct = total_delta / base_total * 100 if base_total else 0
    direction = "SLOWER" if total_delta > 0 else "FASTER"
    total_color = _RED if total_delta > 0 else _GREEN

    base_agg = _aggregate_kernels(base.gpu_kernels)
    tgt_agg = _aggregate_kernels(target.gpu_kernels)

    # Pre-compute all kernel deltas (used by multiple sections)
    all_names = set(base_agg) | set(tgt_agg)
    kernel_deltas = []
    for name in all_names:
        b = base_agg.get(name)
        t = tgt_agg.get(name)
        bt = b.total_us if b else 0
        tt = t.total_us if t else 0
        cat = (b or t).category
        kernel_deltas.append((name, cat, bt, tt, tt - bt))
    kernel_deltas.sort(key=lambda x: abs(x[4]), reverse=True)

    abs_total_delta = abs(total_delta) if total_delta != 0 else 1

    # ── Section 1: Executive Summary ──
    _print_section("Executive Summary", 1)
    print(f"  {_BOLD}{label_a}{_RESET} → {_BOLD}{label_b}{_RESET}")
    print(f"  Total kernel time: {_fmt_us(base_total)} → {_fmt_us(tgt_total)}")
    print(f"  {total_color}{_BOLD}{label_b} is {abs(pct):.1f}% {direction} ({_fmt_delta(total_delta, base_total)}){_RESET}")
    print()

    # ── Section 2: Per-Phase Breakdown ──
    pre_blocks = {k: v for k, v in base.detail_blocks.items() if "(pre)" in k.lower()}
    dec_blocks = {k: v for k, v in base.detail_blocks.items() if "(dec)" in k.lower()}
    tgt_pre_blocks = {k: v for k, v in target.detail_blocks.items() if "(pre)" in k.lower()}
    tgt_dec_blocks = {k: v for k, v in target.detail_blocks.items() if "(dec)" in k.lower()}

    has_phases = bool(pre_blocks or dec_blocks)
    if has_phases:
        _print_section("Phase Breakdown — Prefill vs Decode", 2)

        for phase_label, base_blocks, tgt_blocks in [
            ("Prefill", pre_blocks, tgt_pre_blocks),
            ("Decode", dec_blocks, tgt_dec_blocks),
        ]:
            common = sorted(set(base_blocks) & set(tgt_blocks))
            if not common:
                continue

            phase_base_total = sum(base_blocks[b].kernel_sum_us for b in common)
            phase_tgt_total = sum(tgt_blocks[b].kernel_sum_us for b in common)
            phase_delta = phase_tgt_total - phase_base_total
            phase_color = _color_delta(phase_delta)

            print(f"  {_BOLD}{phase_label}{_RESET}: "
                  f"{_fmt_us(phase_base_total)} → {_fmt_us(phase_tgt_total)}  "
                  f"{phase_color}{_BOLD}{_fmt_delta(phase_delta, phase_base_total)}{_RESET}")

            for block_name in common:
                bb = base_blocks[block_name]
                tb = tgt_blocks[block_name]
                bdelta = tb.kernel_sum_us - bb.kernel_sum_us
                if abs(bdelta) < 50:
                    continue
                bcolor = _color_delta(bdelta)
                print(f"    {block_name:<35s}  "
                      f"{_fmt_us(bb.kernel_sum_us)} → {_fmt_us(tb.kernel_sum_us)}  "
                      f"{bcolor}{_fmt_delta(bdelta, bb.kernel_sum_us)}{_RESET}")

                all_block_cats = sorted(
                    set(bb.categories) | set(tb.categories),
                    key=lambda c: abs(tb.categories.get(c, 0) - bb.categories.get(c, 0)),
                    reverse=True)
                for cat in all_block_cats[:5]:
                    bv = bb.categories.get(cat, 0)
                    tv = tb.categories.get(cat, 0)
                    cdelta = tv - bv
                    if abs(cdelta) < 20:
                        continue
                    ccolor = _color_delta(cdelta)
                    print(f"      {cat:<22s}  "
                          f"{_fmt_us(bv)} → {_fmt_us(tv)}  "
                          f"{ccolor}{_fmt_delta(cdelta, bv)}{_RESET}")

                if bb.kernels or tb.kernels:
                    all_knames = set(bb.kernels) | set(tb.kernels)
                    block_kdiffs = []
                    for kn in all_knames:
                        bk = bb.kernels.get(kn)
                        tk = tb.kernels.get(kn)
                        kbt = bk.total_us if bk else 0
                        ktt = tk.total_us if tk else 0
                        block_kdiffs.append((kn, kbt, ktt, ktt - kbt))
                    block_kdiffs.sort(key=lambda x: abs(x[3]), reverse=True)

                    for kn, kbt, ktt, kdelta in block_kdiffs[:3]:
                        if abs(kdelta) < 20:
                            break
                        kcolor = _color_delta(kdelta)
                        kb_s = _fmt_us(kbt) if kbt > 0 else "--"
                        kt_s = _fmt_us(ktt) if ktt > 0 else "--"
                        if kbt == 0:
                            tag = f"{_YELLOW}NEW{_RESET}"
                        elif ktt == 0:
                            tag = f"{_DIM}GONE{_RESET}"
                        else:
                            tag = f"{kcolor}{_fmt_delta(kdelta, kbt)}{_RESET}"
                        print(f"        {_trunc(kn, 48)}  {kb_s:>8s} → {kt_s:>8s}  {tag}")
            print()

    next_section = 3 if has_phases else 2

    # ── Tab-Level Comparison ──
    tab_comparisons = _compare_tabs(base, target)
    if tab_comparisons:
        _print_section("Tab-Level Comparison — per-module-tab kernel diff", next_section)
        next_section += 1

        matched_tabs = [tc for tc in tab_comparisons
                        if not tc.only_in_base and not tc.only_in_target]
        unmatched_tabs = [tc for tc in tab_comparisons
                          if tc.only_in_base or tc.only_in_target]

        for tc in matched_tabs:
            tc_color = _color_delta(tc.delta_us)
            tab_label = tc.tab_name
            if (tc.base_tab_name and tc.tgt_tab_name
                    and tc.base_tab_name != tc.tgt_tab_name):
                tab_label = f"{tc.base_tab_name} ↔ {tc.tgt_tab_name}"
            print(f"  {_BOLD}{tab_label}{_RESET}  "
                  f"{_fmt_us(tc.base_kernel_sum_us)} → {_fmt_us(tc.tgt_kernel_sum_us)}  "
                  f"{tc_color}{_BOLD}{_fmt_delta(tc.delta_us, tc.base_kernel_sum_us)}{_RESET}")

            for cat, bv, tv, cdelta in tc.category_deltas[:5]:
                ccolor = _color_delta(cdelta)
                print(f"    {cat:<22s}  "
                      f"{_fmt_us(bv)} → {_fmt_us(tv)}  "
                      f"{ccolor}{_fmt_delta(cdelta, bv)}{_RESET}")

            if tc.replacements:
                w_k = 36
                w_t = 8
                for rep in tc.replacements:
                    net_color = _color_delta(rep.net_delta_us)
                    print(f"    {_DIM}┌ {rep.category}{_RESET}")
                    print(f"    │ {_BOLD}{label_a:<{w_k}s}{_RESET} {'':>{w_t}s}    "
                          f"{_BOLD}{label_b:<{w_k}s}{_RESET}")
                    print(f"    │ {'─'*w_k}  {'─'*w_t}    {'─'*w_k}  {'─'*w_t}")
                    max_rows = max(len(rep.gone), len(rep.new))
                    for i in range(max_rows):
                        if i < len(rep.gone):
                            gn_s = _trunc(rep.gone[i][0], w_k)
                            gt_s = _fmt_us(rep.gone[i][1])
                        else:
                            gn_s = " " * w_k
                            gt_s = ""
                        if i < len(rep.new):
                            nn_s = _trunc(rep.new[i][0], w_k)
                            nt_s = _fmt_us(rep.new[i][1])
                        else:
                            nn_s = " " * w_k
                            nt_s = ""
                        print(f"    │ {_RED}{gn_s}{_RESET}  {gt_s:>{w_t}s}  →  "
                              f"{_GREEN}{nn_s}{_RESET}  {nt_s:>{w_t}s}")
                    print(f"    └ net: {net_color}{_BOLD}"
                          f"{_fmt_delta(rep.net_delta_us, rep.gone_total)}{_RESET}")
            else:
                changed = [(n, c, bt, tt, d) for n, c, bt, tt, d in tc.kernel_deltas
                           if abs(d) >= 20]
                for kn, kcat, kbt, ktt, kdelta in changed[:5]:
                    kcolor = _color_delta(kdelta)
                    kb_s = _fmt_us(kbt) if kbt > 0 else "--"
                    kt_s = _fmt_us(ktt) if ktt > 0 else "--"
                    if kbt == 0:
                        tag = f"{_YELLOW}NEW{_RESET}"
                    elif ktt == 0:
                        tag = f"{_DIM}GONE{_RESET}"
                    else:
                        tag = f"{kcolor}{_fmt_delta(kdelta, kbt)}{_RESET}"
                    print(f"    {_trunc(kn, 48)}  {kb_s:>8s} → {kt_s:>8s}  {tag}")

            print()

        if unmatched_tabs:
            print(f"  {_DIM}Tabs only in one file:{_RESET}")
            for tc in unmatched_tabs:
                if tc.only_in_base:
                    print(f"    {tc.tab_name:<35s}  {_DIM}only in {label_a}  "
                          f"({_fmt_us(tc.base_kernel_sum_us)}){_RESET}")
                else:
                    print(f"    {tc.tab_name:<35s}  {_DIM}only in {label_b}  "
                          f"({_fmt_us(tc.tgt_kernel_sum_us)}){_RESET}")
            print()

    # ── Kernel Replacements (global, table format) ──
    replacements = _detect_replacements(base_agg, tgt_agg)
    if replacements:
        _print_section("Kernel Replacements — what changed to what", next_section)
        next_section += 1

        w_kern = 44
        w_time = 10

        for rep in replacements:
            net_color = _color_delta(rep.net_delta_us)
            print(f"  {_BOLD}[{rep.category}]{_RESET}  "
                  f"net impact: {net_color}{_BOLD}{_fmt_delta(rep.net_delta_us, rep.gone_total)}{_RESET}")
            print()

            print(f"    {_BOLD}{label_a + ' (GONE)':<{w_kern}s}{_RESET}  "
                  f"{'Time':>{w_time}s}  │  "
                  f"{_BOLD}{label_b + ' (NEW)':<{w_kern}s}{_RESET}  "
                  f"{'Time':>{w_time}s}")
            print(f"    {'─'*w_kern}  {'─'*w_time}  │  {'─'*w_kern}  {'─'*w_time}")

            max_rows = max(len(rep.gone), len(rep.new))
            for i in range(max_rows):
                if i < len(rep.gone):
                    gn, gus = rep.gone[i]
                    left = f"    {_RED}{_trunc(gn, w_kern)}{_RESET}  {_fmt_us(gus):>{w_time}s}"
                else:
                    left = f"    {'':<{w_kern}s}  {'':>{w_time}s}"

                if i < len(rep.new):
                    nn, nus = rep.new[i]
                    right = f"{_GREEN}{_trunc(nn, w_kern)}{_RESET}  {_fmt_us(nus):>{w_time}s}"
                else:
                    right = f"{'':<{w_kern}s}  {'':>{w_time}s}"

                print(f"{left}  │  {right}")

            print(f"    {'─'*w_kern}  {'─'*w_time}  │  {'─'*w_kern}  {'─'*w_time}")
            print(f"    {'Total':<{w_kern}s}  {_fmt_us(rep.gone_total):>{w_time}s}  │  "
                  f"{'Total':<{w_kern}s}  {_fmt_us(rep.new_total):>{w_time}s}  "
                  f"  {net_color}Δ {_fmt_delta(rep.net_delta_us, rep.gone_total)}{_RESET}")
            print()

    # ── Section 4: Category Drill-Down ──
    _print_section("Category Drill-Down — per-category delta with top kernels", next_section)
    next_section += 1

    base_cat_map = {c.name: c for c in base.categories}
    tgt_cat_map = {c.name: c for c in target.categories}
    all_cats_set = sorted(set(base_cat_map) | set(tgt_cat_map))

    cat_entries = []
    for cat in all_cats_set:
        bc = base_cat_map.get(cat)
        tc = tgt_cat_map.get(cat)
        bt = bc.total_us if bc else 0
        tt = tc.total_us if tc else 0
        cat_entries.append((cat, bt, tt, tt - bt))

    # Sort: decrease first (largest |Δ| first), then increase (largest |Δ| first)
    decreased = sorted([e for e in cat_entries if e[3] < -100], key=lambda x: x[3])
    increased = sorted([e for e in cat_entries if e[3] > 100], key=lambda x: -x[3])
    negligible = [e for e in cat_entries if abs(e[3]) <= 100]

    kernels_by_cat: Dict[str, List[Tuple[str, float, float, float]]] = defaultdict(list)
    for name, cat, bt, tt, delta in kernel_deltas:
        kernels_by_cat[cat].append((name, bt, tt, delta))

    if decreased:
        print(f"  {_GREEN}{_BOLD}▼ Time Decreased (faster){_RESET}")
        print()
    for cat, bt, tt, delta in decreased:
        _print_cat_block(cat, bt, tt, delta, abs_total_delta, kernels_by_cat, label_a, label_b)

    if increased:
        if decreased:
            print(f"  {'─' * 70}")
            print()
        print(f"  {_RED}{_BOLD}▲ Time Increased (slower){_RESET}")
        print()
    for cat, bt, tt, delta in increased:
        _print_cat_block(cat, bt, tt, delta, abs_total_delta, kernels_by_cat, label_a, label_b)

    if negligible:
        print(f"  {_DIM}Negligible: {', '.join(c for c, _, _, _ in negligible)}{_RESET}")
        print()

    # ── Section 5: Kernel Time Changes (same-name kernels only) ──
    _print_section("Kernel Time Changes — same kernel, different time", next_section)
    next_section += 1

    changed_kernels = [(n, c, bt, tt, d) for n, c, bt, tt, d in kernel_deltas
                       if bt > 0 and tt > 0 and abs(d) >= 50]

    if changed_kernels:
        max_changed_abs = max(abs(d) for _, _, _, _, d in changed_kernels)

        print(f"  {_DIM}A = {label_a}  │  B = {label_b}{_RESET}")
        print()
        print(f"  {'#':>3s}  {'Kernel':<44s}  {'Category':>13s}  {'A':>9s}  {'B':>9s}  {'Delta':>20s}  Bar")
        print(f"  {'─'*3}  {'─'*44}  {'─'*13}  {'─'*9}  {'─'*9}  {'─'*20}  {'─'*20}")

        for i, (name, cat, bt, tt, delta) in enumerate(changed_kernels[:20], 1):
            color = _color_delta(delta)
            bar = _bar(delta, max_changed_abs, 20)
            print(f"  {i:>3d}  {_trunc(name, 44)}  {cat:>13s}  "
                  f"{_fmt_us(bt):>9s}  {_fmt_us(tt):>9s}  "
                  f"{color}{_fmt_delta(delta, bt):>20s}{_RESET}  {bar}")

        remaining = len(changed_kernels) - min(len(changed_kernels), 20)
        if remaining > 0:
            print(f"  {'':>3s}  {_DIM}(+{remaining} more kernels with Δ < {_fmt_us(changed_kernels[19][4]) if len(changed_kernels) > 19 else '...'}){_RESET}")
    else:
        print(f"  {_DIM}No same-name kernels with significant time difference.{_RESET}")
    print()

    # ── Section 6: Actionable Summary ──
    _print_section("Actionable Summary", next_section)

    print(f"  {_BOLD}Total delta:{_RESET} {total_color}{_fmt_delta(total_delta, base_total)}{_RESET}")
    print()

    # Replacement summary table
    if replacements:
        print(f"  {_BOLD}Kernel replacements:{_RESET}")
        print()
        w_lk = 36
        w_rk = 36
        w_t = 8
        print(f"    {'Category':<18s}  {label_a+' kernel':<{w_lk}s}  {'Time':>{w_t}s}  →  "
              f"{label_b+' kernel':<{w_rk}s}  {'Time':>{w_t}s}  {'Net Δ':>14s}")
        print(f"    {'─'*18}  {'─'*w_lk}  {'─'*w_t}  ─  {'─'*w_rk}  {'─'*w_t}  {'─'*14}")

        for rep in replacements:
            net_color = _color_delta(rep.net_delta_us)
            max_rows = max(len(rep.gone), len(rep.new))
            for i in range(max_rows):
                cat_str = rep.category if i == 0 else ""
                if i < len(rep.gone):
                    gn, gus = rep.gone[i]
                    left = f"{_trunc(gn, w_lk)}  {_fmt_us(gus):>{w_t}s}"
                else:
                    left = f"{'':<{w_lk}s}  {'':>{w_t}s}"

                if i < len(rep.new):
                    nn, nus = rep.new[i]
                    right = f"{_trunc(nn, w_rk)}  {_fmt_us(nus):>{w_t}s}"
                else:
                    right = f"{'':<{w_rk}s}  {'':>{w_t}s}"

                net_str = f"{net_color}{_fmt_delta(rep.net_delta_us, rep.gone_total)}{_RESET}" if i == 0 else ""
                print(f"    {cat_str:<18s}  {left}  →  {right}  {net_str}")
        print()

    # Top 3 same-kernel contributors
    top_changed = [(n, c, bt, tt, d) for n, c, bt, tt, d in kernel_deltas
                   if bt > 0 and tt > 0][:3]
    if top_changed:
        print(f"  {_BOLD}Top same-kernel time changes:{_RESET}")
        for i, (name, cat, bt, tt, delta) in enumerate(top_changed, 1):
            d_color = _color_delta(delta)
            print(f"    {i}. {d_color}{name}{_RESET} [{cat}]: "
                  f"{_fmt_us(bt)} → {_fmt_us(tt)}  "
                  f"{d_color}{_fmt_delta(delta, bt)}{_RESET}")
        print()


def _print_cat_block(cat: str, bt: float, tt: float, delta: float,
                     abs_total_delta: float,
                     kernels_by_cat: Dict[str, List[Tuple[str, float, float, float]]],
                     label_a: str, label_b: str):
    color = _color_delta(delta)
    pct_of_total = abs(delta) / abs_total_delta * 100 if abs_total_delta > 0 else 0

    print(f"  {_BOLD}{cat}{_RESET}  "
          f"{_fmt_us(bt)} → {_fmt_us(tt)}  "
          f"{color}{_BOLD}{_fmt_delta(delta, bt)}{_RESET}  "
          f"{_DIM}({pct_of_total:.0f}% of total Δ){_RESET}")

    cat_kernels = kernels_by_cat.get(cat, [])
    cat_kernels.sort(key=lambda x: abs(x[3]), reverse=True)
    for kname, kbt, ktt, kdelta in cat_kernels[:5]:
        if abs(kdelta) < 50:
            break
        kcolor = _color_delta(kdelta)
        kb_str = _fmt_us(kbt) if kbt > 0 else "--"
        kt_str = _fmt_us(ktt) if ktt > 0 else "--"
        pct_of_cat = abs(kdelta) / abs(delta) * 100 if delta != 0 else 0
        print(f"    {_trunc(kname, 50)}  {kb_str:>8s} → {kt_str:>8s}  "
              f"{kcolor}{_fmt_delta(kdelta, kbt)}{_RESET}  "
              f"{_DIM}({pct_of_cat:.0f}% of cat Δ){_RESET}")
    print()


# ── Excel report ─────────────────────────────────────────────────────────────


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
_SLOWER_FILL = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
_FASTER_FILL = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
_NEW_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
_GONE_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
_RED_FONT = Font(color="E53E3E")
_GREEN_FONT = Font(color="38A169")
_YELLOW_FONT = Font(color="D69E2E")
_GRAY_FONT = Font(color="A0AEC0")
_BOLD_FONT = Font(bold=True)
_SECTION_FILL = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")


def _write_header(ws, row: int, headers: List[str]):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _delta_font(delta: float, base: float) -> Font:
    if base > 0:
        pct = abs(delta) / base * 100
    else:
        pct = 100 if delta != 0 else 0
    if pct < 1 and abs(delta) < 100:
        return _GRAY_FONT
    return _RED_FONT if delta > 0 else _GREEN_FONT


def _row_fill(delta: float, base: float) -> Optional[PatternFill]:
    if base > 0:
        pct = abs(delta) / base * 100
    else:
        pct = 100 if delta != 0 else 0
    if pct < 1 and abs(delta) < 100:
        return None
    return _SLOWER_FILL if delta > 0 else _FASTER_FILL


def write_excel(base: AnalysisData, target: AnalysisData,
                output_path: str, label_a: str, label_b: str):
    wb = openpyxl.Workbook()

    base_total = base.modules[0].total_us if base.modules else 0
    tgt_total = target.modules[0].total_us if target.modules else 0
    total_delta = tgt_total - base_total

    base_agg = _aggregate_kernels(base.gpu_kernels)
    tgt_agg = _aggregate_kernels(target.gpu_kernels)

    # ── Sheet 1: Delta Attribution ──
    ws = wb.active
    ws.title = "Delta Attribution"
    _write_header(ws, 1, [
        "#", "Kernel", "Category",
        f"{label_a}", f"{label_b}",
        "Delta", "Delta %", "Cumul %", "Status",
    ])

    all_names = set(base_agg) | set(tgt_agg)
    kernel_deltas = []
    for name in all_names:
        b = base_agg.get(name)
        t = tgt_agg.get(name)
        bt = b.total_us if b else 0
        tt = t.total_us if t else 0
        cat = (b or t).category
        kernel_deltas.append((name, cat, bt, tt, tt - bt))
    kernel_deltas.sort(key=lambda x: abs(x[4]), reverse=True)

    abs_total_delta = abs(total_delta) if total_delta != 0 else 1
    cumulative = 0.0
    row = 2

    for i, (name, cat, bt, tt, delta) in enumerate(kernel_deltas, 1):
        cumulative += delta
        cumul_pct = cumulative / total_delta * 100 if total_delta != 0 else 0

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=round(bt / 1e6, 4) if bt > 0 else None)
        ws.cell(row=row, column=5, value=round(tt / 1e6, 4) if tt > 0 else None)
        ws.cell(row=row, column=6, value=round(delta / 1e6, 4))

        if bt > 0 and tt > 0:
            pct = delta / bt * 100
            ws.cell(row=row, column=7, value=f"{pct:+.1f}%")
            status = "changed"
        elif bt == 0:
            ws.cell(row=row, column=7, value="NEW")
            status = "NEW"
        else:
            ws.cell(row=row, column=7, value="GONE")
            status = "GONE"

        ws.cell(row=row, column=8, value=f"{cumul_pct:.1f}%")
        ws.cell(row=row, column=9, value=status)

        font = _delta_font(delta, bt)
        fill = _row_fill(delta, bt)
        if bt == 0:
            font = _YELLOW_FONT
            fill = _NEW_FILL
        elif tt == 0:
            font = _GRAY_FONT
            fill = _GONE_FILL

        for c in [6, 7, 9]:
            ws.cell(row=row, column=c).font = font
        if fill:
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = fill
        row += 1

    col_widths = [5, 50, 16, 14, 14, 14, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Category Drill-Down ──
    ws_cat = wb.create_sheet("Category Drill-Down")

    base_cat_map = {c.name: c for c in base.categories}
    tgt_cat_map = {c.name: c for c in target.categories}
    all_cats = sorted(set(base_cat_map) | set(tgt_cat_map))

    kernels_by_cat: Dict[str, List[Tuple[str, float, float, float]]] = defaultdict(list)
    for name, cat, bt, tt, delta in kernel_deltas:
        kernels_by_cat[cat].append((name, bt, tt, delta))

    cat_entries = []
    for cat in all_cats:
        bc = base_cat_map.get(cat)
        tc = tgt_cat_map.get(cat)
        bt = bc.total_us if bc else 0
        tt = tc.total_us if tc else 0
        cat_entries.append((cat, bt, tt, tt - bt))
    cat_entries.sort(key=lambda x: abs(x[3]), reverse=True)

    _write_header(ws_cat, 1, [
        "Category", f"{label_a} (s)", f"{label_b} (s)",
        "Delta (s)", "Delta %", "% of Total Δ",
    ])
    row = 2
    for cat, bt, tt, delta in cat_entries:
        ws_cat.cell(row=row, column=1, value=cat).font = _BOLD_FONT
        ws_cat.cell(row=row, column=2, value=round(bt / 1e6, 4))
        ws_cat.cell(row=row, column=3, value=round(tt / 1e6, 4))
        ws_cat.cell(row=row, column=4, value=round(delta / 1e6, 4))

        pct = _pct_change(bt, tt)
        ws_cat.cell(row=row, column=5, value=f"{pct:+.1f}%")
        ws_cat.cell(row=row, column=5).font = _delta_font(delta, bt)

        pct_total = abs(delta) / abs_total_delta * 100 if abs_total_delta > 0 else 0
        ws_cat.cell(row=row, column=6, value=f"{pct_total:.1f}%")

        fill = _row_fill(delta, bt)
        if fill:
            for c in range(1, 7):
                ws_cat.cell(row=row, column=c).fill = fill
        row += 1

        # Top kernels within this category
        cat_ks = kernels_by_cat.get(cat, [])
        cat_ks.sort(key=lambda x: abs(x[3]), reverse=True)
        for kname, kbt, ktt, kdelta in cat_ks[:10]:
            if abs(kdelta) < 50:
                break
            ws_cat.cell(row=row, column=1, value=f"  └ {kname}")
            ws_cat.cell(row=row, column=2, value=round(kbt / 1e6, 4) if kbt > 0 else None)
            ws_cat.cell(row=row, column=3, value=round(ktt / 1e6, 4) if ktt > 0 else None)
            ws_cat.cell(row=row, column=4, value=round(kdelta / 1e6, 4))
            kpct = _pct_change(kbt, ktt)
            if kbt == 0:
                ws_cat.cell(row=row, column=5, value="NEW").font = _YELLOW_FONT
            elif ktt == 0:
                ws_cat.cell(row=row, column=5, value="GONE").font = _GRAY_FONT
            else:
                ws_cat.cell(row=row, column=5, value=f"{kpct:+.1f}%")
                ws_cat.cell(row=row, column=5).font = _delta_font(kdelta, kbt)
            row += 1

        row += 1

    for i, w in enumerate([55, 14, 14, 14, 12, 14], 1):
        ws_cat.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Kernel Replacements ──
    replacements = _detect_replacements(base_agg, tgt_agg)
    if replacements:
        ws_rep = wb.create_sheet("Replacements")
        _write_header(ws_rep, 1, [
            "Category", "Direction", "Kernel", "Time (s)", "Net Impact (s)",
        ])
        row = 2
        for rep in replacements:
            ws_rep.cell(row=row, column=1, value=rep.category).font = _BOLD_FONT
            ws_rep.cell(row=row, column=5, value=round(rep.net_delta_us / 1e6, 4))
            ws_rep.cell(row=row, column=5).font = _delta_font(rep.net_delta_us, rep.gone_total)
            row += 1
            for name, us in rep.gone:
                ws_rep.cell(row=row, column=2, value="GONE")
                ws_rep.cell(row=row, column=2).font = _GRAY_FONT
                ws_rep.cell(row=row, column=3, value=name)
                ws_rep.cell(row=row, column=4, value=round(us / 1e6, 4))
                for c in range(1, 6):
                    ws_rep.cell(row=row, column=c).fill = _GONE_FILL
                row += 1
            for name, us in rep.new:
                ws_rep.cell(row=row, column=2, value="NEW")
                ws_rep.cell(row=row, column=2).font = _YELLOW_FONT
                ws_rep.cell(row=row, column=3, value=name)
                ws_rep.cell(row=row, column=4, value=round(us / 1e6, 4))
                for c in range(1, 6):
                    ws_rep.cell(row=row, column=c).fill = _NEW_FILL
                row += 1
            row += 1

        for i, w in enumerate([18, 8, 50, 14, 14], 1):
            ws_rep.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Phase Breakdown ──
    pre_blocks_b = {k: v for k, v in base.detail_blocks.items() if "(pre)" in k.lower()}
    dec_blocks_b = {k: v for k, v in base.detail_blocks.items() if "(dec)" in k.lower()}
    pre_blocks_t = {k: v for k, v in target.detail_blocks.items() if "(pre)" in k.lower()}
    dec_blocks_t = {k: v for k, v in target.detail_blocks.items() if "(dec)" in k.lower()}

    if pre_blocks_b or dec_blocks_b:
        ws_ph = wb.create_sheet("Phase Breakdown")
        _write_header(ws_ph, 1, [
            "Phase / Block / Category", f"{label_a} (us)", f"{label_b} (us)",
            "Delta (us)", "Delta %",
        ])
        row = 2

        for phase_label, bb_map, tb_map in [
            ("Prefill", pre_blocks_b, pre_blocks_t),
            ("Decode", dec_blocks_b, dec_blocks_t),
        ]:
            common = sorted(set(bb_map) & set(tb_map))
            if not common:
                continue

            phase_bt = sum(bb_map[b].kernel_sum_us for b in common)
            phase_tt = sum(tb_map[b].kernel_sum_us for b in common)
            phase_delta = phase_tt - phase_bt

            ws_ph.cell(row=row, column=1, value=phase_label).font = _BOLD_FONT
            ws_ph.cell(row=row, column=2, value=round(phase_bt, 1))
            ws_ph.cell(row=row, column=3, value=round(phase_tt, 1))
            ws_ph.cell(row=row, column=4, value=round(phase_delta, 1))
            pct = _pct_change(phase_bt, phase_tt)
            ws_ph.cell(row=row, column=5, value=f"{pct:+.1f}%")
            ws_ph.cell(row=row, column=5).font = _delta_font(phase_delta, phase_bt)
            for c in range(1, 6):
                ws_ph.cell(row=row, column=c).fill = _SECTION_FILL
            row += 1

            for block_name in common:
                bb = bb_map[block_name]
                tb = tb_map[block_name]
                bdelta = tb.kernel_sum_us - bb.kernel_sum_us

                ws_ph.cell(row=row, column=1, value=f"  {block_name}")
                ws_ph.cell(row=row, column=2, value=round(bb.kernel_sum_us, 1))
                ws_ph.cell(row=row, column=3, value=round(tb.kernel_sum_us, 1))
                ws_ph.cell(row=row, column=4, value=round(bdelta, 1))
                bpct = _pct_change(bb.kernel_sum_us, tb.kernel_sum_us)
                ws_ph.cell(row=row, column=5, value=f"{bpct:+.1f}%")
                ws_ph.cell(row=row, column=5).font = _delta_font(bdelta, bb.kernel_sum_us)
                row += 1

                all_block_cats = sorted(
                    set(bb.categories) | set(tb.categories),
                    key=lambda c: abs(tb.categories.get(c, 0) - bb.categories.get(c, 0)),
                    reverse=True)
                for cat in all_block_cats:
                    bv = bb.categories.get(cat, 0)
                    tv = tb.categories.get(cat, 0)
                    cdelta = tv - bv
                    if abs(cdelta) < 10:
                        continue
                    ws_ph.cell(row=row, column=1, value=f"    {cat}")
                    ws_ph.cell(row=row, column=2, value=round(bv, 1))
                    ws_ph.cell(row=row, column=3, value=round(tv, 1))
                    ws_ph.cell(row=row, column=4, value=round(cdelta, 1))
                    cpct = _pct_change(bv, tv)
                    ws_ph.cell(row=row, column=5, value=f"{cpct:+.1f}%")
                    ws_ph.cell(row=row, column=5).font = _delta_font(cdelta, bv)
                    row += 1
            row += 1

        for i, w in enumerate([40, 16, 16, 16, 12], 1):
            ws_ph.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 5: Tab-Level Comparison ──
    tab_comparisons = _compare_tabs(base, target)
    if tab_comparisons:
        ws_tab = wb.create_sheet("Tab Comparison")
        _write_header(ws_tab, 1, [
            "Tab / Kernel", "Category", "Status",
            f"{label_a} (us)", f"{label_b} (us)", "Delta (us)", "Delta %",
        ])
        row = 2

        for tc in tab_comparisons:
            if tc.only_in_base:
                ws_tab.cell(row=row, column=1, value=tc.tab_name).font = _BOLD_FONT
                ws_tab.cell(row=row, column=3, value=f"only in {label_a}")
                ws_tab.cell(row=row, column=3).font = _GRAY_FONT
                ws_tab.cell(row=row, column=4, value=round(tc.base_kernel_sum_us, 1))
                for c in range(1, 8):
                    ws_tab.cell(row=row, column=c).fill = _GONE_FILL
                row += 1
                continue
            if tc.only_in_target:
                ws_tab.cell(row=row, column=1, value=tc.tab_name).font = _BOLD_FONT
                ws_tab.cell(row=row, column=3, value=f"only in {label_b}")
                ws_tab.cell(row=row, column=3).font = _YELLOW_FONT
                ws_tab.cell(row=row, column=5, value=round(tc.tgt_kernel_sum_us, 1))
                for c in range(1, 8):
                    ws_tab.cell(row=row, column=c).fill = _NEW_FILL
                row += 1
                continue

            tab_label = tc.tab_name
            if (tc.base_tab_name and tc.tgt_tab_name
                    and tc.base_tab_name != tc.tgt_tab_name):
                tab_label = f"{tc.base_tab_name} ↔ {tc.tgt_tab_name}"
            ws_tab.cell(row=row, column=1, value=tab_label).font = _BOLD_FONT
            ws_tab.cell(row=row, column=4, value=round(tc.base_kernel_sum_us, 1))
            ws_tab.cell(row=row, column=5, value=round(tc.tgt_kernel_sum_us, 1))
            ws_tab.cell(row=row, column=6, value=round(tc.delta_us, 1))
            tab_pct = _pct_change(tc.base_kernel_sum_us, tc.tgt_kernel_sum_us)
            ws_tab.cell(row=row, column=7, value=f"{tab_pct:+.1f}%")
            ws_tab.cell(row=row, column=7).font = _delta_font(tc.delta_us, tc.base_kernel_sum_us)
            fill = _row_fill(tc.delta_us, tc.base_kernel_sum_us)
            if fill:
                for c in range(1, 8):
                    ws_tab.cell(row=row, column=c).fill = _SECTION_FILL
            row += 1

            for rep in tc.replacements:
                for name, us in rep.gone:
                    ws_tab.cell(row=row, column=1, value=f"  {name}")
                    ws_tab.cell(row=row, column=2, value=rep.category)
                    ws_tab.cell(row=row, column=3, value="GONE").font = _GRAY_FONT
                    ws_tab.cell(row=row, column=4, value=round(us, 1))
                    for c in range(1, 8):
                        ws_tab.cell(row=row, column=c).fill = _GONE_FILL
                    row += 1
                for name, us in rep.new:
                    ws_tab.cell(row=row, column=1, value=f"  {name}")
                    ws_tab.cell(row=row, column=2, value=rep.category)
                    ws_tab.cell(row=row, column=3, value="NEW").font = _YELLOW_FONT
                    ws_tab.cell(row=row, column=5, value=round(us, 1))
                    for c in range(1, 8):
                        ws_tab.cell(row=row, column=c).fill = _NEW_FILL
                    row += 1

            gone_names = {n for rep in tc.replacements for n, _ in rep.gone}
            new_names = {n for rep in tc.replacements for n, _ in rep.new}
            for kn, kcat, kbt, ktt, kdelta in tc.kernel_deltas:
                if kn in gone_names or kn in new_names:
                    continue
                if abs(kdelta) < 20:
                    continue
                ws_tab.cell(row=row, column=1, value=f"  {kn}")
                ws_tab.cell(row=row, column=2, value=kcat)
                ws_tab.cell(row=row, column=3, value="changed")
                ws_tab.cell(row=row, column=4, value=round(kbt, 1) if kbt > 0 else None)
                ws_tab.cell(row=row, column=5, value=round(ktt, 1) if ktt > 0 else None)
                ws_tab.cell(row=row, column=6, value=round(kdelta, 1))
                kpct = _pct_change(kbt, ktt) if kbt > 0 else 0
                ws_tab.cell(row=row, column=7, value=f"{kpct:+.1f}%" if kbt > 0 else "NEW")
                ws_tab.cell(row=row, column=7).font = _delta_font(kdelta, kbt)
                row += 1

            row += 1

        for i, w in enumerate([50, 16, 12, 16, 16, 16, 12], 1):
            ws_tab.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 6: Module Summary (compact) ──
    ws_mod = wb.create_sheet("Module Summary")
    _write_header(ws_mod, 1, [
        "Module", f"{label_a} (s)", f"{label_b} (s)", "Delta (s)", "Delta %",
    ])
    tgt_mod_map = {m.name: m for m in target.modules}
    row = 2
    for bm in base.modules:
        tm = tgt_mod_map.get(bm.name)
        if not tm:
            continue
        ws_mod.cell(row=row, column=1, value=bm.name)
        ws_mod.cell(row=row, column=2, value=round(bm.total_us / 1e6, 4))
        ws_mod.cell(row=row, column=3, value=round(tm.total_us / 1e6, 4))
        diff = tm.total_us - bm.total_us
        ws_mod.cell(row=row, column=4, value=round(diff / 1e6, 4))
        pct = diff / bm.total_us * 100 if bm.total_us else 0
        cell = ws_mod.cell(row=row, column=5, value=f"{pct:+.1f}%")
        cell.font = _delta_font(diff, bm.total_us)
        fill = _row_fill(diff, bm.total_us)
        if fill:
            for c in range(1, 6):
                ws_mod.cell(row=row, column=c).fill = fill
        row += 1

    for i, w in enumerate([40, 14, 14, 14, 12], 1):
        ws_mod.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"  {_GREEN}Excel report saved to: {output_path}{_RESET}")
    print()


# ── CLI ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Compare two trace_module_analyzer Excel reports — action-oriented diff.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
    python compare_analysis.py bf16/analysis.xlsx fp8/analysis.xlsx
    python compare_analysis.py baseline.xlsx optimized.xlsx -o comparison.xlsx
    python compare_analysis.py a.xlsx b.xlsx --labels BF16 FP8 -o diff.xlsx
""",
    )
    parser.add_argument("baseline", help="Baseline analysis.xlsx (file A)")
    parser.add_argument("target", help="Target analysis.xlsx (file B)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output Excel diff report (.xlsx)")
    parser.add_argument("--labels", nargs=2, default=None, metavar=("A", "B"),
                        help="Labels for the two files (default: auto-detect from path)")

    args = parser.parse_args()

    if args.labels:
        label_a, label_b = args.labels
    else:
        label_a = Path(args.baseline).parent.name or "Baseline"
        label_b = Path(args.target).parent.name or "Target"

    print(f"\n  Parsing {label_a}: {args.baseline}")
    base = parse_analysis(args.baseline)
    print(f"  Parsing {label_b}: {args.target}")
    target = parse_analysis(args.target)

    print_report(base, target, label_a, label_b)

    if args.output:
        write_excel(base, target, args.output, label_a, label_b)


if __name__ == "__main__":
    main()
