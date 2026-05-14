#!/usr/bin/env python3
"""
Interactive tab-by-tab kernel comparison of two analysis.xlsx files.

Usage:
    python compare_tabs.py <file_a.xlsx> <file_b.xlsx>

Prompts the user to pick a common detail tab, then prints a side-by-side
per-module kernel diff sorted by largest absolute time delta.
"""

import argparse
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

_META_SHEETS = {"Summary", "Overview", "Module Tree", "GPU Kernels", "Model Info (WIP)"}

_BOLD = "\033[1m"
_DIM = "\033[2m"
_GREEN = "\033[32m"
_RED = "\033[31m"
_CYAN = "\033[36m"
_YELLOW = "\033[33m"
_RESET = "\033[0m"

NAME_W = 44
TIME_W = 10
CAT_W = 14
SIDE_W = NAME_W + 2 + TIME_W + 2 + CAT_W  # 72


@dataclass
class KernelRow:
    module: str
    input_dims: str
    kernel_name: str
    duration_us: float
    pct_wall: float
    category: str
    source: str


def _base_module(name: str) -> str:
    return re.sub(r"_\d+$", "", name)


def _trunc(s: str, n: int = NAME_W) -> str:
    if len(s) <= n:
        return s
    return s[: n - 3] + "..."


def _fmt_time(us: float) -> str:
    if us >= 1000:
        return f"{us / 1000:.1f}ms"
    if us >= 100:
        return f"{us:.0f}us"
    if us >= 10:
        return f"{us:.1f}us"
    return f"{us:.1f}us"


def _delta_str(delta: float, base: float) -> str:
    sign = "+" if delta >= 0 else "-"
    t = _fmt_time(abs(delta))
    if base > 0:
        pct = delta / base * 100
        return f"Δ {sign}{t} ({pct:+.1f}%)"
    return f"Δ {sign}{t}"


def _parse_header(rows: list) -> Tuple[str, str]:
    title = rows[0][0] or ""
    timing = rows[1][0] or ""
    return title, timing


def _parse_categories(rows: list) -> Dict[str, float]:
    cats = {}
    for r in rows[7:17]:
        if r[0] and r[0] != "Total" and r[2] is not None:
            cats[r[0]] = float(r[2])
    return cats


def _parse_kernels(rows: list) -> List[KernelRow]:
    kernels = []
    for r in rows[19:]:
        if r[0] is None or r[0] == "Module":
            continue
        dur = r[3] if r[3] is not None else 0.0
        pct = r[4] if r[4] is not None else 0.0
        kernels.append(
            KernelRow(
                module=r[0],
                input_dims=str(r[1]) if r[1] else "",
                kernel_name=str(r[2]) if r[2] else "",
                duration_us=float(dur),
                pct_wall=float(pct),
                category=str(r[5]) if r[5] else "",
                source=str(r[9]) if r[9] else "",
            )
        )
    return kernels


def _group_by_base_module(kernels: List[KernelRow]) -> Dict[str, List[KernelRow]]:
    groups: Dict[str, List[KernelRow]] = OrderedDict()
    for k in kernels:
        base = _base_module(k.module)
        groups.setdefault(base, []).append(k)
    return groups


def _dominant_category(kernels: List[KernelRow]) -> str:
    if not kernels:
        return ""
    cat_time: Dict[str, float] = {}
    for k in kernels:
        cat_time[k.category] = cat_time.get(k.category, 0) + k.duration_us
    return max(cat_time, key=lambda c: cat_time[c])


def _default_label(path: str) -> str:
    return Path(path).parent.name


def _side_by_side_block(
    label_a: str,
    label_b: str,
    ka: List[KernelRow],
    kb: List[KernelRow],
    sum_a: float,
    sum_b: float,
    delta: float,
) -> List[str]:
    """Build side-by-side lines for one module."""
    out: List[str] = []
    sep = " │ "
    dash_l = "─" * NAME_W + "  " + "─" * TIME_W + "  " + "─" * CAT_W
    dash_r = "─" * NAME_W + "  " + "─" * TIME_W + "  " + "─" * CAT_W

    hdr_l = f"{label_a} (GONE)"
    hdr_r = f"{label_b} (NEW)"
    out.append(
        f"  {hdr_l:<{NAME_W}}  {'Time':>{TIME_W}}  {'Category':<{CAT_W}}"
        f"{sep} {hdr_r:<{NAME_W}}  {'Time':>{TIME_W}}  {'Category':<{CAT_W}}"
    )
    out.append(f"  {dash_l}{sep} {dash_r}")

    ka_sorted = sorted(ka, key=lambda k: k.duration_us, reverse=True)
    kb_sorted = sorted(kb, key=lambda k: k.duration_us, reverse=True)
    n = max(len(ka_sorted), len(kb_sorted))

    for i in range(n):
        if i < len(ka_sorted):
            kn = _trunc(ka_sorted[i].kernel_name)
            kt = _fmt_time(ka_sorted[i].duration_us)
            kc = ka_sorted[i].category
            left = f"{kn:<{NAME_W}}  {kt:>{TIME_W}}  {kc:<{CAT_W}}"
        else:
            left = " " * SIDE_W

        if i < len(kb_sorted):
            kn = _trunc(kb_sorted[i].kernel_name)
            kt = _fmt_time(kb_sorted[i].duration_us)
            kc = kb_sorted[i].category
            right = f"{kn:<{NAME_W}}  {kt:>{TIME_W}}  {kc:<{CAT_W}}"
        else:
            right = " " * SIDE_W

        out.append(f"  {left}{sep} {right}")

    out.append(f"  {dash_l}{sep} {dash_r}")

    tl = f"{'Total':<{NAME_W}}  {_fmt_time(sum_a):>{TIME_W}}  {'':<{CAT_W}}"
    tr = f"{'Total':<{NAME_W}}  {_fmt_time(sum_b):>{TIME_W}}"
    ds = _delta_str(delta, sum_a)
    out.append(f"  {tl}{sep} {tr}    {ds}")

    return out


def main():
    parser = argparse.ArgumentParser(description="Compare analysis.xlsx tabs interactively")
    parser.add_argument("file_a", help="First analysis.xlsx")
    parser.add_argument("file_b", help="Second analysis.xlsx")
    parser.add_argument("--label-a", default=None, help="Label for file A (default: parent dir name)")
    parser.add_argument("--label-b", default=None, help="Label for file B (default: parent dir name)")
    parser.add_argument("-o", "--output-dir", default="/home/yichiche/agent-box/profile",
                        help="Directory to save comparison report")
    args = parser.parse_args()

    label_a = args.label_a or _default_label(args.file_a)
    label_b = args.label_b or _default_label(args.file_b)

    wb_a = openpyxl.load_workbook(args.file_a, read_only=True)
    wb_b = openpyxl.load_workbook(args.file_b, read_only=True)

    tabs_a = set(wb_a.sheetnames) - _META_SHEETS
    tabs_b = set(wb_b.sheetnames) - _META_SHEETS
    common = sorted(tabs_a & tabs_b)

    if not common:
        print("No common detail tabs found between the two files.")
        sys.exit(1)

    print(f"\n{_BOLD}Common tabs between the two files:{_RESET}")
    for i, tab in enumerate(common, 1):
        print(f"  {_CYAN}{i}{_RESET}. {tab}")

    only_a = sorted(tabs_a - tabs_b)
    only_b = sorted(tabs_b - tabs_a)
    if only_a:
        print(f"\n  {_DIM}Only in {label_a}: {', '.join(only_a)}{_RESET}")
    if only_b:
        print(f"  {_DIM}Only in {label_b}: {', '.join(only_b)}{_RESET}")

    print()
    choice = input(f"Select tab to compare (1-{len(common)}): ").strip()
    try:
        idx = int(choice) - 1
        if idx < 0 or idx >= len(common):
            raise ValueError
    except ValueError:
        print("Invalid selection.")
        sys.exit(1)

    tab_name = common[idx]
    print(f"\n{_BOLD}Comparing tab: {tab_name}{_RESET}\n")

    rows_a = list(wb_a[tab_name].iter_rows(values_only=True))
    rows_b = list(wb_b[tab_name].iter_rows(values_only=True))

    title_a, timing_a = _parse_header(rows_a)
    title_b, timing_b = _parse_header(rows_b)

    cats_a = _parse_categories(rows_a)
    cats_b = _parse_categories(rows_b)

    kernels_a = _parse_kernels(rows_a)
    kernels_b = _parse_kernels(rows_b)

    groups_a = _group_by_base_module(kernels_a)
    groups_b = _group_by_base_module(kernels_b)

    all_bases = list(OrderedDict.fromkeys(list(groups_a.keys()) + list(groups_b.keys())))

    lines: list[str] = []

    def emit(s: str = ""):
        lines.append(s)
        print(s)

    emit(f"{'=' * 120}")
    emit(f"  {label_a}: {title_a}")
    emit(f"       {timing_a}")
    emit(f"  {label_b}: {title_b}")
    emit(f"       {timing_b}")
    emit(f"{'=' * 120}")

    # Category comparison
    all_cats = sorted(
        set(list(cats_a.keys()) + list(cats_b.keys())),
        key=lambda c: abs(cats_a.get(c, 0) - cats_b.get(c, 0)),
        reverse=True,
    )
    emit(f"\n{_BOLD}Category Breakdown:{_RESET}")
    emit(
        f"  {'Category':<20} {label_a + ' (us)':>12} {label_b + ' (us)':>12}"
        f" {'Delta (us)':>12} {'Change':>10}"
    )
    emit(f"  {'-' * 20} {'-' * 12} {'-' * 12} {'-' * 12} {'-' * 10}")
    total_a = sum(cats_a.values())
    total_b = sum(cats_b.values())
    for cat in all_cats:
        va = cats_a.get(cat, 0)
        vb = cats_b.get(cat, 0)
        delta = vb - va
        if va > 0:
            pct = delta / va * 100
            pct_str = f"{pct:+.0f}%"
        else:
            pct_str = "new"
        color = _RED if delta > 0 else _GREEN if delta < 0 else ""
        emit(
            f"  {cat:<20} {va:>12.1f} {vb:>12.1f}"
            f" {color}{delta:>+12.1f}{_RESET} {pct_str:>10}"
        )
    delta_total = total_b - total_a
    color_total = _RED if delta_total > 0 else _GREEN if delta_total < 0 else ""
    emit(
        f"  {'TOTAL':<20} {total_a:>12.1f} {total_b:>12.1f}"
        f" {color_total}{delta_total:>+12.1f}{_RESET}"
    )

    # Per-module comparison sorted by |delta|
    module_diffs: list[Tuple[str, float, float, float, List[KernelRow], List[KernelRow]]] = []
    for base in all_bases:
        ka = groups_a.get(base, [])
        kb = groups_b.get(base, [])
        sum_a = sum(k.duration_us for k in ka)
        sum_b = sum(k.duration_us for k in kb)
        delta = sum_b - sum_a
        module_diffs.append((base, sum_a, sum_b, delta, ka, kb))

    module_diffs.sort(key=lambda x: x[3])

    emit("")

    for base, sum_a, sum_b, delta, ka, kb in module_diffs:
        all_kernels = list(ka) + list(kb)
        dom_cat = _dominant_category(all_kernels)
        color = _RED if delta > 0 else _GREEN if delta < 0 else ""
        ds = _delta_str(delta, sum_a)

        emit(f"    {_BOLD}[{dom_cat}] {base}{_RESET}  net impact: {color}{ds}{_RESET}")
        emit("")

        block = _side_by_side_block(label_a, label_b, ka, kb, sum_a, sum_b, delta)
        for line in block:
            emit(f"    {line}")

        emit("")

    wb_a.close()
    wb_b.close()

    # Save report (strip ANSI codes)
    ansi_re = re.compile(r"\033\[[0-9;]*m")
    safe_tab = re.sub(r"[^\w\-]", "_", tab_name)
    out_path = Path(args.output_dir) / f"compare_{safe_tab}.txt"
    out_path.write_text("\n".join(ansi_re.sub("", line) for line in lines) + "\n")
    print(f"\n{_DIM}Report saved to {out_path}{_RESET}")


if __name__ == "__main__":
    main()
