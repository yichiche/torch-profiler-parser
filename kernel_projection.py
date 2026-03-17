#!/usr/bin/env python3
"""
Kernel Projection Tool — Interactive what-if analysis for kernel performance.

Reads analysis.xlsx (output of trace_module_analyzer.py), shows top kernels by
time occupancy, and lets users interactively specify kernel improvements to
project their impact on TTFT, ITL, E2E latency, and throughput.

Usage:
    python kernel_projection.py /path/to/analysis.xlsx [--output-tokens N] [--top N]

Examples:
    # LLM model with prefill/decode phases
    python kernel_projection.py /path/to/profile-baseline/analysis.xlsx --output-tokens 128

    # Diffusion model (single phase)
    python kernel_projection.py /path/to/wen2.2/analysis.xlsx
"""

import argparse
import collections
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

# Import parse_excel and data classes from compare_traces
_BENCHMARK_DIR = str(Path(__file__).resolve().parent.parent / "benchmark")
if _BENCHMARK_DIR not in sys.path:
    sys.path.insert(0, _BENCHMARK_DIR)

from compare_traces import KernelInfo, ModuleInfo, parse_excel

# ── Constants ────────────────────────────────────────────────────────────────

_KNOWN_SHEETS = {"Summary", "Overview", "Module Tree", "GPU Kernels", "Model Info (WIP)"}

_PHASE_LABELS = {"pre": "prefill", "dec": "decode", "": "all"}

# Terminal colors
_BOLD = "\033[1m"
_DIM = "\033[2m"
_GREEN = "\033[32m"
_YELLOW = "\033[33m"
_CYAN = "\033[36m"
_RED = "\033[31m"
_RESET = "\033[0m"


# ── Data structures ─────────────────────────────────────────────────────────


@dataclass
class GPUKernelRow:
    """One row from the GPU Kernels sheet."""

    name: str
    category: str
    total_duration_us: float
    count: int
    avg_us: float
    pct_of_total: float
    module_types_str: str


@dataclass
class DetailKernelRef:
    """Where a kernel appears in a specific detail sheet."""

    module_type: str
    phase: str  # 'pre', 'dec', ''
    kernel_indices: List[int]  # indices into ModuleInfo.kernels
    total_duration_us: float  # sum of matched kernel durations in this module
    pct_of_wall_time: float  # total_duration / module wall_time * 100


@dataclass
class TopKernel:
    """A top kernel with cross-references to detail sheets."""

    rank: int
    name: str
    category: str
    total_duration_us: float
    count: int
    pct_of_total: float
    detail_refs: List[DetailKernelRef]
    phases: Set[str] = field(default_factory=set)


@dataclass
class Modification:
    """A user-applied kernel modification."""

    kernel_name: str
    improvement_pct: float  # 20 means 20% faster


@dataclass
class ModuleProjection:
    """Projection result for one (module_type, phase)."""

    module_type: str
    phase: str
    original_kernel_sum_us: float
    projected_kernel_sum_us: float
    original_wall_time_us: float
    projected_wall_time_us: float
    delta_wall_time_us: float
    delta_wall_time_pct: float
    kernel_deltas: List[Tuple[str, float, float]]  # (name, old_dur, new_dur)


# ── Overview + Module Tree parsers ───────────────────────────────────────────


@dataclass
class OverviewNode:
    """A module type from the Overview sheet."""

    module_type: str
    depth: int
    count: int
    mean_us: float
    total_us: float
    pct_of_parent: float
    parent: Optional[str] = None
    children: List[str] = field(default_factory=list)


@dataclass
class PhaseRootModule:
    """A root module contributing to a phase's per-step time."""

    module_type: str
    phase: str
    per_step_us: float  # from Overview mean
    pct_of_phase: float = 0.0  # computed after grouping
    has_detail_sheet: bool = False
    # For projection
    projected_per_step_us: float = 0.0
    change_pct: float = 0.0


@dataclass
class PhaseModel:
    """Complete model for one phase with all root modules."""

    phase: str
    root_modules: List[PhaseRootModule]
    total_step_us: float


def parse_overview_tree(filepath: str) -> Tuple[Dict[str, OverviewNode], Dict[str, Set[str]]]:
    """Parse Overview sheet to build module hierarchy.

    Returns:
        nodes: {module_type: OverviewNode} (first occurrence used for root stats)
        ancestors: {module_type: set of ancestor module_types}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    nodes: Dict[str, OverviewNode] = {}
    ancestors: Dict[str, Set[str]] = {}
    stack: List[Tuple[int, str]] = []

    if "Overview" not in wb.sheetnames:
        wb.close()
        return nodes, ancestors

    ws = wb["Overview"]
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:
        if not row[0] or row[1] is None:
            continue
        label = str(row[0]).strip()
        clean = re.sub(r'^[\s│├└─]+', '', label).strip()
        if not clean or clean == "(self)":
            continue

        depth = int(row[1])
        count = int(row[2]) if row[2] else 1
        mean_us = float(row[3]) if row[3] else 0.0
        total_us = float(row[7]) if row[7] else 0.0
        pct_parent = float(row[8]) if row[8] is not None else 0.0

        while stack and stack[-1][0] >= depth:
            stack.pop()

        parent = stack[-1][1] if stack else None
        ancestor_set = {s[1] for s in stack}

        if clean not in ancestors:
            ancestors[clean] = ancestor_set
        else:
            ancestors[clean] |= ancestor_set

        if depth == 0:
            nodes[clean] = OverviewNode(
                module_type=clean, depth=depth, count=count,
                mean_us=mean_us, total_us=total_us,
                pct_of_parent=pct_parent, parent=parent,
            )
        elif parent and parent in nodes:
            nodes[parent].children.append(clean)

        # Also store child nodes for pct_of_parent propagation
        if clean not in nodes:
            nodes[clean] = OverviewNode(
                module_type=clean, depth=depth, count=count,
                mean_us=mean_us, total_us=total_us,
                pct_of_parent=pct_parent, parent=parent,
            )

        stack.append((depth, clean))

    wb.close()
    return nodes, ancestors


def parse_module_tree_phases(filepath: str) -> Dict[str, str]:
    """Parse Module Tree sheet to get phase for each root module type."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    phases: Dict[str, str] = {}

    if "Module Tree" not in wb.sheetnames:
        wb.close()
        return phases

    ws = wb["Module Tree"]
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:
        if not row[0]:
            continue
        name = str(row[0]).strip()
        phase = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        # Root instances have no tree-prefix indentation
        if name.startswith(" ") or name.startswith("├") or name.startswith("└") or name.startswith("│"):
            continue

        # Extract module type (remove _N suffix)
        type_match = re.match(r'^(.+?)_\d+$', name)
        mtype = type_match.group(1) if type_match else name

        if phase and phase.lower() not in ("none", ""):
            phase_short = "pre" if "prefill" in phase.lower() else "dec" if "decode" in phase.lower() else ""
            if phase_short and mtype not in phases:
                phases[mtype] = phase_short

    wb.close()
    return phases


def build_phase_models(
    filepath: str,
    modules: Dict[Tuple[str, str], ModuleInfo],
    overview_nodes: Dict[str, OverviewNode],
    ancestors: Dict[str, Set[str]],
) -> Dict[str, PhaseModel]:
    """Build phase models with correct per-step totals from Overview data.

    Uses Overview root modules for accurate phase totals, and Module Tree
    for phase assignment.

    Works for both LLM models (pre/dec phases) and single-phase models
    (e.g. diffusion models like Wan where phase="").
    """
    # Detect if this is a single-phase model (no pre/dec phases in detail sheets)
    all_phases = {phase for (_, phase) in modules}
    is_single_phase = not (all_phases & {"pre", "dec"})

    # Get phase for each root module type from Module Tree
    root_phases = parse_module_tree_phases(filepath)

    # Also infer phases from detail sheets for roots that have them
    for (mtype, phase) in modules:
        if phase:
            root_phases.setdefault(mtype, phase)

    # Build root modules per phase
    phase_roots: Dict[str, List[PhaseRootModule]] = collections.defaultdict(list)
    detail_types = {mtype for (mtype, _) in modules}

    # For single-phase models, use "" as the canonical phase
    default_phase = "" if is_single_phase else ""

    for mtype, node in overview_nodes.items():
        if node.depth != 0:
            continue

        phase = root_phases.get(mtype, "")
        if not phase and not is_single_phase:
            # Try to infer from children's detail sheet phases
            for (dtype, dphase) in modules:
                if dphase and dtype in (ancestors.get(dtype, set()) | {dtype}):
                    node_ancestors = ancestors.get(dtype, set())
                    if mtype in node_ancestors:
                        phase = dphase
                        break
        if not phase and not is_single_phase:
            continue  # can't determine phase for LLM model, skip

        # For single-phase models, use "" as phase
        if is_single_phase:
            phase = ""

        has_detail = (mtype, phase) in modules
        phase_roots[phase].append(
            PhaseRootModule(
                module_type=mtype,
                phase=phase,
                per_step_us=node.mean_us,
                has_detail_sheet=has_detail,
            )
        )

    # Also add detail-sheet-only roots not in Overview (LogitsProcessor, Sampler, etc.)
    for (mtype, phase) in modules:
        # Check if already added from Overview
        already = any(rm.module_type == mtype for rm in phase_roots.get(phase, []))
        if not already:
            minfo = modules[(mtype, phase)]
            # Check if this is a root (not a child of another detail sheet module)
            is_child = False
            for (other_type, other_phase) in modules:
                if other_phase == phase and other_type != mtype:
                    if mtype in ancestors.get(mtype, set()):
                        pass  # self-reference, skip
                    other_ancestors = ancestors.get(mtype, set())
                    if other_type in other_ancestors:
                        is_child = True
                        break
            if not is_child:
                phase_roots[phase].append(
                    PhaseRootModule(
                        module_type=mtype,
                        phase=phase,
                        per_step_us=minfo.wall_time_us,
                        has_detail_sheet=True,
                    )
                )

    # Build phase models with totals and percentages
    models = {}
    for phase, roots in phase_roots.items():
        total = sum(r.per_step_us for r in roots)
        for r in roots:
            r.pct_of_phase = (r.per_step_us / total * 100) if total > 0 else 0
            r.projected_per_step_us = r.per_step_us  # default: unchanged
        models[phase] = PhaseModel(phase=phase, root_modules=roots, total_step_us=total)

    return models


def project_phase_models(
    phase_models: Dict[str, PhaseModel],
    modules: Dict[Tuple[str, str], ModuleInfo],
    modifications: List[Modification],
    overview_nodes: Dict[str, OverviewNode],
    ancestors: Dict[str, Set[str]],
) -> Dict[str, PhaseModel]:
    """Apply modifications to phase models and compute projected per-step times."""
    import copy
    projected = copy.deepcopy(phase_models)

    # First, compute per-module detail sheet changes
    module_changes: Dict[Tuple[str, str], float] = {}  # (mtype, phase) -> change_pct
    for (mtype, phase), minfo in modules.items():
        proj = project_module(minfo, modifications)
        if proj:
            module_changes[(mtype, phase)] = proj.delta_wall_time_pct

    # Apply changes to phase model root modules
    for phase, model in projected.items():
        for rm in model.root_modules:
            if (rm.module_type, phase) in module_changes:
                # Root has a detail sheet with direct change
                rm.change_pct = module_changes[(rm.module_type, phase)]
                rm.projected_per_step_us = rm.per_step_us * (1 + rm.change_pct / 100)
            else:
                # Root has no detail sheet — propagate from children
                # Find children with detail sheet changes
                total_child_impact = 0.0
                for (ctype, cphase), cpct in module_changes.items():
                    if cphase != phase:
                        continue
                    # Check if this child is a descendant of this root
                    child_ancestors = ancestors.get(ctype, set())
                    if rm.module_type in child_ancestors:
                        # Get child's % of this parent from Overview
                        child_node = overview_nodes.get(ctype)
                        if child_node and child_node.pct_of_parent > 0:
                            # Use pct_of_parent from the direct child relationship
                            total_child_impact += child_node.pct_of_parent / 100 * cpct
                if total_child_impact != 0:
                    rm.change_pct = total_child_impact
                    rm.projected_per_step_us = rm.per_step_us * (1 + rm.change_pct / 100)

        model.total_step_us = sum(r.projected_per_step_us for r in model.root_modules)

    return projected


def get_root_detail_modules(
    modules: Dict[Tuple[str, str], ModuleInfo],
    ancestors: Dict[str, Set[str]],
) -> Set[Tuple[str, str]]:
    """Identify detail sheet modules that are NOT children of other detail sheet modules."""
    detail_types_by_phase: Dict[str, Set[str]] = collections.defaultdict(set)
    for mtype, phase in modules:
        detail_types_by_phase[phase].add(mtype)

    roots = set()
    for (mtype, phase), minfo in modules.items():
        anc = ancestors.get(mtype, set())
        same_phase_detail_types = detail_types_by_phase.get(phase, set())
        if not (anc & same_phase_detail_types):
            roots.add((mtype, phase))
    return roots


# ── GPU Kernels sheet parser ─────────────────────────────────────────────────


def parse_gpu_kernels_sheet(filepath: str) -> List[GPUKernelRow]:
    """Parse the GPU Kernels sheet for global kernel ranking."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    rows_out = []

    if "GPU Kernels" not in wb.sheetnames:
        wb.close()
        return rows_out

    ws = wb["GPU Kernels"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return rows_out

    # Skip header row
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            category = str(row[1]).strip() if row[1] else ""
            total_dur = float(row[2]) if row[2] else 0.0
            count = int(row[3]) if row[3] else 0
            avg = float(row[4]) if row[4] else 0.0
            pct = float(row[5]) if row[5] else 0.0
            mod_types = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            rows_out.append(
                GPUKernelRow(
                    name=name,
                    category=category,
                    total_duration_us=total_dur,
                    count=count,
                    avg_us=avg,
                    pct_of_total=pct,
                    module_types_str=mod_types,
                )
            )
        except (ValueError, TypeError):
            continue

    wb.close()
    return rows_out


# ── Cross-reference builder ──────────────────────────────────────────────────


def build_top_kernels(
    gpu_kernels: List[GPUKernelRow],
    modules: Dict[Tuple[str, str], ModuleInfo],
    top_n: int = 10,
) -> List[TopKernel]:
    """Build top-N kernel list with cross-references to detail sheets."""
    top = []
    for rank, gk in enumerate(gpu_kernels[:top_n], 1):
        detail_refs = []
        phases = set()

        for (mtype, phase), minfo in modules.items():
            matched_indices = []
            matched_dur = 0.0
            for i, k in enumerate(minfo.kernels):
                if k.kernel_name == gk.name:
                    matched_indices.append(i)
                    matched_dur += k.duration_us

            if matched_indices:
                pct = (matched_dur / minfo.wall_time_us * 100) if minfo.wall_time_us > 0 else 0.0
                detail_refs.append(
                    DetailKernelRef(
                        module_type=mtype,
                        phase=phase,
                        kernel_indices=matched_indices,
                        total_duration_us=matched_dur,
                        pct_of_wall_time=pct,
                    )
                )
                phases.add(phase)

        top.append(
            TopKernel(
                rank=rank,
                name=gk.name,
                category=gk.category,
                total_duration_us=gk.total_duration_us,
                count=gk.count,
                pct_of_total=gk.pct_of_total,
                detail_refs=detail_refs,
                phases=phases,
            )
        )
    return top


# ── Projection engine ────────────────────────────────────────────────────────


def project_module(
    minfo: ModuleInfo,
    modifications: List[Modification],
) -> Optional[ModuleProjection]:
    """Apply modifications to a module and compute projected wall time."""
    kernel_deltas = []
    total_delta = 0.0

    for ki in minfo.kernels:
        new_dur = ki.duration_us
        for mod in modifications:
            if ki.kernel_name == mod.kernel_name:
                new_dur = ki.duration_us * (1 - mod.improvement_pct / 100)
                break
        if new_dur != ki.duration_us:
            kernel_deltas.append((ki.kernel_name, ki.duration_us, new_dur))
            total_delta += new_dur - ki.duration_us

    if not kernel_deltas:
        return None

    new_kernel_sum = minfo.kernel_sum_us + total_delta

    # Scaled wall time: preserve overlap ratio
    if minfo.kernel_sum_us > 0:
        ratio = minfo.wall_time_us / minfo.kernel_sum_us
        new_wall_time = new_kernel_sum * ratio
    else:
        new_wall_time = minfo.wall_time_us

    delta_wt = new_wall_time - minfo.wall_time_us
    delta_pct = (delta_wt / minfo.wall_time_us * 100) if minfo.wall_time_us > 0 else 0.0

    return ModuleProjection(
        module_type=minfo.module_type,
        phase=minfo.phase,
        original_kernel_sum_us=minfo.kernel_sum_us,
        projected_kernel_sum_us=new_kernel_sum,
        original_wall_time_us=minfo.wall_time_us,
        projected_wall_time_us=new_wall_time,
        delta_wall_time_us=delta_wt,
        delta_wall_time_pct=delta_pct,
        kernel_deltas=kernel_deltas,
    )


def compute_phase_totals(
    modules: Dict[Tuple[str, str], ModuleInfo],
    root_keys: Set[Tuple[str, str]],
) -> Dict[str, float]:
    """Compute total wall time per phase from root-level detail sheets only."""
    totals: Dict[str, float] = collections.defaultdict(float)
    for (mtype, phase), minfo in modules.items():
        if (mtype, phase) in root_keys:
            totals[phase] += minfo.wall_time_us
    return dict(totals)


def compute_projections(
    modules: Dict[Tuple[str, str], ModuleInfo],
    modifications: List[Modification],
    root_keys: Set[Tuple[str, str]],
) -> Tuple[List[ModuleProjection], Dict[str, Tuple[float, float, float]]]:
    """Apply all modifications and return projections.

    Returns:
        module_projections: list of per-module projection results (ALL affected modules)
        phase_summary: {phase: (original_us, projected_us, delta_pct)} using root modules only
    """
    module_projections = []
    for (mtype, phase), minfo in modules.items():
        proj = project_module(minfo, modifications)
        if proj:
            module_projections.append(proj)

    # Aggregate per phase — only root modules to avoid double-counting
    phase_original: Dict[str, float] = collections.defaultdict(float)
    phase_projected: Dict[str, float] = collections.defaultdict(float)

    for (mtype, phase), minfo in modules.items():
        if (mtype, phase) not in root_keys:
            continue
        phase_original[phase] += minfo.wall_time_us
        # Check if we have a projection for this module
        proj_match = [p for p in module_projections if p.module_type == mtype and p.phase == phase]
        if proj_match:
            phase_projected[phase] += proj_match[0].projected_wall_time_us
        else:
            phase_projected[phase] += minfo.wall_time_us

    phase_summary = {}
    for phase in phase_original:
        orig = phase_original[phase]
        proj = phase_projected[phase]
        delta_pct = ((proj - orig) / orig * 100) if orig > 0 else 0.0
        phase_summary[phase] = (orig, proj, delta_pct)

    return module_projections, phase_summary


# ── Display helpers ──────────────────────────────────────────────────────────


def fmt_us(us: float) -> str:
    """Format microseconds with comma separators."""
    if us >= 1_000_000:
        return f"{us:,.0f}"
    elif us >= 1_000:
        return f"{us:,.0f}"
    else:
        return f"{us:,.1f}"


def fmt_pct(pct: float, sign: bool = True) -> str:
    """Format percentage."""
    if sign and pct <= 0:
        return f"{_GREEN}{pct:+.1f}%{_RESET}"
    elif sign and pct > 0:
        return f"{_RED}{pct:+.1f}%{_RESET}"
    return f"{pct:.1f}%"


def phase_label(phase: str) -> str:
    return _PHASE_LABELS.get(phase, phase)


def detect_model_type(modules: Dict[Tuple[str, str], ModuleInfo]) -> str:
    """Detect model type from phases present in detail sheets."""
    phases = {phase for (_, phase) in modules.keys()}
    if "pre" in phases or "dec" in phases:
        return "LLM"
    return "single-phase"


def print_header(
    filepath: str,
    modules: Dict[Tuple[str, str], ModuleInfo],
    root_keys: Set[Tuple[str, str]],
):
    """Print tool header with model detection info."""
    model_type = detect_model_type(modules)
    phase_totals = compute_phase_totals(modules, root_keys)

    print(f"\n{_BOLD}{'=' * 72}{_RESET}")
    print(f"{_BOLD}  Kernel Projection Tool{_RESET}")
    print(f"  Source: {filepath}")

    if model_type == "LLM":
        phases_found = []
        if "pre" in phase_totals:
            phases_found.append("prefill")
        if "dec" in phase_totals:
            phases_found.append("decode")
        print(f"  Model type: {_CYAN}LLM{_RESET} (phases: {', '.join(phases_found)})")
        _PO = {"pre": 0, "dec": 1}
        for phase, total in sorted(phase_totals.items(), key=lambda x: _PO.get(x[0], 99)):
            if phase:
                sheets = sum(1 for (_, p) in modules if p == phase)
                print(f"    {phase_label(phase):>7s} step time: {fmt_us(total)} us ({sheets} detail sheets)")
    else:
        total = sum(phase_totals.values())
        sheets = len(modules)
        print(f"  Model type: {_CYAN}Single-phase{_RESET} (no prefill/decode)")
        print(f"    Step time: {fmt_us(total)} us ({sheets} detail sheets)")

    print(f"{'=' * 72}\n")


def print_top_kernels(top_kernels: List[TopKernel]):
    """Print the top-N kernel table."""
    print(f"{_BOLD}{'─' * 72}{_RESET}")
    print(f"{_BOLD}  Top {len(top_kernels)} Kernels by Time{_RESET}")
    print(f"{'─' * 72}")

    # Header
    print(
        f"  {'#':>3s}  {'Kernel Name':<42s} {'Category':<13s} "
        f"{'% Total':>7s}  {'Phase':<8s}"
    )
    print(f"  {'─' * 3}  {'─' * 42} {'─' * 13} {'─' * 7}  {'─' * 8}")

    for tk in top_kernels:
        # Truncate kernel name for display
        display_name = tk.name
        if len(display_name) > 42:
            display_name = display_name[:39] + "..."

        # Phase info
        if tk.phases == {"pre", "dec"} or tk.phases == {"pre", "dec", ""}:
            phase_str = "pre+dec"
        elif tk.phases:
            phase_str = "+".join(sorted(tk.phases)) or "all"
        else:
            phase_str = "all"

        print(
            f"  {tk.rank:>3d}  {display_name:<42s} {tk.category:<13s} "
            f"{tk.pct_of_total:>6.1f}%  {phase_str:<8s}"
        )

    print()


def print_kernel_detail(tk: TopKernel, modules: Dict[Tuple[str, str], ModuleInfo]):
    """Print detailed info about a selected kernel."""
    print(f"\n{_BOLD}  Selected: {tk.name}{_RESET}")
    print(f"  Category: {tk.category}  |  Total: {fmt_us(tk.total_duration_us)} us  |  Count: {tk.count}")

    if tk.detail_refs:
        print(f"\n  Appears in:")
        for ref in tk.detail_refs:
            pl = phase_label(ref.phase)
            minfo = modules.get((ref.module_type, ref.phase))
            wt = fmt_us(minfo.wall_time_us) if minfo else "?"
            print(
                f"    {ref.module_type} ({pl}): "
                f"{fmt_us(ref.total_duration_us)} us "
                f"({ref.pct_of_wall_time:.1f}% of module wall time {wt} us, "
                f"{len(ref.kernel_indices)} occurrence{'s' if len(ref.kernel_indices) > 1 else ''})"
            )
    else:
        print(f"  {_YELLOW}Not found in any detail sheet (global-only kernel){_RESET}")
    print()


def print_projection_with_deduction(
    modules: Dict[Tuple[str, str], ModuleInfo],
    modifications: List[Modification],
    model_type: str,
    output_tokens: Optional[int],
    root_keys: Set[Tuple[str, str]],
    original_pm: Dict[str, PhaseModel],
    projected_pm: Dict[str, PhaseModel],
    verbose: bool = False,
):
    """Print projection results with full deduction chain.

    Shows the complete calculation process:
      Step 1: Kernel improvement → module wall time change (with kernel % of module)
      Step 2: Module wall time change → phase step time (with module % of phase)
      Step 3: Phase step time → TTFT / ITL / E2E

    Works for both LLM (prefill/decode) and non-LLM (single-phase) models.
    """
    print(f"\n{_BOLD}  ── Projection Results ──{_RESET}\n")

    # ── Step 1: Kernel → Module ──
    # Compute per-module projections for affected detail-sheet modules
    module_projections = []
    for (mtype, phase), minfo in modules.items():
        proj = project_module(minfo, modifications)
        if proj:
            module_projections.append(proj)

    if not module_projections:
        print(f"  {_DIM}No modules affected by these changes.{_RESET}\n")
        return

    print(f"  {_BOLD}Step 1: Kernel improvement → Module wall time{_RESET}\n")
    for mp in module_projections:
        pl = phase_label(mp.phase)
        is_child = (mp.module_type, mp.phase) not in root_keys
        child_tag = f" {_DIM}(child){_RESET}" if is_child else ""
        print(f"  {_CYAN}{mp.module_type} ({pl}){_RESET}{child_tag}")

        # Show each affected kernel with its % of module
        for kname, old_dur, new_dur in mp.kernel_deltas:
            k_pct = (old_dur / mp.original_wall_time_us * 100) if mp.original_wall_time_us > 0 else 0
            k_imp = ((old_dur - new_dur) / old_dur * 100) if old_dur > 0 else 0
            dname = kname if len(kname) <= 50 else kname[:47] + "..."
            print(
                f"    {_DIM}kernel: {dname}{_RESET}\n"
                f"    {_DIM}  {fmt_us(old_dur)} → {fmt_us(new_dur)} us "
                f"({k_imp:.0f}% faster, {k_pct:.1f}% of module wall time){_RESET}"
            )
        # Module-level summary
        print(
            f"    {_DIM}kernel sum: {fmt_us(mp.original_kernel_sum_us)} → "
            f"{fmt_us(mp.projected_kernel_sum_us)} us  "
            f"(overlap ratio: {mp.original_wall_time_us / mp.original_kernel_sum_us:.2f}x){_RESET}"
            if mp.original_kernel_sum_us > 0 else ""
        )
        print(
            f"    wall time: {fmt_us(mp.original_wall_time_us)} → "
            f"{fmt_us(mp.projected_wall_time_us)} us  "
            f"({fmt_pct(mp.delta_wall_time_pct)})"
        )
        print()

    # ── Step 2: Module → Phase step time ──
    # Use phase models which include ALL root modules (with and without detail sheets)
    # Order: prefill first, then decode (matches request lifecycle)
    _PHASE_ORDER = {"pre": 0, "dec": 1, "": 2}
    phases_to_show = sorted(original_pm.keys(), key=lambda p: _PHASE_ORDER.get(p, 99))
    if not phases_to_show:
        return

    if model_type == "LLM":
        phase_display = {"pre": "TTFT (prefill step)", "dec": "ITL (decode step)"}
    else:
        phase_display = {p: "Step time" for p in phases_to_show}

    print(f"  {_BOLD}Step 2: Module wall time → Phase step time{_RESET}\n")

    phase_results: Dict[str, Tuple[float, float, float]] = {}  # phase -> (orig, proj, delta_pct)

    for phase in phases_to_show:
        if phase not in original_pm or phase not in projected_pm:
            continue
        orig_model = original_pm[phase]
        proj_model = projected_pm[phase]
        metric_name = phase_display.get(phase, phase_label(phase))

        print(f"  {_CYAN}{metric_name}{_RESET}")
        if model_type == "LLM":
            print(f"    {_DIM}= sum of root {phase_label(phase)} module wall times{_RESET}")
        else:
            print(f"    {_DIM}= sum of root module wall times{_RESET}")

        # Show each root module with its % of phase
        for orig_rm, proj_rm in zip(orig_model.root_modules, proj_model.root_modules):
            pct = orig_rm.pct_of_phase
            detail_tag = "" if orig_rm.has_detail_sheet else f" {_DIM}(no detail sheet){_RESET}"
            if abs(proj_rm.change_pct) > 0.01:
                print(
                    f"    {_DIM}  {orig_rm.module_type} ({pct:.1f}%): "
                    f"{fmt_us(orig_rm.per_step_us)} → {fmt_us(proj_rm.projected_per_step_us)} us "
                    f"({fmt_pct(proj_rm.change_pct)}){detail_tag}{_RESET}"
                )
            else:
                print(
                    f"    {_DIM}  {orig_rm.module_type} ({pct:.1f}%): "
                    f"{fmt_us(orig_rm.per_step_us)} us (unchanged){detail_tag}{_RESET}"
                )

        orig_total = orig_model.total_step_us
        proj_total = proj_model.total_step_us
        delta_pct = ((proj_total - orig_total) / orig_total * 100) if orig_total > 0 else 0.0
        phase_results[phase] = (orig_total, proj_total, delta_pct)

        print(
            f"    {_DIM}  total: {fmt_us(orig_total)} → {fmt_us(proj_total)} us{_RESET}"
        )
        print(f"    {metric_name}: {fmt_pct(delta_pct)}")
        print()

    # ── Step 3: Phase → E2E metrics ──
    if model_type == "LLM":
        print(f"  {_BOLD}Step 3: Phase step time → E2E latency{_RESET}\n")
        if output_tokens and "pre" in phase_results and "dec" in phase_results:
            pre_orig, pre_proj, _ = phase_results["pre"]
            dec_orig, dec_proj, _ = phase_results["dec"]

            e2e_orig = pre_orig + (output_tokens - 1) * dec_orig
            e2e_proj = pre_proj + (output_tokens - 1) * dec_proj
            e2e_delta_pct = ((e2e_proj - e2e_orig) / e2e_orig * 100) if e2e_orig > 0 else 0.0

            print(
                f"    {_DIM}E2E = TTFT + (output_tokens - 1) x ITL{_RESET}\n"
                f"    {_DIM}    = {fmt_us(pre_proj)} + ({output_tokens} - 1) x {fmt_us(dec_proj)}{_RESET}\n"
                f"    {_DIM}    = {fmt_us(e2e_proj)} us{_RESET}"
            )
            print(
                f"    E2E ({output_tokens} output tokens): {fmt_pct(e2e_delta_pct)}  "
                f"({fmt_us(e2e_orig)} → {fmt_us(e2e_proj)} us)"
            )

            if e2e_proj > 0 and e2e_orig > 0:
                tp_change = (e2e_orig / e2e_proj - 1) * 100
                print(
                    f"    {_DIM}Throughput = old_e2e / new_e2e - 1 "
                    f"= {fmt_us(e2e_orig)} / {fmt_us(e2e_proj)} - 1{_RESET}"
                )
                print(f"    Throughput: {_GREEN}+{tp_change:.1f}%{_RESET}")
        else:
            if not output_tokens:
                print(f"    {_DIM}(Use --output-tokens N for E2E and throughput projections){_RESET}")
    else:
        # Single-phase: step 2 already shows the final result
        if len(phase_results) == 1:
            phase = list(phase_results.keys())[0]
            orig_total, proj_total, delta_pct = phase_results[phase]
            if orig_total > 0 and proj_total > 0:
                tp_change = (orig_total / proj_total - 1) * 100
                print(
                    f"  {_BOLD}Throughput:{_RESET} {_GREEN}+{tp_change:.1f}%{_RESET}  "
                    f"{_DIM}(= {fmt_us(orig_total)} / {fmt_us(proj_total)} - 1){_RESET}"
                )

    print()


def print_cumulative_summary(
    modules: Dict[Tuple[str, str], ModuleInfo],
    all_modifications: List[Modification],
    model_type: str,
    output_tokens: Optional[int],
    root_keys: Set[Tuple[str, str]],
    phase_models: Dict[str, PhaseModel],
    overview_nodes: Dict[str, OverviewNode],
    ancestors: Dict[str, Set[str]],
):
    """Print cumulative summary of all applied modifications."""
    if not all_modifications:
        print(f"\n  {_DIM}No modifications applied yet.{_RESET}\n")
        return

    print(f"\n{_BOLD}{'=' * 72}{_RESET}")
    print(f"{_BOLD}  Cumulative Projection Summary ({len(all_modifications)} modifications){_RESET}")
    print(f"{'=' * 72}\n")

    # List modifications
    print(f"  Applied changes:")
    for i, mod in enumerate(all_modifications, 1):
        dname = mod.kernel_name if len(mod.kernel_name) <= 50 else mod.kernel_name[:47] + "..."
        print(f"    {i}. {dname}: {mod.improvement_pct:.0f}% faster")
    print()

    projected_pm = project_phase_models(
        phase_models, modules, all_modifications, overview_nodes, ancestors,
    )
    print_projection_with_deduction(
        modules, all_modifications, model_type, output_tokens, root_keys,
        phase_models, projected_pm,
        verbose=True,
    )


# ── Interactive loop ─────────────────────────────────────────────────────────


def interactive_loop(
    filepath: str,
    modules: Dict[Tuple[str, str], ModuleInfo],
    top_kernels: List[TopKernel],
    output_tokens: Optional[int],
    root_keys: Set[Tuple[str, str]],
    phase_models: Dict[str, PhaseModel],
    overview_nodes: Dict[str, OverviewNode],
    ancestors: Dict[str, Set[str]],
):
    """Main interactive loop."""
    model_type = detect_model_type(modules)
    all_modifications: List[Modification] = []

    print_top_kernels(top_kernels)

    while True:
        prompt_parts = [
            f"  Select kernel [{_BOLD}1-{len(top_kernels)}{_RESET}]",
        ]
        if all_modifications:
            prompt_parts.append(f"'{_BOLD}c{_RESET}' cumulative")
            prompt_parts.append(f"'{_BOLD}r{_RESET}' reset")
        prompt_parts.append(f"'{_BOLD}s{_RESET}' show table")
        prompt_parts.append(f"'{_BOLD}q{_RESET}' quit")
        prompt = ", ".join(prompt_parts) + ": "

        try:
            choice = input(prompt).strip().lower()
        except (EOFError, KeyboardInterrupt):
            print()
            break

        if choice == "q":
            if all_modifications:
                print_cumulative_summary(
                    modules, all_modifications, model_type, output_tokens, root_keys,
                    phase_models, overview_nodes, ancestors,
                )
            break
        elif choice == "s":
            print_top_kernels(top_kernels)
            continue
        elif choice == "c":
            print_cumulative_summary(
                modules, all_modifications, model_type, output_tokens, root_keys,
                phase_models, overview_nodes, ancestors,
            )
            continue
        elif choice == "r":
            all_modifications.clear()
            print(f"  {_GREEN}All modifications reset.{_RESET}\n")
            continue

        # Parse kernel selection
        try:
            idx = int(choice)
            if idx < 1 or idx > len(top_kernels):
                print(f"  {_RED}Invalid selection. Choose 1-{len(top_kernels)}.{_RESET}")
                continue
        except ValueError:
            print(f"  {_RED}Invalid input.{_RESET}")
            continue

        tk = top_kernels[idx - 1]
        print_kernel_detail(tk, modules)

        if not tk.detail_refs:
            print(f"  {_YELLOW}Cannot project: kernel not in any detail sheet.{_RESET}\n")
            continue

        # Get improvement percentage
        try:
            imp_input = input(f"  Improvement % (e.g. 20 = 20% faster, 0 to skip): ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break

        if not imp_input or imp_input == "0":
            continue

        try:
            improvement = float(imp_input)
        except ValueError:
            print(f"  {_RED}Invalid number.{_RESET}")
            continue

        if improvement < 0 or improvement > 100:
            print(f"  {_YELLOW}Warning: unusual improvement value ({improvement}%).{_RESET}")

        mod = Modification(kernel_name=tk.name, improvement_pct=improvement)

        # Compute projection with phase models
        projected_pm = project_phase_models(
            phase_models, modules, [mod], overview_nodes, ancestors,
        )
        print_projection_with_deduction(
            modules, [mod], model_type, output_tokens, root_keys,
            phase_models, projected_pm,
        )

        # Add to cumulative
        all_modifications.append(mod)
        print(
            f"  {_GREEN}Change applied. "
            f"({len(all_modifications)} total modifications, 'c' for cumulative summary){_RESET}\n"
        )


# ── CLI ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Interactive kernel performance projection tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python kernel_projection.py analysis.xlsx
  python kernel_projection.py analysis.xlsx --output-tokens 128
  python kernel_projection.py analysis.xlsx --top 20
        """,
    )
    parser.add_argument("excel_path", help="Path to analysis.xlsx from trace_module_analyzer.py")
    parser.add_argument(
        "--output-tokens",
        type=int,
        default=None,
        help="Number of output tokens for E2E projection (LLM models)",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=10,
        help="Number of top kernels to display (default: 10)",
    )
    args = parser.parse_args()

    path = args.excel_path
    if not Path(path).exists():
        print(f"Error: {path} not found", file=sys.stderr)
        sys.exit(1)

    # Parse Excel
    print(f"Loading {path}...")
    modules = parse_excel(path)
    gpu_kernels = parse_gpu_kernels_sheet(path)

    if not modules:
        print("Error: no detail sheets found in Excel file.", file=sys.stderr)
        sys.exit(1)

    if not gpu_kernels:
        print("Warning: GPU Kernels sheet not found, using detail sheets for ranking.")
        # Fall back: build kernel ranking from detail sheets
        kernel_agg: Dict[str, Tuple[float, int, str, Set[str]]] = {}
        for (mtype, phase), minfo in modules.items():
            for k in minfo.kernels:
                if k.kernel_name not in kernel_agg:
                    kernel_agg[k.kernel_name] = (0.0, 0, k.category, set())
                old = kernel_agg[k.kernel_name]
                kernel_agg[k.kernel_name] = (
                    old[0] + k.duration_us,
                    old[1] + 1,
                    old[2],
                    old[3] | {mtype},
                )
        gpu_kernels = sorted(
            [
                GPUKernelRow(
                    name=name,
                    category=cat,
                    total_duration_us=dur,
                    count=cnt,
                    avg_us=dur / cnt if cnt else 0,
                    pct_of_total=0,
                    module_types_str=", ".join(mtypes),
                )
                for name, (dur, cnt, cat, mtypes) in kernel_agg.items()
            ],
            key=lambda x: x.total_duration_us,
            reverse=True,
        )
        total = sum(g.total_duration_us for g in gpu_kernels)
        for g in gpu_kernels:
            g.pct_of_total = (g.total_duration_us / total * 100) if total > 0 else 0

    # Parse module hierarchy for parent-child detection
    overview_nodes, ancestors = parse_overview_tree(path)
    root_keys = get_root_detail_modules(modules, ancestors)

    # Build phase models (uses Overview for complete root module list)
    phase_models = build_phase_models(path, modules, overview_nodes, ancestors)

    # Build top-N with cross-references
    top_kernels = build_top_kernels(gpu_kernels, modules, top_n=args.top)

    # Print header and enter interactive loop
    print_header(path, modules, root_keys)

    # For LLM models, prompt for output_tokens if not provided
    output_tokens = args.output_tokens
    model_type = detect_model_type(modules)
    if model_type == "LLM" and output_tokens is None:
        try:
            ans = input(
                f"  Output tokens for E2E projection (Enter to skip): "
            ).strip()
            if ans:
                output_tokens = int(ans)
        except (ValueError, EOFError, KeyboardInterrupt):
            pass

    interactive_loop(
        path, modules, top_kernels, output_tokens, root_keys,
        phase_models, overview_nodes, ancestors,
    )


if __name__ == "__main__":
    main()
